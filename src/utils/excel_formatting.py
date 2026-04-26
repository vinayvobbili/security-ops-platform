"""
Excel Formatting Utilities

Provides consistent professional formatting for Excel reports across the application.
"""

import logging
import os

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle

logger = logging.getLogger(__name__)


def apply_professional_formatting(file_path, column_widths=None, wrap_columns=None, date_columns=None):
    """
    Apply professional formatting to Excel files including:
    - Bold white headers on blue background
    - Zebra striping (alternating row colors)
    - Text wrapping for specified columns
    - Date formatting for specified columns
    - Cell borders throughout
    - Frozen header row
    - Auto-filter
    
    Args:
        file_path: Path to Excel file
        column_widths: Dict mapping column names (lowercase) to widths
        wrap_columns: Set of column names (lowercase) that should wrap text
        date_columns: Set of column names (lowercase) that contain dates
    """
    try:
        # Default column widths
        default_column_widths = {
            'hostname': 25,
            'host_id': 20,
            'current_tags': 80,
            'invalid_tags': 60,
            'last_seen': 20,
            'status': 15,
            'cs_host_category': 20,
            'snow_environment': 15,
            'snow_lifecyclestatus': 20,
            'comment': 50,
            'environment': 15,
            'platform': 15,
            'tags': 60,
            'device_id': 20,
            'first_seen': 20,
            'local_ip': 15,
            'external_ip': 15,
            'os_version': 30,
            'kernel_version': 30,
            'system_manufacturer': 20,
            'system_product_name': 25,
        }

        # Use provided column widths or defaults
        if column_widths:
            default_column_widths.update(column_widths)
        column_widths = default_column_widths

        # Default wrap columns
        default_wrap_columns = {'current_tags', 'invalid_tags', 'comment', 'tags'}
        if wrap_columns:
            default_wrap_columns.update(wrap_columns)
        wrap_columns = default_wrap_columns

        # Default date columns
        default_date_columns = {'last_seen', 'first_seen'}
        if date_columns:
            default_date_columns.update(date_columns)
        date_columns = default_date_columns

        workbook = load_workbook(file_path)
        worksheet = workbook.active

        # Get header row to map column names to letters
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')

        # Zebra stripe colors
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Create date style with borders
        date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY HH:MM')
        date_style.border = thin_border

        # Format headers
        for col_idx, header in enumerate(header_row, 1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            cell = worksheet.cell(row=1, column=col_idx)

            # Set column width
            if header and header.lower() in column_widths:
                worksheet.column_dimensions[col_letter].width = column_widths[header.lower()]
            else:
                worksheet.column_dimensions[col_letter].width = 15

            # Format header cell
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Format data rows
        for row_idx in range(2, worksheet.max_row + 1):
            # Zebra striping - every other row
            is_alternate_row = (row_idx % 2 == 0)

            for col_idx, header in enumerate(header_row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Add borders to all cells
                cell.border = thin_border

                # Zebra striping
                if is_alternate_row:
                    cell.fill = light_fill

                # Text wrapping for long content columns
                if header and header.lower() in wrap_columns:
                    cell.alignment = wrap_alignment

                # Date formatting
                if header and header.lower() in date_columns and cell.value:
                    cell.style = date_style

        # Freeze the header row
        worksheet.freeze_panes = 'A2'

        # Add auto filter to the data range
        if worksheet.max_row > 1:  # Only add filter if there's data beyond headers
            worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"

        # Add watermark: "By <author>" in bottom-right cell + page footer
        _watermark_author = os.environ.get("WATERMARK_AUTHOR", "")
        if _watermark_author:
            wm_text = f"By {_watermark_author}"
            wm_row = worksheet.max_row + 2
            wm_col = worksheet.max_column
            wm_cell = worksheet.cell(row=wm_row, column=wm_col)
            wm_cell.value = wm_text
            wm_cell.font = Font(size=8, italic=True, color="9E9E9E")
            wm_cell.alignment = Alignment(horizontal='right')
            worksheet.oddFooter.right.text = wm_text
            worksheet.oddFooter.right.size = 8
            worksheet.oddFooter.right.font = "Calibri,Italic"
            worksheet.oddFooter.right.color = "9E9E9E"

        workbook.save(file_path)
        logger.info(f"Applied professional formatting to {file_path}")

    except Exception as e:
        logger.warning(f"Could not format Excel file {file_path}: {e}")


def add_tanium_hyperlinks(file_path, portal_url=None, tanium_id_column='Tanium ID',
                          action_id_column='Action ID', scheduled_action_id_column='Scheduled Action ID',
                          source_column='Source', portal_urls_by_source=None):
    """Add clickable hyperlinks to Tanium ID and/or Action ID columns in an Excel file.

    Supports either a single portal_url for all rows, or per-source URLs via
    portal_urls_by_source (e.g. {'Cloud': 'https://...', 'On-Prem': 'https://...'}).

    Args:
        file_path: Path to the Excel file
        portal_url: Single Tanium portal base URL for all rows (used if portal_urls_by_source is None)
        tanium_id_column: Header name for the Tanium endpoint ID column (None to skip)
        action_id_column: Header name for the Action ID column (None to skip)
        source_column: Header name for the Source column (used with portal_urls_by_source)
        portal_urls_by_source: Dict mapping source name to portal URL (e.g. {'Cloud': 'https://...'})
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Build header -> column index map
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col_map = {str(h).strip(): idx + 1 for idx, h in enumerate(header_row) if h}

        tanium_col = col_map.get(tanium_id_column) if tanium_id_column else None
        action_col = col_map.get(action_id_column) if action_id_column else None
        sched_action_col = col_map.get(scheduled_action_id_column) if scheduled_action_id_column else None
        source_col = col_map.get(source_column) if portal_urls_by_source and source_column else None

        if not tanium_col and not action_col and not sched_action_col:
            logger.debug(f"No Tanium ID or Action ID columns found in {file_path}, skipping hyperlinks")
            return

        for row_idx in range(2, ws.max_row + 1):
            # Determine portal URL for this row
            if portal_urls_by_source and source_col:
                source_val = str(ws.cell(row=row_idx, column=source_col).value or '').strip()
                portal = portal_urls_by_source.get(source_val, portal_url or '')
            else:
                portal = portal_url or ''
            if not portal:
                continue
            portal = portal.rstrip('/')

            if tanium_col:
                cell = ws.cell(row=row_idx, column=tanium_col)
                val = str(cell.value or '').strip().rstrip('0').rstrip('.') if '.' in str(cell.value or '') else str(cell.value or '').strip()
                if val:
                    cell.hyperlink = (
                        f"{portal}/ui/reporting/single-endpoint-view"
                        f"?eid={val}&tab=sev-endpoint-overview"
                    )
                    cell.style = 'Hyperlink'
            if action_col:
                cell = ws.cell(row=row_idx, column=action_col)
                val = str(cell.value or '').strip().rstrip('0').rstrip('.') if '.' in str(cell.value or '') else str(cell.value or '').strip()
                if val and val != 'N/A':
                    cell.hyperlink = f"{portal}/ui/console/actions/scheduled-actions/{val}"
                    cell.style = 'Hyperlink'
            if sched_action_col:
                cell = ws.cell(row=row_idx, column=sched_action_col)
                val = str(cell.value or '').strip().rstrip('0').rstrip('.') if '.' in str(cell.value or '') else str(cell.value or '').strip()
                if val and val != 'N/A':
                    cell.hyperlink = f"{portal}/ui/console/actions/scheduled-actions/{val}"
                    cell.style = 'Hyperlink'

        wb.save(file_path)
        logger.info(f"Added Tanium hyperlinks to {file_path}")

    except Exception as e:
        logger.warning(f"Could not add Tanium hyperlinks to {file_path}: {e}")
