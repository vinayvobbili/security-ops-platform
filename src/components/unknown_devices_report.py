"""
Unknown Devices Report

Generates a report categorizing devices that cannot be properly classified:
- Unknown due to missing data in Snow (entry exists but category is missing)
- Unknown due to no Snow entry (device not found in ServiceNow)
- Unknown due to an error (API errors during lookup)

This report helps identify:
- Which teams need to correct their ServiceNow entries
- Issues on the security side (API errors, connectivity)

Usage:
    python -m src.components.unknown_devices_report
    python -m src.components.unknown_devices_report --output /path/to/output.xlsx
"""

import argparse
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from services.crowdstrike import CrowdStrikeClient, DATA_DIR, process_unique_hosts
from services.service_now import enrich_host_report

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Color scheme
COLORS = {
    'header_blue': '4472C4',
    'missing_data': 'FFC000',      # Orange - teams need to fix
    'no_entry': 'ED7D31',          # Dark orange - teams need to add
    'error': 'A5A5A5',             # Gray - our side issue
    'known': '70AD47',             # Green - properly categorized
    'light_gray': 'F2F2F2',
}


def categorize_unknown_devices(enriched_df: pd.DataFrame) -> dict:
    """
    Categorize devices based on their ServiceNow enrichment status.

    Returns:
        Dict with categorized DataFrames and the known (properly categorized) devices
    """
    missing_snow_data = []
    no_snow_entry = []
    error = []
    known = []

    for idx, row in enriched_df.iterrows():
        snow_status = str(row.get('SNOW_status', '')).strip()
        snow_category = str(row.get('SNOW_category', '')).strip()
        snow_error = str(row.get('SNOW_error', '')).strip()

        # Check for API error first (our side issue)
        # Note: NaN becomes "nan" when converted to string, so check for that
        if snow_status == 'ServiceNow API Error' or (snow_error and snow_error.lower() not in ('', 'nan', 'none')):
            error.append(idx)
        # Check for not found in Snow (teams need to add entry)
        elif snow_status == 'Not Found':
            no_snow_entry.append(idx)
        # Check for missing category data (teams need to fix entry)
        elif not snow_category or snow_category.lower() in ('', 'nan', 'none'):
            missing_snow_data.append(idx)
        else:
            # Properly categorized
            known.append(idx)

    return {
        'missing_snow_data': enriched_df.loc[missing_snow_data].copy() if missing_snow_data else pd.DataFrame(),
        'no_snow_entry': enriched_df.loc[no_snow_entry].copy() if no_snow_entry else pd.DataFrame(),
        'error': enriched_df.loc[error].copy() if error else pd.DataFrame(),
        'known': enriched_df.loc[known].copy() if known else pd.DataFrame(),
    }


def get_breakdown_by_field(df: pd.DataFrame, field: str) -> pd.DataFrame:
    """Get count breakdown by a specific field."""
    if df.empty or field not in df.columns:
        return pd.DataFrame(columns=[field, 'Count'])

    counts = df[field].fillna('(Not Specified)').value_counts().reset_index()
    counts.columns = [field, 'Count']
    return counts.sort_values('Count', ascending=False)


def apply_formatting_to_sheet(ws, has_data=True, max_format_rows=5000):
    """Apply professional formatting to a worksheet.

    Args:
        ws: Worksheet to format
        has_data: Whether the sheet has data rows
        max_format_rows: Maximum rows to apply detailed formatting (for performance)
    """
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=COLORS['header_blue'], end_color=COLORS['header_blue'], fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    light_fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")

    # Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format data rows with zebra striping (limit to max_format_rows for performance)
    # For large datasets (>5000 rows), skip cell-by-cell formatting to avoid hanging
    rows_to_format = min(ws.max_row + 1, max_format_rows + 1)
    if ws.max_row > max_format_rows:
        logger.info(f"  Sheet has {ws.max_row:,} rows - applying formatting to first {max_format_rows:,} rows only")

    for row_idx in range(2, rows_to_format):
        for cell in ws[row_idx]:
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill

    # Auto-width columns based on header and sample of data (not all rows)
    sample_rows = min(ws.max_row, 100)  # Sample first 100 rows for width calculation
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        for row_idx in range(1, sample_rows + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    # Freeze header row and add auto-filter
    ws.freeze_panes = 'A2'
    if has_data and ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"


def create_executive_summary_sheet(wb, categories: dict, total_devices: int):
    """Create an executive summary sheet with charts."""
    ws = wb.create_sheet("Executive Summary", 0)

    # Title
    ws['A1'] = "Unknown Devices Report"
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['header_blue'])
    ws.merge_cells('A1:E1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')

    # Key metrics section
    ws['A4'] = "KEY METRICS"
    ws['A4'].font = Font(bold=True, size=12)

    # Summary data for chart
    ws['A6'] = "Category"
    ws['B6'] = "Count"
    ws['C6'] = "Percentage"
    ws['D6'] = "Action Required"

    counts = {
        'missing_snow_data': len(categories['missing_snow_data']),
        'no_snow_entry': len(categories['no_snow_entry']),
        'error': len(categories['error']),
        'known': len(categories['known']),
    }
    total_unknown = counts['missing_snow_data'] + counts['no_snow_entry'] + counts['error']

    rows_data = [
        ("Missing Data in Snow", counts['missing_snow_data'], "Teams: Update ServiceNow entries"),
        ("No Snow Entry", counts['no_snow_entry'], "Teams: Add device to ServiceNow"),
        ("API/Lookup Errors", counts['error'], "Security: Investigate connectivity"),
        ("Properly Categorized", counts['known'], "No action needed"),
    ]

    for i, (label, count, action) in enumerate(rows_data, start=7):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = count
        ws[f'C{i}'] = f"{(count/total_devices*100):.1f}%" if total_devices > 0 else "0%"
        ws[f'D{i}'] = action

    # Total row
    ws['A11'] = "TOTAL DEVICES"
    ws['B11'] = total_devices
    ws['A11'].font = Font(bold=True)
    ws['B11'].font = Font(bold=True)

    ws['A12'] = "TOTAL UNKNOWN"
    ws['B12'] = total_unknown
    ws['C12'] = f"{(total_unknown/total_devices*100):.1f}%" if total_devices > 0 else "0%"
    ws['A12'].font = Font(bold=True, color="C00000")
    ws['B12'].font = Font(bold=True, color="C00000")
    ws['C12'].font = Font(bold=True, color="C00000")

    # Apply formatting to summary table
    header_fill = PatternFill(start_color=COLORS['header_blue'], end_color=COLORS['header_blue'], fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[6]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

    # Color-code the category rows
    category_colors = [
        COLORS['missing_data'],
        COLORS['no_entry'],
        COLORS['error'],
        COLORS['known'],
    ]

    for i, color in enumerate(category_colors, start=7):
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{i}']
            cell.border = thin_border
        # Color indicator in first column
        ws[f'A{i}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Create pie chart for unknown categories only
    if total_unknown > 0:
        # Data for pie chart (only unknown categories)
        ws['G6'] = "Unknown Category"
        ws['H6'] = "Count"
        ws['G7'] = "Missing Data in Snow"
        ws['H7'] = counts['missing_snow_data']
        ws['G8'] = "No Snow Entry"
        ws['H8'] = counts['no_snow_entry']
        ws['G9'] = "API/Lookup Errors"
        ws['H9'] = counts['error']

        pie = PieChart()
        pie.title = "Unknown Devices Breakdown"
        labels = Reference(ws, min_col=7, min_row=7, max_row=9)
        data = Reference(ws, min_col=8, min_row=6, max_row=9)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 12
        pie.height = 8

        # Add data labels
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        pie.dataLabels.showCatName = False

        # Set colors for pie slices
        colors_list = [COLORS['missing_data'], COLORS['no_entry'], COLORS['error']]
        for i, color in enumerate(colors_list):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            pie.series[0].data_points.append(pt)

        ws.add_chart(pie, "G12")

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 12

    return ws


def create_breakdown_sheet(wb, df: pd.DataFrame, sheet_name: str, breakdown_field: str):
    """Create a sheet with breakdown by specified field and a bar chart."""
    if df.empty:
        return

    breakdown = get_breakdown_by_field(df, breakdown_field)
    if breakdown.empty:
        return

    ws = wb.create_sheet(sheet_name)

    # Write breakdown data
    ws['A1'] = breakdown_field.replace('SNOW_', '').replace('_', ' ').title()
    ws['B1'] = 'Count'

    for i, (_, row) in enumerate(breakdown.iterrows(), start=2):
        ws[f'A{i}'] = row[breakdown_field]
        ws[f'B{i}'] = row['Count']

    apply_formatting_to_sheet(ws)

    # Create bar chart if we have data
    if len(breakdown) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"Devices by {breakdown_field.replace('SNOW_', '').replace('_', ' ').title()}"
        chart.y_axis.title = "Count"

        data = Reference(ws, min_col=2, min_row=1, max_row=min(len(breakdown) + 1, 16))
        cats = Reference(ws, min_col=1, min_row=2, max_row=min(len(breakdown) + 1, 16))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 15
        chart.height = 10

        # Position chart to the right of data
        ws.add_chart(chart, "D2")


def generate_report(output_path: Path = None) -> dict:
    """
    Generate the unknown devices report with charts and professional formatting.

    Args:
        output_path: Optional path for output Excel file

    Returns:
        Dict with summary counts and output file path
    """
    today_date = datetime.now().strftime('%m-%d-%Y')
    output_dir = DATA_DIR / today_date

    # Step 1: Fetch all hosts from CrowdStrike
    logger.info("Step 1: Fetching all hosts from CrowdStrike...")
    cs_client = CrowdStrikeClient()
    cs_client.fetch_all_hosts_and_write_to_xlsx()

    # Step 2: Process to get unique hosts
    logger.info("Step 2: Processing unique hosts...")
    hosts_file = output_dir / "all_cs_hosts.xlsx"
    df = pd.read_excel(hosts_file, engine="openpyxl")
    unique_hosts = process_unique_hosts(df)

    unique_hosts_file = output_dir / "unique_cs_hosts.xlsx"
    unique_hosts_file.parent.mkdir(parents=True, exist_ok=True)
    unique_hosts.to_excel(unique_hosts_file, index=False, engine="openpyxl")

    # Step 3: Enrich with ServiceNow data
    logger.info("Step 3: Enriching with ServiceNow data...")
    enriched_file = enrich_host_report(str(unique_hosts_file))

    # Step 4: Read enriched data and categorize
    logger.info("Step 4: Categorizing unknown devices...")
    enriched_df = pd.read_excel(enriched_file, engine="openpyxl")
    categories = categorize_unknown_devices(enriched_df)

    # Step 5: Generate output report
    logger.info("Step 5: Generating report with charts...")
    if output_path is None:
        output_path = output_dir / f"unknown_devices_report_{today_date}.xlsx"

    # Select relevant columns for detail sheets
    report_columns = [
        'hostname', 'host_id', 'last_seen', 'status', 'cs_host_category',
        'SNOW_category', 'SNOW_status', 'SNOW_error', 'SNOW_ciClass',
        'SNOW_environment', 'SNOW_country', 'SNOW_lifecycleStatus'
    ]

    def filter_columns(df_in):
        cols = [c for c in report_columns if c in df_in.columns]
        return df_in[cols] if cols else df_in

    # Create workbook
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Placeholder sheet (will be replaced by Executive Summary)
        pd.DataFrame({'Placeholder': [1]}).to_excel(writer, sheet_name='Temp', index=False)

        # Write detail sheets
        if not categories['missing_snow_data'].empty:
            filter_columns(categories['missing_snow_data']).to_excel(
                writer, sheet_name='Missing Snow Data', index=False
            )
        if not categories['no_snow_entry'].empty:
            filter_columns(categories['no_snow_entry']).to_excel(
                writer, sheet_name='No Snow Entry', index=False
            )
        if not categories['error'].empty:
            filter_columns(categories['error']).to_excel(
                writer, sheet_name='API Errors', index=False
            )

    # Load workbook to add charts and formatting
    wb = load_workbook(output_path)

    # Remove placeholder sheet
    if 'Temp' in wb.sheetnames:
        del wb['Temp']

    # Create executive summary with charts
    create_executive_summary_sheet(wb, categories, len(enriched_df))

    # Create breakdown sheets for actionable insights
    # For "Missing Snow Data" - breakdown by country to identify which teams
    if not categories['missing_snow_data'].empty:
        create_breakdown_sheet(wb, categories['missing_snow_data'],
                             'Missing Data by Country', 'SNOW_country')

    # For "No Snow Entry" - breakdown by CS host category
    if not categories['no_snow_entry'].empty:
        create_breakdown_sheet(wb, categories['no_snow_entry'],
                             'No Entry by CS Category', 'cs_host_category')

    # Apply formatting to detail sheets
    for sheet_name in ['Missing Snow Data', 'No Snow Entry', 'API Errors']:
        if sheet_name in wb.sheetnames:
            apply_formatting_to_sheet(wb[sheet_name])

    wb.save(output_path)

    result = {
        'output_file': str(output_path),
        'total_devices': len(enriched_df),
        'missing_snow_data': len(categories['missing_snow_data']),
        'no_snow_entry': len(categories['no_snow_entry']),
        'error': len(categories['error']),
        'known': len(categories['known']),
        'total_unknown': (
            len(categories['missing_snow_data']) +
            len(categories['no_snow_entry']) +
            len(categories['error'])
        )
    }

    logger.info(f"Report generated: {output_path}")
    logger.info("Summary:")
    logger.info(f"  - Total devices processed: {result['total_devices']}")
    logger.info(f"  - Properly categorized: {result['known']}")
    logger.info(f"  - Unknown - missing Snow data: {result['missing_snow_data']} (Teams need to update)")
    logger.info(f"  - Unknown - no Snow entry: {result['no_snow_entry']} (Teams need to add)")
    logger.info(f"  - Unknown - API errors: {result['error']} (Our side)")
    logger.info(f"  - Total unknown: {result['total_unknown']}")

    return result


def main():
    parser = argparse.ArgumentParser(
        description="Generate unknown devices report by category"
    )
    parser.add_argument(
        '--output', '-o',
        type=str,
        help='Output Excel file path (default: data/transient/epp_device_tagging/<date>/unknown_devices_report_<date>.xlsx)'
    )
    args = parser.parse_args()

    output_path = Path(args.output) if args.output else None
    result = generate_report(output_path)

    print("\n" + "=" * 70)
    print("UNKNOWN DEVICES REPORT - EXECUTIVE SUMMARY")
    print("=" * 70)
    print(f"Total devices processed:         {result['total_devices']:>10,}")
    print(f"Properly categorized:            {result['known']:>10,}")
    print("-" * 70)
    print("UNKNOWN DEVICES (Action Required):")
    print("-" * 70)
    print(f"  Missing data in Snow:          {result['missing_snow_data']:>10,}  <- Teams: Update entries")
    print(f"  No Snow entry:                 {result['no_snow_entry']:>10,}  <- Teams: Add to ServiceNow")
    print(f"  API/Lookup errors:             {result['error']:>10,}  <- Security: Check connectivity")
    print("-" * 70)
    print(f"TOTAL UNKNOWN:                   {result['total_unknown']:>10,}")
    print("=" * 70)
    print(f"\nReport saved to: {result['output_file']}")
    print("\nSheets included:")
    print("  - Executive Summary (with charts)")
    print("  - Missing Snow Data (detail)")
    print("  - No Snow Entry (detail)")
    print("  - API Errors (detail)")
    print("  - Breakdown by Country/Category")


if __name__ == "__main__":
    main()
