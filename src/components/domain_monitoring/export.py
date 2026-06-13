"""Excel export for the Domain Monitoring dashboard.

Flattens a daily ``results.json`` into a professionally formatted, multi-sheet
workbook for reporting and tracking — a Summary sheet plus one sheet per finding
class (lookalikes, RF watchlist, brand-keyword impersonation, outstanding SSL
threats).
"""

import logging
from datetime import datetime
from typing import Any, Dict, List

logger = logging.getLogger(__name__)


def _lookalike_rows(results: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = []
    for parent, data in (results.get("domains") or {}).items():
        look = data.get("lookalikes") or {}
        seen = set()
        for bucket, status in (("new_domains", "New"), ("became_active", "Became Active"),
                               ("reregistered", "Re-registered")):
            for d in look.get(bucket, []):
                dom = d.get("domain")
                if not dom or dom in seen:
                    continue
                seen.add(dom)
                rows.append({
                    "Domain": dom,
                    "Parent Brand": parent,
                    "Status": status,
                    "Risk": d.get("risk_category") or d.get("risk") or "",
                    "Registrar": d.get("registrar") or "",
                    "IP": ", ".join(d.get("dns_a", []) or []) if isinstance(d.get("dns_a"), list) else (d.get("ip") or ""),
                    "RF Risk Score": d.get("rf_risk_score") if d.get("rf_risk_score") is not None else "",
                    "VT Malicious": (d.get("vt_stats") or {}).get("malicious", ""),
                    "Defensive": "Yes" if d.get("is_defensive") else "No",
                })
    return rows


def _rf_watchlist_rows(results: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = []
    wl = (results.get("rf_watchlist") or {}).get("domains", [])
    for d in wl:
        rows.append({
            "Domain": d.get("domain", ""),
            "RF Risk Score": d.get("rf_risk_score") if d.get("rf_risk_score") is not None else "",
            "RF Risk Level": d.get("rf_risk_level") or "",
            "RF Rules": ", ".join(d.get("rf_rules", []) or []),
            "Reputation": "MALICIOUS" if d.get("reputation_malicious") else "",
        })
    return rows


def _brand_keyword_rows(results: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = []
    bk = (results.get("brand_keyword_impersonation") or {}).get("new_domains", [])
    for d in bk:
        rows.append({
            "Domain": d.get("domain", ""),
            "Brand Keyword": d.get("brand_keyword", ""),
            "Issuer": d.get("issuer") or "",
            "First Seen": d.get("not_before") or d.get("discovered_at") or "",
        })
    return rows


def _summary_rows(results: Dict[str, Any], report_date: str) -> List[tuple]:
    return [
        ("Domain Monitoring Report", ""),
        ("", ""),
        ("Report Date", report_date),
        ("Scan Time", results.get("scan_time", "")),
        ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("New Lookalike Domains", results.get("total_new_lookalikes", 0)),
        ("Became Active", results.get("total_became_active", 0)),
        ("RF Watchlist Domains", results.get("total_rf_watchlist", 0)),
        ("RF Watchlist High Risk", results.get("total_rf_watchlist_high_risk", 0)),
        ("Brand-Keyword Impersonations", results.get("total_brand_keyword_impersonation", 0)),
        ("CT / SSL Findings", results.get("total_ct_findings", 0)),
        ("WHOIS Changes", results.get("total_whois_changes", 0)),
        ("VT High Risk", results.get("total_vt_high_risk", 0)),
        ("abuse.ch Malicious", results.get("total_abusech_malicious", 0)),
        ("AbuseIPDB Malicious", results.get("total_abuseipdb_malicious", 0)),
    ]


def build_export_workbook(results: Dict[str, Any], out_path: str, report_date: str) -> str:
    """Build the multi-sheet xlsx export at ``out_path`` and return the path."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Font, Side
    from src.utils.excel_formatting import apply_professional_formatting

    sheets = {
        "Lookalikes": _lookalike_rows(results),
        "RF Watchlist": _rf_watchlist_rows(results),
        "Brand-Keyword Impersonation": _brand_keyword_rows(results),
        "Outstanding SSL Threats": _ssl_threat_rows(),
    }

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Placeholder sheet replaced below by the styled Summary; ensures the file
        # always has at least one finding sheet even when every list is empty.
        wrote_any = False
        for name, rows in sheets.items():
            if rows:
                pd.DataFrame(rows).to_excel(writer, sheet_name=name[:31], index=False)
                wrote_any = True
        if not wrote_any:
            pd.DataFrame([{"Status": "No findings in this report"}]).to_excel(
                writer, sheet_name="Findings", index=False)

    apply_professional_formatting(out_path)

    # Prepend a styled Summary sheet.
    wb = load_workbook(out_path)
    summary = wb.create_sheet("Summary", 0)
    for row in _summary_rows(results, report_date):
        summary.append(row)
    summary["A1"].font = Font(bold=True, size=14, color="0046AD")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for r in range(3, summary.max_row + 1):
        label = summary.cell(row=r, column=1)
        if label.value:
            label.font = Font(bold=True)
            for c in (1, 2):
                summary.cell(row=r, column=c).border = thin
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 28
    wb.save(out_path)

    logger.info(f"Built domain monitoring export at {out_path}")
    return out_path


def build_monthly_report_workbook(rollup: Dict[str, Any], out_path: str, month: str) -> str:
    """Build the monthly Domain Monitoring & Brand Protection report xlsx.

    Mirrors the manual report's sheets: Summary, All Domains (with triage
    columns), By Brand, Takedown vs Monitoring, Irrelevant Domains, Weekly Trend.
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Font, Side
    from src.utils.excel_formatting import apply_professional_formatting

    def _finding_rows(items):
        rows = []
        for f in items:
            rows.append({
                "Domain": f.get("domain", ""),
                "First Seen": (f.get("first_seen") or "")[:10],
                "Brand": f.get("brand") or "",
                "Source": f.get("source") or "",
                "Status": f.get("status") or "",
                "Assignee": f.get("assignee") or "",
                "PhishFort Incident": f.get("phishfort_incident_id") or "",
                "XSOAR": f.get("xsoar_id") or "",
                "Risk": f.get("risk_score") if f.get("risk_score") is not None else "",
                "Notes": f.get("notes") or "",
            })
        return rows or [{"Domain": "No findings this month"}]

    sheets = {
        "All Domains": _finding_rows([f for f in rollup.get("findings", []) if f.get("status") != "irrelevant"]),
        "By Brand": [{"Impersonating Brand": b, "Count": c} for b, c in rollup.get("by_brand", {}).items()]
                     or [{"Impersonating Brand": "(none)", "Count": 0}],
        "Takedown vs Monitoring": [
            {"Category": "Monitoring", "Count": rollup.get("monitoring", 0)},
            {"Category": "Takedown", "Count": rollup.get("takedowns", 0)},
        ],
        "Weekly Trend": [{"Week": w, "Count": c} for w, c in rollup.get("weekly_trend", {}).items()]
                        or [{"Week": "(none)", "Count": 0}],
        "Irrelevant Domains": _finding_rows(rollup.get("irrelevant_findings", [])),
    }

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, rows in sheets.items():
            pd.DataFrame(rows).to_excel(writer, sheet_name=name[:31], index=False)

    apply_professional_formatting(out_path)

    wb = load_workbook(out_path)
    summary = wb.create_sheet("Summary", 0)
    summary_rows = [
        ("Domain Monitoring & Brand Protection Report", ""),
        ("", ""),
        ("Reporting Month", month),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Total Domains Reviewed", rollup.get("total_findings", 0)),
        ("Relevant (Monitoring + Takedown)", rollup.get("relevant", 0)),
        ("Takedowns Raised", rollup.get("takedowns", 0)),
        ("Under Monitoring", rollup.get("monitoring", 0)),
        ("Irrelevant / Triaged Out", rollup.get("irrelevant", 0)),
        ("Brands Impersonated", len(rollup.get("by_brand", {}))),
    ]
    for row in summary_rows:
        summary.append(row)
    summary["A1"].font = Font(bold=True, size=14, color="0046AD")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for r in range(3, summary.max_row + 1):
        label = summary.cell(row=r, column=1)
        if label.value:
            label.font = Font(bold=True)
            for c in (1, 2):
                summary.cell(row=r, column=c).border = thin
    summary.column_dimensions["A"].width = 40
    summary.column_dimensions["B"].width = 26
    wb.save(out_path)

    logger.info(f"Built monthly report for {month} at {out_path}")
    return out_path


def _ssl_threat_rows() -> List[Dict[str, Any]]:
    """Outstanding (unacknowledged) SSL-cert threats from the CT monitor store."""
    try:
        from services.cert_transparency import get_outstanding_threats
        threats = get_outstanding_threats()
    except Exception as e:
        logger.warning(f"Could not load outstanding SSL threats for export: {e}")
        return []
    rows = []
    for t in threats:
        rows.append({
            "Domain": t.get("domain", ""),
            "Brand": t.get("brand", ""),
            "Issuer": t.get("issuer", ""),
            "Discovered": t.get("discovered_at", ""),
            "Status": t.get("status", ""),
        })
    return rows
