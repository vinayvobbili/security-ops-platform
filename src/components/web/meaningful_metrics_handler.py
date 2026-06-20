"""Meaningful Metrics Handler for Web Dashboard."""

import json
import logging
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List

import pandas as pd
import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font

from services.xsoar import TicketHandler, XsoarEnvironment
from src.config import XsoarConfig
from src.utils.excel_formatting import apply_professional_formatting

logger = logging.getLogger(__name__)

MAX_EXPORT_DATE_RANGE_DAYS = 366


def _load_incidents(base_dir: str, eastern, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Load incidents from cache or live XSOAR fetch depending on filters.

    Uses cache for slider-based date ranges (last N days).
    Uses live XSOAR fetch for custom date ranges (arbitrary start/end).
    """
    if filters.get('useCustomDate') and filters.get('customDateStart') and filters.get('customDateEnd'):
        from src.components.ticket_cache import TicketCache
        logger.info(f"Export: live XSOAR fetch for {filters['customDateStart']} to {filters['customDateEnd']}")
        all_incidents = TicketCache.fetch_for_range(filters['customDateStart'], filters['customDateEnd'])
    else:
        today_date = datetime.now(eastern).strftime('%m-%d-%Y')
        root_directory = Path(base_dir).parent
        cache_file = root_directory / 'data' / 'transient' / 'secOps' / today_date / 'past_90_days_tickets.json'
        if not cache_file.exists():
            raise FileNotFoundError('Cache file not found')
        with open(cache_file, 'r') as f:
            cached_data = json.load(f)
        all_incidents = cached_data['data'] if isinstance(cached_data, dict) and 'data' in cached_data else cached_data

    return apply_filters_to_incidents(all_incidents, filters)


def _validate_export_date_range(filters: Dict[str, Any]) -> None:
    """Raise ValueError if custom date range exceeds the maximum allowed span."""
    if filters.get('useCustomDate'):
        start = filters.get('customDateStart', '')
        end = filters.get('customDateEnd', '')
        if start and end:
            try:
                span = (datetime.strptime(end, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')).days
                if span > MAX_EXPORT_DATE_RANGE_DAYS:
                    raise ValueError(
                        f'Custom date range cannot exceed {MAX_EXPORT_DATE_RANGE_DAYS} days (selected {span} days)'
                    )
            except ValueError as e:
                if 'Custom date range' in str(e):
                    raise


def get_meaningful_metrics_data(base_dir: str, eastern: pytz.tzinfo.BaseTzInfo) -> Dict[str, Any]:
    """Get cached security incident data for dashboard.

    Args:
        base_dir: Base directory of the web application
        eastern: Pytz timezone object for US/Eastern

    Returns:
        Dictionary with success status, data, total_count, and data_generated_at

    Raises:
        FileNotFoundError: If cache file not found
    """
    logger.info("Loading meaningful metrics data from cache")

    today_date = datetime.now(eastern).strftime('%m-%d-%Y')
    root_directory = Path(base_dir).parent
    cache_file = root_directory / 'data' / 'transient' / 'secOps' / today_date / 'past_90_days_tickets.json'

    if not cache_file.exists():
        raise FileNotFoundError('Cache file not found')

    with open(cache_file, 'r') as f:
        cached_data = json.load(f)

    if isinstance(cached_data, dict) and 'data' in cached_data:
        return {
            'success': True,
            'data': cached_data['data'],
            'total_count': cached_data.get('total_count', len(cached_data['data'])),
            'data_generated_at': cached_data.get('data_generated_at')
        }
    else:
        return {
            'success': True,
            'data': cached_data,
            'total_count': len(cached_data),
            'data_generated_at': None
        }


def apply_filters_to_incidents(incidents: List[Dict[str, Any]], filters: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Apply filters to incidents data - mirrors JavaScript filtering logic.

    Args:
        incidents: List of incident dictionaries
        filters: Dictionary of filter criteria

    Returns:
        Filtered list of incidents
    """
    logger.debug(f"Applying filters to {len(incidents)} incidents")

    date_range = filters.get('dateRange', 30)
    use_custom_date = filters.get('useCustomDate', False)
    custom_date_start = filters.get('customDateStart', '')
    custom_date_end = filters.get('customDateEnd', '')

    # Parse custom date boundaries once
    custom_start_dt = None
    custom_end_dt = None
    if use_custom_date and custom_date_start and custom_date_end:
        try:
            custom_start_dt = datetime.strptime(custom_date_start, '%Y-%m-%d')
            custom_end_dt = datetime.strptime(custom_date_end, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            use_custom_date = False

    mttr_filter = filters.get('mttrFilter', 0)
    mttc_filter = filters.get('mttcFilter', 0)
    age_filter = filters.get('ageFilter', 0)
    countries = filters.get('countries', [])
    regions = filters.get('regions', [])
    impacts = filters.get('impacts', [])
    severities = filters.get('severities', [])
    ticket_types = filters.get('ticketTypes', [])
    statuses = filters.get('statuses', [])
    automation_levels = filters.get('automationLevels', [])
    assignment = filters.get('assignment', 'assigned')
    if assignment not in ('assigned', 'unassigned', 'both'):
        assignment = 'assigned'

    filtered_incidents = []

    for item in incidents:
        # Date filter
        if use_custom_date and custom_start_dt and custom_end_dt:
            created_val = item.get('created')
            if not created_val:
                continue
            try:
                if isinstance(created_val, datetime):
                    created_dt = created_val.replace(tzinfo=None)
                else:
                    created_dt = datetime.fromisoformat(created_val.replace('Z', '+00:00')).replace(tzinfo=None)
            except (ValueError, AttributeError, TypeError):
                continue
            if created_dt < custom_start_dt or created_dt > custom_end_dt:
                continue
        elif item.get('created_days_ago') is not None and item.get('created_days_ago') > date_range:
            continue

        # Location filters
        if countries or regions:
            location_match = False

            if countries:
                has_no_country = not item.get('affected_country') or item.get('affected_country') == 'Unknown' or item.get('affected_country', '').strip() == ''
                should_show_no_country = 'No Country' in countries and has_no_country
                should_show_with_country = any(c != 'No Country' and c == item.get('affected_country') for c in countries)
                location_match = should_show_no_country or should_show_with_country

            if regions:
                has_no_region = not item.get('affected_region') or item.get('affected_region') == 'Unknown' or item.get('affected_region', '').strip() == ''
                should_show_no_region = 'No Region' in regions and has_no_region
                should_show_with_region = any(r != 'No Region' and r == item.get('affected_region') for r in regions)
                location_match = should_show_no_region or should_show_with_region

            if not location_match:
                continue

        # Impact filter
        if impacts:
            has_no_impact = not item.get('impact') or item.get('impact') == 'Unknown' or item.get('impact', '').strip() == ''
            should_show_no_impact = 'No Impact' in impacts and has_no_impact
            should_show_with_impact = any(i != 'No Impact' and i == item.get('impact') for i in impacts)

            if not should_show_no_impact and not should_show_with_impact:
                continue

        # Severity filter
        if severities and str(item.get('severity')) not in severities:
            continue

        # Ticket type filter
        if ticket_types and item.get('type') not in ticket_types:
            continue

        # Status filter
        if statuses and str(item.get('status')) not in statuses:
            continue

        # Automation level filter
        if automation_levels:
            has_no_level = not item.get('automation_level') or item.get('automation_level') == 'Unknown' or item.get('automation_level', '').strip() == ''
            should_show_no_level = 'No Level' in automation_levels and has_no_level
            should_show_with_level = any(l != 'No Level' and l == item.get('automation_level') for l in automation_levels)

            if not should_show_no_level and not should_show_with_level:
                continue

        # Assignment filter (analyst-worked vs unassigned)
        if assignment in ('assigned', 'unassigned'):
            owner_val = item.get('owner') or ''
            is_assigned = bool(owner_val.strip()) if isinstance(owner_val, str) else bool(owner_val)
            if assignment == 'assigned' and not is_assigned:
                continue
            if assignment == 'unassigned' and is_assigned:
                continue

        # MTTR filter
        if mttr_filter > 0:
            mttr_seconds = item.get('time_to_respond_secs')
            if mttr_seconds is None or mttr_seconds == 0:
                continue

            if mttr_filter == 1 and mttr_seconds > 180:
                continue
            if mttr_filter == 2 and mttr_seconds <= 180:
                continue
            if mttr_filter == 3 and mttr_seconds <= 300:
                continue

        # MTTC filter
        if mttc_filter > 0:
            if not item.get('has_hostname'):
                continue

            mttc_seconds = item.get('time_to_contain_secs')
            if mttc_seconds is None or mttc_seconds == 0:
                continue

            if mttc_filter == 1 and mttc_seconds > 300:
                continue
            if mttc_filter == 2 and mttc_seconds > 900:
                continue
            if mttc_filter == 3 and mttc_seconds <= 900:
                continue

        # Age filter
        if age_filter > 0:
            aging_days = item.get('currently_aging_days')
            if aging_days is None or aging_days == '':
                continue
            try:
                if float(aging_days) <= age_filter:
                    continue
            except (ValueError, TypeError):
                continue

        filtered_incidents.append(item)

    logger.debug(f"Filtered to {len(filtered_incidents)} incidents")
    return filtered_incidents


def export_meaningful_metrics(
        base_dir: str,
        eastern: pytz.tzinfo.BaseTzInfo,
        filters: Dict[str, Any],
        visible_columns: List[str],
        column_labels: Dict[str, str],
        include_notes: bool = False
) -> str:
    """Server-side Excel export with professional formatting.

    Args:
        base_dir: Base directory of the web application
        eastern: Pytz timezone object for US/Eastern
        filters: Filter criteria to apply
        visible_columns: List of column IDs to include
        column_labels: Map of column IDs to display labels
        include_notes: Whether to enrich with notes

    Returns:
        Path to temporary Excel file

    Raises:
        FileNotFoundError: If cache file not found
        ValueError: If no incidents to export
    """
    logger.info(f"Exporting meaningful metrics with filters: {filters}")
    _validate_export_date_range(filters)

    incidents = _load_incidents(base_dir, eastern, filters)

    if not incidents:
        raise ValueError('No incidents to export')

    # Enrich with notes if requested
    if include_notes:
        logger.info(f"Enriching {len(incidents)} filtered tickets with notes (on-demand)...")
        incidents = _enrich_incidents_with_notes(incidents)

    # Prepare rows for export
    max_cell_length = 32767
    rows = _prepare_export_rows(incidents, visible_columns, column_labels, max_cell_length)

    # Create DataFrame and export to Excel
    df = pd.DataFrame(rows)

    # Create temporary file
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        temp_path = tmp.name
        df.to_excel(temp_path, index=False, engine='openpyxl')

    # Add hyperlinks and formatting
    _add_hyperlinks_and_formatting(temp_path)

    logger.info(f"Export complete: {temp_path}")
    return temp_path


class ExportConfig:
    """Configuration for meaningful metrics export operations.

    Note: Worker count (MAX_WORKERS) is inherited from XsoarConfig.
    To change worker count and connection pool size, update XsoarConfig.MAX_WORKERS.
    """

    # Log progress every N tickets (for visibility during long exports)
    PROGRESS_LOG_INTERVAL = 25

    # Timeout per ticket when fetching notes (seconds)
    # API has 30s timeout globally, add buffer for processing
    # Default: 45 (allows for API timeout + processing overhead)
    TIMEOUT_PER_TICKET = 45


def _enrich_incidents_with_notes(incidents: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Fetch notes for incidents in parallel with detailed performance tracking."""
    start_time = time.time()
    logger.info(f"Starting note enrichment for {len(incidents)} filtered tickets...")
    logger.info(f"Configuration: workers={XsoarConfig.MAX_WORKERS}, "
                f"timeout={ExportConfig.TIMEOUT_PER_TICKET}s per ticket")

    ticket_handler = TicketHandler(XsoarEnvironment.PROD)
    enriched_incidents = []
    failed_count = 0
    success_count = 0

    # Track timing statistics
    fetch_times = []
    future_start_times = {}

    def fetch_notes_for_incident(incident):
        """Fetch notes with timing tracking."""
        incident_id = incident.get('id')
        fetch_start = time.time()

        if not incident_id:
            incident['notes'] = []
            return incident, 0.0

        try:
            logger.debug(f"Fetching notes for ticket {incident_id}...")
            notes = ticket_handler.get_user_notes(incident_id)
            incident['notes'] = notes if notes else []
            fetch_duration = time.time() - fetch_start

            if notes:
                logger.debug(f"Ticket {incident_id}: fetched {len(notes)} notes in {fetch_duration:.2f}s")
            else:
                logger.debug(f"Ticket {incident_id}: no notes returned in {fetch_duration:.2f}s")

            return incident, fetch_duration
        except Exception as excep:
            fetch_duration = time.time() - fetch_start
            error_type = type(excep).__name__

            if '429' in str(excep) or 'rate limit' in str(excep).lower():
                logger.warning(f"Ticket {incident_id}: Rate limited after {fetch_duration:.2f}s")
            elif 'timeout' in str(excep).lower():
                logger.warning(f"Ticket {incident_id}: Timeout after {fetch_duration:.2f}s ({error_type})")
            else:
                logger.warning(f"Ticket {incident_id}: {error_type} after {fetch_duration:.2f}s - {str(excep)[:100]}")

            incident['notes'] = []
            return incident, fetch_duration

    logger.info(f"Submitting {len(incidents)} tasks to {XsoarConfig.MAX_WORKERS} workers...")

    with ThreadPoolExecutor(max_workers=XsoarConfig.MAX_WORKERS) as executor:
        # Submit all tasks and track start times
        futures = {}
        for incident in incidents:
            future = executor.submit(fetch_notes_for_incident, incident)
            futures[future] = incident
            future_start_times[future] = time.time()

        logger.info("All tasks submitted, waiting for completions...")
        processing_start = time.time()
        completed_count = 0

        # Process completions
        for future in as_completed(futures.keys()):
            completed_count += 1
            incident = futures[future]
            incident_id = incident.get('id', 'unknown')

            try:
                result, fetch_duration = future.result(timeout=ExportConfig.TIMEOUT_PER_TICKET)
                enriched_incidents.append(result)
                fetch_times.append(fetch_duration)

                if result.get('notes'):
                    success_count += 1
                else:
                    failed_count += 1

                # Log progress at intervals
                if completed_count % ExportConfig.PROGRESS_LOG_INTERVAL == 0:
                    elapsed = time.time() - processing_start
                    rate = completed_count / elapsed if elapsed > 0 else 0
                    logger.info(f"Progress: {completed_count}/{len(incidents)} tickets "
                                f"({success_count} with notes, {failed_count} without) - "
                                f"Rate: {rate:.2f} tickets/sec")

            except TimeoutError:
                elapsed_time = time.time() - future_start_times[future]
                logger.error(f"Ticket {incident_id}: timed out after {elapsed_time:.1f}s "
                             f"(limit: {ExportConfig.TIMEOUT_PER_TICKET}s)")
                incident['notes'] = []
                enriched_incidents.append(incident)
                failed_count += 1
                future.cancel()

            except Exception as ex:
                elapsed_time = time.time() - future_start_times[future]
                logger.error(f"Ticket {incident_id}: exception after {elapsed_time:.1f}s - {type(ex).__name__}: {ex}")
                incident['notes'] = []
                enriched_incidents.append(incident)
                failed_count += 1

    total_elapsed = time.time() - start_time

    # Calculate statistics
    avg_fetch_time = sum(fetch_times) / len(fetch_times) if fetch_times else 0
    max_fetch_time = max(fetch_times) if fetch_times else 0
    min_fetch_time = min(fetch_times) if fetch_times else 0
    overall_rate = len(enriched_incidents) / total_elapsed if total_elapsed > 0 else 0

    logger.info("=" * 60)
    logger.info("Note Enrichment Complete")
    logger.info("=" * 60)
    logger.info(f"Total tickets: {len(enriched_incidents)}")
    logger.info(f"Success (with notes): {success_count} ({success_count / len(enriched_incidents) * 100:.1f}%)")
    logger.info(f"Failed (no notes): {failed_count} ({failed_count / len(enriched_incidents) * 100:.1f}%)")
    logger.info(f"Total time: {total_elapsed:.1f}s")
    logger.info(f"Overall rate: {overall_rate:.2f} tickets/sec")
    logger.info(f"Fetch time - Avg: {avg_fetch_time:.2f}s, Min: {min_fetch_time:.2f}s, Max: {max_fetch_time:.2f}s")
    logger.info("=" * 60)

    # Warn if failure rate is high
    if failed_count > len(enriched_incidents) * 0.3:
        failure_pct = (failed_count / len(enriched_incidents)) * 100
        logger.warning(f"HIGH FAILURE RATE: {failure_pct:.1f}% - Consider reducing MAX_WORKERS or increasing TIMEOUT_PER_TICKET")
        logger.warning(f"Current config: MAX_WORKERS={XsoarConfig.MAX_WORKERS}, TIMEOUT_PER_TICKET={ExportConfig.TIMEOUT_PER_TICKET}s")

    return enriched_incidents


_SEVERITY_MAP = {0: 'Unknown', 1: 'Low', 2: 'Medium', 3: 'High', 4: 'Critical'}
_STATUS_MAP = {0: 'Pending', 1: 'Active', 2: 'Closed'}
_COLUMN_PATH_MAP = {
    'timetorespond': 'time_to_respond_secs',
    'timetocontain': 'time_to_contain_secs',
}


def _prepare_single_export_row(
        incident: Dict[str, Any],
        visible_columns: List[str],
        column_labels: Dict[str, str],
        max_cell_length: int,
        sanitize_stats: Dict[str, Any] = None,
) -> List[Any]:
    """Transform a single incident into a list of cell values ordered by visible_columns.

    If `sanitize_stats` is provided, the number of cells that contained
    Excel-illegal control characters is recorded there along with the ticket id,
    so the caller can surface the issue instead of silently cleaning it.
    """
    row: List[Any] = []
    ticket_stripped = False
    for col_id in visible_columns:
        data_field = _COLUMN_PATH_MAP.get(col_id, col_id)
        value = incident.get(data_field)

        if col_id == 'notes':
            value = _format_notes(value, max_cell_length)
        elif col_id == 'severity':
            value = _SEVERITY_MAP.get(value, 'Unknown')
        elif col_id == 'status':
            value = _STATUS_MAP.get(value, 'Unknown')
        elif col_id in ('timetorespond', 'timetocontain'):
            value = _format_duration(value)
        elif col_id in ('created', 'modified', 'closed', 'updated') and value:
            value = _format_date(value)
        elif isinstance(value, list):
            value = ', '.join(str(v) for v in value)

        if isinstance(value, str) and len(value) > max_cell_length:
            truncation_msg = '\n\n[... Content truncated due to Excel cell size limit ...]'
            value = value[:max_cell_length - len(truncation_msg)] + truncation_msg

        if isinstance(value, str):
            cleaned = ILLEGAL_CHARACTERS_RE.sub('', value)
            if cleaned != value:
                if sanitize_stats is not None:
                    sanitize_stats['stripped_cells'] = sanitize_stats.get('stripped_cells', 0) + 1
                    ticket_stripped = True
                logger.warning(
                    f"Stripped Excel-illegal chars from ticket "
                    f"{incident.get('id')} column '{col_id}' "
                    f"({len(value) - len(cleaned)} chars removed)"
                )
                value = cleaned

        row.append(value if value is not None else '')

    if ticket_stripped and sanitize_stats is not None:
        ids = sanitize_stats.setdefault('ticket_ids', [])
        tid = incident.get('id')
        if tid and tid not in ids:
            ids.append(tid)

    return row


def _prepare_export_rows(
        incidents: List[Dict[str, Any]],
        visible_columns: List[str],
        column_labels: Dict[str, str],
        max_cell_length: int
) -> List[Dict[str, Any]]:
    """Prepare incident rows for Excel export (dict form, used by the sync DataFrame path)."""
    rows = []
    for incident in incidents:
        values = _prepare_single_export_row(incident, visible_columns, column_labels, max_cell_length)
        rows.append({column_labels.get(c, c): v for c, v in zip(visible_columns, values)})
    return rows


def _format_notes(notes: Any, max_cell_length: int) -> str:
    """Format notes with truncation."""
    if not isinstance(notes, list) or not notes:
        return ''

    # Check if this is an error marker from failed fetch
    if len(notes) == 1 and isinstance(notes[0], dict) and notes[0].get('_fetch_error'):
        error_msg = notes[0].get('error_message', 'Unknown error')
        return f"[ERROR: Unable to fetch notes - {error_msg}]"

    truncation_message = '\n\n[... Content truncated due to Excel cell size limit. Please view full notes in the web interface ...]'
    reserved_length = len(truncation_message) + 100

    notes_text = ''
    total_length = 0
    truncated = False

    for idx, note in enumerate(notes):
        note_text = note.get('note_text', '')
        author = note.get('author', '')
        timestamp = note.get('created_at', '')
        formatted_note = f"{idx + 1}. Note: {note_text}\nAuthor: {author}\nTimestamp: {timestamp}"
        separator = '\n\n' if idx > 0 else ''
        next_chunk = separator + formatted_note

        if total_length + len(next_chunk) + reserved_length > max_cell_length:
            truncated = True
            break

        notes_text += next_chunk
        total_length += len(next_chunk)

    if truncated:
        notes_text += truncation_message

    return notes_text


def _format_duration(seconds: Any) -> str:
    """Format duration in seconds as MM:SS."""
    if seconds is None or seconds == 0:
        return ''

    try:
        total_seconds = int(float(seconds))
        minutes = total_seconds // 60
        secs = total_seconds % 60
        return f"{minutes}:{secs:02d}"
    except (ValueError, TypeError):
        return ''


def _format_date(value: Any) -> str:
    """Format date as MM/DD/YYYY HH:MM AM ET."""
    try:
        if isinstance(value, str):
            try:
                dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
            except (ValueError, AttributeError):
                dt = datetime.strptime(value, '%Y-%m-%dT%H:%M:%S.%fZ')
        else:
            dt = value

        if dt.tzinfo is None:
            dt = pytz.utc.localize(dt)

        et_tz = pytz.timezone('US/Eastern')
        dt_et = dt.astimezone(et_tz)

        return dt_et.strftime('%m/%d/%Y %I:%M %p ET')
    except Exception as date_err:
        logger.warning(f"Could not format date {value}: {date_err}")
        return str(value)


def _add_hyperlinks_and_formatting(temp_path: str) -> None:
    """Add hyperlinks to ID column and apply professional formatting."""
    wb = load_workbook(temp_path)
    ws = wb.active

    # Find ID column
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    id_col_idx = None
    for idx, header in enumerate(header_row, 1):
        if header and header.lower() == 'id':
            id_col_idx = idx
            break

    # Add hyperlinks to ID cells
    if id_col_idx:
        blue_font = Font(color="0046AD", underline="single")
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=id_col_idx)
            if cell.value:
                ticket_id = cell.value
                ticket_url = f"https://orchestrator.crtx.us.paloaltonetworks.com/Custom/caseinfoid/{ticket_id}"
                cell.hyperlink = ticket_url
                cell.font = blue_font
                cell.value = ticket_id

    wb.save(temp_path)

    # Apply professional formatting
    column_widths = {
        'id': 15,
        'name': 30,
        'severity': 15,
        'status': 15,
        'country': 20,
        'impact': 25,
        'type': 25,
        'owner': 25,
        'created': 25,
        'user notes': 80,
        'notes': 80
    }

    wrap_columns = {'notes', 'impact', 'name', 'user notes'}
    apply_professional_formatting(temp_path, column_widths, wrap_columns, date_columns=set())


def export_meaningful_metrics_async(
        base_dir: str,
        eastern: pytz.tzinfo.BaseTzInfo,
        filters: Dict[str, Any],
        visible_columns: List[str],
        column_labels: Dict[str, str],
        include_notes: bool = False,
        progress_callback=None,
        warning_callback=None,
) -> str:
    """Async version of export with progress tracking.

    Args:
        base_dir: Base directory of the web application
        eastern: Pytz timezone object for US/Eastern
        filters: Filter criteria to apply
        visible_columns: List of column IDs to include
        column_labels: Map of column IDs to display labels
        include_notes: Whether to enrich with notes
        progress_callback: Optional callback(current, total) for progress tracking

    Returns:
        Path to temporary Excel file

    Raises:
        FileNotFoundError: If cache file not found
        ValueError: If no incidents to export
    """
    logger.info(f"Async export started with filters: {filters}")
    _validate_export_date_range(filters)

    incidents = _load_incidents(base_dir, eastern, filters)

    if not incidents:
        raise ValueError('No incidents to export')

    total = len(incidents)
    if progress_callback:
        progress_callback(0, total)

    # Stream rows directly to an openpyxl write-only workbook so we never hold
    # the full enriched-incidents list or a pandas DataFrame in memory. This is
    # the fix for year-sized exports with notes, where accumulating everything
    # before writing was blowing past the VM's RAM and getting OOM-killed.
    max_cell_length = 32767
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append([column_labels.get(c, c) for c in visible_columns])

    sanitize_stats: Dict[str, Any] = {'stripped_cells': 0, 'ticket_ids': []}

    if include_notes:
        logger.info(f"Enriching {total} filtered tickets with notes (async, streaming)...")
        _enrich_and_stream_notes_to_ws(
            incidents, ws, visible_columns, column_labels, max_cell_length,
            progress_callback, warning_callback, sanitize_stats,
        )
    else:
        for idx, incident in enumerate(incidents, start=1):
            ws.append(_prepare_single_export_row(
                incident, visible_columns, column_labels, max_cell_length, sanitize_stats,
            ))
            if progress_callback and (idx % 500 == 0 or idx == total):
                progress_callback(idx, total)

    if warning_callback and sanitize_stats['stripped_cells']:
        sample = sanitize_stats['ticket_ids'][:5]
        more = len(sanitize_stats['ticket_ids']) - len(sample)
        warning_callback(
            f"Stripped Excel-illegal control characters from "
            f"{sanitize_stats['stripped_cells']} cell(s) across "
            f"{len(sanitize_stats['ticket_ids'])} ticket(s). "
            f"Sample ticket IDs: {sample}"
            + (f" (+{more} more)" if more > 0 else "")
        )

    # Release the raw-incidents list before openpyxl materializes the file —
    # _add_hyperlinks_and_formatting re-loads the xlsx, and we want as much
    # headroom as possible for that step.
    incidents = None

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        temp_path = tmp.name
    wb.save(temp_path)
    wb = None

    _add_hyperlinks_and_formatting(temp_path)

    logger.info(f"Async export complete: {temp_path}")
    return temp_path


def _enrich_incidents_with_notes_async(incidents: List[Dict[str, Any]], progress_callback=None) -> List[Dict[str, Any]]:
    """Async version of note enrichment with progress tracking."""
    start_time = time.time()
    processing_start = time.time()

    ticket_handler = TicketHandler(XsoarEnvironment.PROD)
    enriched_incidents = []
    failed_count = 0
    success_count = 0

    # Notify start of processing
    if progress_callback:
        progress_callback(0, len(incidents))

    def fetch_notes_for_incident(incident):
        """Fetch notes for a single incident with simple error handling."""
        incident_id = incident.get('id')
        if not incident_id:
            incident['notes'] = []
            return incident
        try:
            # max_retries=0: No retries - if API call fails/times out, just skip it
            # API has 30s timeout configured globally in xsoar.py
            notes = ticket_handler.get_user_notes(incident_id, max_retries=0)
            incident['notes'] = notes if notes else []
        except Exception as excep:
            # Mark as failed fetch with a special error marker
            # This will show up in the Excel export so users know the fetch failed
            logger.warning(f"Failed to fetch notes for ticket {incident_id}: {excep}")
            incident['notes'] = [{'_fetch_error': True, 'error_message': str(excep)}]
        return incident

    # Parallel execution with bounded queue - only MAX_WORKERS futures in flight at once
    from collections import deque

    with ThreadPoolExecutor(max_workers=XsoarConfig.MAX_WORKERS) as executor:
        pending = deque(incidents)
        futures = {}

        # Submit initial batch (up to MAX_WORKERS)
        for _ in range(min(XsoarConfig.MAX_WORKERS, len(pending))):
            incident = pending.popleft()
            future = executor.submit(fetch_notes_for_incident, incident)
            futures[future] = incident

        completed = 0

        # Process futures as they complete, submitting new ones to keep queue full
        while futures:
            try:
                # Wait max 60 seconds for ANY future to complete
                done_futures = []
                for future in as_completed(list(futures.keys()), timeout=60):
                    result = future.result()
                    enriched_incidents.append(result)

                    # Check if fetch was successful
                    notes = result.get('notes', [])
                    is_error = (len(notes) == 1 and isinstance(notes[0], dict) and notes[0].get('_fetch_error'))
                    has_notes = notes and not is_error

                    if has_notes:
                        success_count += 1
                    else:
                        failed_count += 1

                    completed += 1
                    done_futures.append(future)

                    if progress_callback:
                        progress_callback(completed, len(incidents))

                    if completed % ExportConfig.PROGRESS_LOG_INTERVAL == 0:
                        elapsed = time.time() - processing_start
                        rate = completed / elapsed if elapsed > 0 else 0
                        logger.info(f"Progress: {completed}/{len(incidents)} tickets "
                                    f"({success_count} with notes, {failed_count} without) - "
                                    f"Rate: {rate:.2f} tickets/sec")

                # Remove completed futures and submit new ones
                for future in done_futures:
                    del futures[future]
                    # Submit next incident if any remain
                    if pending:
                        incident = pending.popleft()
                        new_future = executor.submit(fetch_notes_for_incident, incident)
                        futures[new_future] = incident

            except TimeoutError:
                # The active batch hasn't completed in 60s. Don't kill the export —
                # mark the stuck futures as failed for progress/accounting purposes,
                # refill from pending, and keep going. The hung worker threads will
                # drain naturally as urllib3's own retry/timeout fires (worst case
                # ~120s), at which point the queued resubmissions start running.
                stuck_count = len(futures)
                logger.warning(
                    f"Export stall: {stuck_count} active futures hung for 60+ seconds. "
                    f"Marking them as failed and continuing with the remaining "
                    f"{len(pending)} pending tickets."
                )
                for future, incident in list(futures.items()):
                    incident_id = incident.get('id') if incident else 'Unknown'
                    logger.warning(f"Abandoning hung ticket: {incident_id}")
                    if incident:
                        incident['notes'] = [{
                            '_fetch_error': True,
                            'error_message': 'Notes fetch hung - abandoned after 60s of no progress'
                        }]
                        enriched_incidents.append(incident)
                    failed_count += 1
                    completed += 1
                    del futures[future]

                if progress_callback:
                    progress_callback(completed, len(incidents))

                # Refill the in-flight set from pending so the loop has something to wait on.
                # The new submissions queue inside the executor and start running once
                # worker threads free up (as the abandoned hung calls eventually return).
                refill = min(XsoarConfig.MAX_WORKERS, len(pending))
                for _ in range(refill):
                    incident = pending.popleft()
                    new_future = executor.submit(fetch_notes_for_incident, incident)
                    futures[new_future] = incident

                # If nothing is left to wait on, exit the loop cleanly
                if not futures:
                    break
                # Otherwise fall through and re-enter the wait on the new batch

    total_elapsed = time.time() - start_time
    overall_rate = len(enriched_incidents) / total_elapsed if total_elapsed > 0 else 0

    logger.info("=" * 60)
    logger.info("Note Enrichment Complete")
    logger.info("=" * 60)
    logger.info(f"Total tickets: {len(enriched_incidents)}")
    logger.info(f"Success (with notes): {success_count} ({success_count / len(enriched_incidents) * 100:.1f}%)")
    logger.info(f"Failed (no notes): {failed_count} ({failed_count / len(enriched_incidents) * 100:.1f}%)")
    logger.info(f"Total time: {total_elapsed:.1f}s")
    logger.info(f"Overall rate: {overall_rate:.2f} tickets/sec")
    logger.info("=" * 60)

    return enriched_incidents


def _enrich_and_stream_notes_to_ws(
        incidents: List[Dict[str, Any]],
        ws,
        visible_columns: List[str],
        column_labels: Dict[str, str],
        max_cell_length: int,
        progress_callback=None,
        warning_callback=None,
        sanitize_stats: Dict[str, Any] = None,
) -> None:
    """Enrich each incident with notes and append it to the worksheet immediately.

    Peak memory stays O(MAX_WORKERS) instead of O(N): once a future completes
    we transform it to a row, append to the write-only worksheet, and drop
    every reference to it. Mirrors the stall-handling of
    _enrich_incidents_with_notes_async so a 33-minute hung batch gets abandoned
    instead of wedging the export.
    """
    from collections import deque

    start_time = time.time()
    processing_start = time.time()
    ticket_handler = TicketHandler(XsoarEnvironment.PROD)
    total = len(incidents)
    failed_count = 0
    success_count = 0
    fetch_failed_tickets: List[str] = []  # Tickets whose notes fetch threw
    stalled_tickets: List[str] = []  # Tickets abandoned after 60s no-progress

    def fetch_notes_for_incident(incident):
        incident_id = incident.get('id')
        if not incident_id:
            incident['notes'] = []
            return incident
        try:
            notes = ticket_handler.get_user_notes(incident_id, max_retries=0)
            incident['notes'] = notes if notes else []
        except Exception as excep:
            logger.warning(f"Failed to fetch notes for ticket {incident_id}: {excep}")
            incident['notes'] = [{'_fetch_error': True, 'error_message': str(excep)}]
        return incident

    def _append_row(incident):
        ws.append(_prepare_single_export_row(
            incident, visible_columns, column_labels, max_cell_length, sanitize_stats,
        ))

    with ThreadPoolExecutor(max_workers=XsoarConfig.MAX_WORKERS) as executor:
        pending = deque(incidents)
        # Release the caller's reference so we're the only holder of the remaining tickets
        incidents = None

        futures = {}
        for _ in range(min(XsoarConfig.MAX_WORKERS, len(pending))):
            incident = pending.popleft()
            future = executor.submit(fetch_notes_for_incident, incident)
            futures[future] = incident

        completed = 0

        while futures:
            try:
                drained = 0
                for future in as_completed(list(futures.keys()), timeout=60):
                    result = future.result()

                    notes = result.get('notes', [])
                    is_error = (len(notes) == 1 and isinstance(notes[0], dict) and notes[0].get('_fetch_error'))
                    if notes and not is_error:
                        success_count += 1
                    else:
                        failed_count += 1
                        if is_error:
                            tid = result.get('id')
                            if tid:
                                fetch_failed_tickets.append(str(tid))

                    _append_row(result)
                    # Drop references so the enriched ticket can be GC'd — this is
                    # the whole point of the streaming path.
                    del futures[future]
                    result = None
                    drained += 1

                    completed += 1
                    if progress_callback:
                        progress_callback(completed, total)

                    if completed % ExportConfig.PROGRESS_LOG_INTERVAL == 0:
                        elapsed = time.time() - processing_start
                        rate = completed / elapsed if elapsed > 0 else 0
                        logger.info(f"Progress: {completed}/{total} tickets "
                                    f"({success_count} with notes, {failed_count} without) - "
                                    f"Rate: {rate:.2f} tickets/sec")

                # Refill in-flight set from pending, one per future drained this round
                for _ in range(drained):
                    if not pending:
                        break
                    incident = pending.popleft()
                    new_future = executor.submit(fetch_notes_for_incident, incident)
                    futures[new_future] = incident

            except TimeoutError:
                stuck_count = len(futures)
                logger.warning(
                    f"Export stall: {stuck_count} active futures hung for 60+ seconds. "
                    f"Marking them as failed and continuing with the remaining "
                    f"{len(pending)} pending tickets."
                )
                for future, incident in list(futures.items()):
                    incident_id = incident.get('id') if incident else 'Unknown'
                    logger.warning(f"Abandoning hung ticket: {incident_id}")
                    if incident:
                        incident['notes'] = [{
                            '_fetch_error': True,
                            'error_message': 'Notes fetch hung - abandoned after 60s of no progress'
                        }]
                        _append_row(incident)
                        if incident_id and incident_id != 'Unknown':
                            stalled_tickets.append(str(incident_id))
                    failed_count += 1
                    completed += 1
                    del futures[future]

                if progress_callback:
                    progress_callback(completed, total)

                refill = min(XsoarConfig.MAX_WORKERS, len(pending))
                for _ in range(refill):
                    incident = pending.popleft()
                    new_future = executor.submit(fetch_notes_for_incident, incident)
                    futures[new_future] = incident

                if not futures:
                    break

    total_elapsed = time.time() - start_time
    overall_rate = total / total_elapsed if total_elapsed > 0 else 0
    logger.info("=" * 60)
    logger.info("Note Enrichment (streaming) Complete")
    logger.info("=" * 60)
    logger.info(f"Total tickets: {total}")
    if total > 0:
        logger.info(f"Success (with notes): {success_count} ({success_count / total * 100:.1f}%)")
        logger.info(f"Failed (no notes): {failed_count} ({failed_count / total * 100:.1f}%)")
    logger.info(f"Total time: {total_elapsed:.1f}s")
    logger.info(f"Overall rate: {overall_rate:.2f} tickets/sec")
    logger.info("=" * 60)

    if warning_callback:
        if fetch_failed_tickets:
            sample = fetch_failed_tickets[:5]
            more = len(fetch_failed_tickets) - len(sample)
            warning_callback(
                f"Failed to fetch notes for {len(fetch_failed_tickets)} ticket(s); "
                f"their rows show an error marker. Sample IDs: {sample}"
                + (f" (+{more} more)" if more > 0 else "")
            )
        if stalled_tickets:
            sample = stalled_tickets[:5]
            more = len(stalled_tickets) - len(sample)
            warning_callback(
                f"Abandoned {len(stalled_tickets)} ticket(s) after 60s of no "
                f"progress on notes fetch. Sample IDs: {sample}"
                + (f" (+{more} more)" if more > 0 else "")
            )
