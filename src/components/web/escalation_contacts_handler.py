"""Escalation Contacts Handler — SQLite CRUD + Excel migration + ChromaDB rebuild."""

import json
import logging
import sqlite3
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

logger = logging.getLogger(__name__)

# Paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
DATA_DIR = PROJECT_ROOT / "data" / "escalation_contacts"
DB_PATH = DATA_DIR / "escalation_contacts.db"
EXCEL_FILE = PROJECT_ROOT / "local_pdfs_docs" / "Updated_Escalations Paths Global JV Contact listing.xlsx"
REGIONAL_EXCEL_FILE = PROJECT_ROOT / "data" / "transient" / "Regional Contact List.xlsx"
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Order in which to display the worksheet tabs (skip empty / non-data sheets)
REGIONAL_SHEET_ORDER = [
    "Emergency Contact Process",
    "Level 1 Contact List",
    "Cybersecurity",
    "GRCA",
    "CASM",
    "Cloud, AI Data Security",
    "IAM",
    "GS-SAI",
    "Corp Sec",
    "Global Technology",
    "JVs, and Subs",
    "App Owners",
    "Regional Contacts",
    "Vendors",
    "MISC",
]
# Sheets that are documentation (not tabular contact data) — render as text blocks
REGIONAL_DOC_SHEETS = {"Emergency Contact Process"}


@contextmanager
def _get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db():
    with _get_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                region TEXT NOT NULL,
                team TEXT NOT NULL,
                name TEXT NOT NULL,
                title TEXT DEFAULT '',
                email TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                comments TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Migrate: add `comments` column to pre-existing DBs that don't have it
        existing_cols = {r["name"] for r in conn.execute("PRAGMA table_info(contacts)").fetchall()}
        if "comments" not in existing_cols:
            conn.execute("ALTER TABLE contacts ADD COLUMN comments TEXT DEFAULT ''")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sheet_tabs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_name TEXT UNIQUE NOT NULL,
                display_order INTEGER NOT NULL DEFAULT 0,
                columns_json TEXT NOT NULL DEFAULT '[]',
                is_doc INTEGER NOT NULL DEFAULT 0,
                source_file TEXT DEFAULT '',
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sheet_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_id INTEGER NOT NULL,
                row_index INTEGER NOT NULL,
                values_json TEXT NOT NULL DEFAULT '[]',
                FOREIGN KEY (sheet_id) REFERENCES sheet_tabs(id) ON DELETE CASCADE
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sheet_rows_sheet ON sheet_rows(sheet_id, row_index)")
    logger.info("Escalation contacts DB initialized at %s", DB_PATH)


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------

def get_all_contacts() -> Dict[str, List[Dict[str, Any]]]:
    """Return contacts grouped by region, then by team.

    Structure: {region: [{team, contacts: [{id, name, title, email, phone}]}]}
    """
    with _get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM contacts ORDER BY region, sort_order, team, name"
        ).fetchall()

    grouped: Dict[str, Dict[str, List[Dict]]] = {}
    for row in rows:
        r = dict(row)
        region = r["region"]
        team = r["team"]
        grouped.setdefault(region, {})
        grouped[region].setdefault(team, [])
        grouped[region][team].append({
            "id": r["id"],
            "name": r["name"],
            "title": r["title"],
            "email": r["email"],
            "phone": r["phone"],
            "comments": r["comments"] if "comments" in r.keys() else "",
        })

    # Convert to list-of-dicts format for easier template iteration
    result: Dict[str, List[Dict[str, Any]]] = {}
    for region, teams in grouped.items():
        result[region] = [
            {"team": team, "contacts": contacts}
            for team, contacts in teams.items()
        ]
    return result


def get_contact(contact_id: int) -> Optional[Dict[str, Any]]:
    with _get_connection() as conn:
        row = conn.execute("SELECT * FROM contacts WHERE id = ?", (contact_id,)).fetchone()
        return dict(row) if row else None


def create_contact(region: str, team: str, name: str,
                   title: str = "", email: str = "", phone: str = "",
                   comments: str = "") -> int:
    with _get_connection() as conn:
        # Auto sort_order: max + 1 within the same region/team
        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) FROM contacts WHERE region = ? AND team = ?",
            (region, team),
        ).fetchone()[0]
        cursor = conn.execute(
            """INSERT INTO contacts (region, team, name, title, email, phone, comments, sort_order)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (region, team, name, title, email, phone, comments, max_sort + 1),
        )
        return cursor.lastrowid


def update_contact(contact_id: int, **fields) -> bool:
    allowed = {"name", "title", "email", "phone", "region", "team", "comments"}
    updates = {k: v for k, v in fields.items() if k in allowed and v is not None}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [contact_id]
    with _get_connection() as conn:
        conn.execute(
            f"UPDATE contacts SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            values,
        )
    return True


def delete_contact(contact_id: int) -> bool:
    with _get_connection() as conn:
        cursor = conn.execute("DELETE FROM contacts WHERE id = ?", (contact_id,))
        return cursor.rowcount > 0


def get_regions() -> List[str]:
    """Return ordered list of distinct regions."""
    with _get_connection() as conn:
        rows = conn.execute(
            "SELECT DISTINCT region FROM contacts ORDER BY sort_order, region"
        ).fetchall()
        return [r["region"] for r in rows]


def contact_count() -> int:
    with _get_connection() as conn:
        return conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]


# ---------------------------------------------------------------------------
# Excel Migration
# ---------------------------------------------------------------------------

def _clean(val) -> str:
    """Return cleaned string or '' for NaN."""
    if pd.isna(val):
        return ""
    return str(val).strip()


def _insert_batch(conn, rows: List[tuple]):
    conn.executemany(
        """INSERT INTO contacts (region, team, name, title, email, phone, sort_order)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        rows,
    )


def _parse_global(xl) -> List[tuple]:
    """Parse the Global sheet — only truly global contacts (skip regional duplicates)."""
    df = pd.read_excel(xl, sheet_name="Global")
    rows = []
    order = 0
    current_team = "General"
    # Teams that match dedicated region sheets — skip these to avoid duplicates
    regional_teams = {"EMEA", "LATAM", "ASIA", "APAC", "JAPAN"}

    for _, r in df.iterrows():
        # Column layout: the company(0) | Name(1) | Unnamed:2(2) | Email Address(3) | ... | ESXI(6..9)
        col_org = _clean(r.iloc[0])       # column 0: org/JV name (Versant, Legal Plans, etc.)
        col_name = _clean(r.iloc[1]) if len(r) > 1 else ""   # column 1: sub-team (Tier 1, Escalation)
        col_contact = _clean(r.iloc[2]) if len(r) > 2 else ""
        col_email = _clean(r.iloc[3]) if len(r) > 3 else ""

        # Column 0 sets team for JV/org rows (Versant, Legal Plans, Raven, etc.)
        if col_org:
            current_team = col_org.replace("\n", " ").strip()

        # Column 1 overrides team for sub-sections (Tier 1, Escalation)
        if col_name:
            current_team = col_name

        if current_team in regional_teams:
            continue

        # Handle multi-line contacts (e.g. "Carol Andreu\nDavid Harper")
        if col_contact:
            import re
            contacts_list = [c.strip() for c in re.split(r'[\n,]+', col_contact) if c.strip()]
            emails_list = [e.strip() for e in re.split(r'[\n,]+', col_email) if e.strip()] if col_email else []

            for i, name in enumerate(contacts_list):
                email = emails_list[i] if i < len(emails_list) else ""
                rows.append(("Global", current_team, name, "", email, "", order))
                order += 1
            # Emails without a matching contact name
            for email in emails_list[len(contacts_list):]:
                rows.append(("Global", current_team, email.split("@")[0], "", email, "", order))
                order += 1
        elif col_email and not col_contact:
            # Email-only row (no contact name)
            import re
            for email in re.split(r'[\n,;]+', col_email):
                email = email.strip()
                if email:
                    rows.append(("Global", current_team, email.split("@")[0], "", email, "", order))
                    order += 1

        # Right-side columns (ESXI Escalation Contacts)
        esxi_label = _clean(r.iloc[6]) if len(r) > 6 else ""
        esxi_name = _clean(r.iloc[8]) if len(r) > 8 else ""
        esxi_region = _clean(r.iloc[9]) if len(r) > 9 else ""
        if esxi_label and "Escalation" in esxi_label:
            continue
        if esxi_name:
            team_label = f"ESXI Escalation - {esxi_region}" if esxi_region else "ESXI Escalation"
            parts = esxi_name.split("+", 1)
            ename = parts[0].strip()
            ephone = ("+" + parts[1].strip()) if len(parts) > 1 else ""
            rows.append(("Global", team_label, ename, "", "", ephone, order))
            order += 1

    return rows


def _parse_asia(xl) -> List[tuple]:
    """Parse the ASIA sheet — regional + country POCs."""
    df = pd.read_excel(xl, sheet_name="ASIA")
    rows = []
    order = 0
    current_team = "ASIA-CIRT"

    for _, r in df.iterrows():
        col0 = _clean(r.iloc[0])
        col1 = _clean(r.iloc[1]) if len(r) > 1 else ""

        # Section header: "Country Point of Contact - X"
        if col0.startswith("Country Point of Contact"):
            current_team = col0.replace("Country Point of Contact - ", "").replace("Country Point of Contact -", "").strip()
            if not current_team:
                current_team = "Unknown"
            continue

        # Sub-header rows (Name, Contact, etc.)
        if col0.lower() in ("name", "contact", ""):
            if not col0:
                continue
            # Check if it's a header row
            if col1.lower() in ("contact", "email", "email address", ""):
                continue

        # Skip "ASIA POC" header columns on the right
        if col0 in ("ASIA POC", "ASIA-CIRT DL"):
            if col0 == "ASIA-CIRT DL":
                rows.append(("APAC", "ASIA-CIRT", "ASIA-CIRT DL", "", col1, "", order))
                order += 1
            continue

        if not col0:
            continue

        # Data row: col0=name, col1=email
        name = col0
        email = col1
        # Right-side POC columns
        poc_name = _clean(r.iloc[4]) if len(r) > 4 else ""
        poc_phone = _clean(r.iloc[5]) if len(r) > 5 else ""

        rows.append(("APAC", current_team, name, "", email, "", order))
        order += 1

        if poc_name and poc_name not in ("ASIA POC", "Contact", name):
            rows.append(("APAC", "ASIA POC", poc_name, "", "", poc_phone, order))
            order += 1

    return rows


def _parse_latam(xl) -> List[tuple]:
    """Parse the LATAM sheet."""
    df = pd.read_excel(xl, sheet_name="LATAM")
    rows = []
    order = 0
    current_team = "LATAM-CIRT"

    for _, r in df.iterrows():
        col0 = _clean(r.iloc[0])
        col1 = _clean(r.iloc[1]) if len(r) > 1 else ""
        col2 = _clean(r.iloc[2]) if len(r) > 2 else ""

        # Country name in col0
        if col0 and "@" not in col0:
            current_team = col0

        # Contact name in col1 with email in col2
        if col1 and "@" not in col1:
            email = col2
            rows.append(("LATAM", current_team, col1, "", email, "", order))
            order += 1
        elif col1 and "@" in col1:
            # col1 is an email (country row where email is in col1)
            rows.append(("LATAM", current_team, current_team, "", col1, "", order))
            order += 1

        # Right-side POC columns
        poc_name = _clean(r.iloc[5]) if len(r) > 5 else ""
        poc_phone = _clean(r.iloc[6]) if len(r) > 6 else ""
        if poc_name and poc_name not in ("LATAM POC", "Contacts"):
            rows.append(("LATAM", "LATAM POC", poc_name, "", "", poc_phone, order))
            order += 1

    return rows


def _parse_emea(xl) -> List[tuple]:
    """Parse the EMEA sheet."""
    df = pd.read_excel(xl, sheet_name="EMEA")
    rows = []
    order = 0

    for _, r in df.iterrows():
        # Left side: EMEA-CIRT members
        col1 = _clean(r.iloc[1]) if len(r) > 1 else ""
        col2 = _clean(r.iloc[2]) if len(r) > 2 else ""
        if col1:
            rows.append(("EMEA", "EMEA-CIRT", col1, "", col2, "", order))
            order += 1

        # Right side: EMEA POC
        poc_name = _clean(r.iloc[5]) if len(r) > 5 else ""
        poc_phone = _clean(r.iloc[6]) if len(r) > 6 else ""
        if poc_name and poc_name not in ("EMEA", "EMEA "):
            rows.append(("EMEA", "EMEA POC", poc_name, "", "", poc_phone, order))
            order += 1

    return rows


def _parse_apac(xl) -> List[tuple]:
    """Parse the APAC sheet — country → semicolon/newline separated emails."""
    df = pd.read_excel(xl, sheet_name="APAC")
    rows = []
    order = 0

    for _, r in df.iterrows():
        country = _clean(r.iloc[0])
        emails_raw = _clean(r.iloc[1]) if len(r) > 1 else ""
        if not country or not emails_raw:
            continue
        # Split on semicolons, newlines, whitespace combos
        import re
        emails = [e.strip() for e in re.split(r'[;\n]+', emails_raw) if e.strip()]
        for email in emails:
            if "@" in email:
                # Extract name from email prefix
                name_part = email.split("@")[0].replace(".", " ").replace("_", " ").replace("-", " ").title()
                rows.append(("APAC", country, name_part, "", email.strip(), "", order))
                order += 1

    return rows


def _parse_japan(xl) -> List[tuple]:
    """Parse the JAPAN sheet."""
    df = pd.read_excel(xl, sheet_name="JAPAN")
    rows = []
    order = 0

    for _, r in df.iterrows():
        col1 = _clean(r.iloc[1]) if len(r) > 1 else ""
        col2 = _clean(r.iloc[2]) if len(r) > 2 else ""
        if col1:
            rows.append(("JAPAN", "JAPAN-CIRT", col1, "", col2, "", order))
            order += 1

    return rows


def migrate_from_excel(force: bool = False) -> int:
    """One-time seed from Excel file. Skips if DB already has data unless forced.

    Returns number of contacts inserted.
    """
    if not force and contact_count() > 0:
        logger.info("Contacts DB already populated (%d rows), skipping migration", contact_count())
        return 0

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    xl = pd.ExcelFile(EXCEL_FILE)

    all_rows: List[tuple] = []
    parsers = {
        "Global": _parse_global,
        "ASIA": _parse_asia,
        "LATAM": _parse_latam,
        "EMEA": _parse_emea,
        "APAC": _parse_apac,
        "JAPAN": _parse_japan,
    }

    for sheet, parser in parsers.items():
        if sheet in xl.sheet_names:
            try:
                sheet_rows = parser(xl)
                all_rows.extend(sheet_rows)
                logger.info("Parsed %d contacts from %s sheet", len(sheet_rows), sheet)
            except Exception as e:
                logger.error("Error parsing %s sheet: %s", sheet, e, exc_info=True)

    if not all_rows:
        logger.warning("No contacts parsed from Excel")
        return 0

    with _get_connection() as conn:
        if force:
            conn.execute("DELETE FROM contacts")
        _insert_batch(conn, all_rows)

    logger.info("Migrated %d contacts from Excel to SQLite", len(all_rows))
    return len(all_rows)


# ---------------------------------------------------------------------------
# Regional Contact List — generic per-worksheet tab storage
# ---------------------------------------------------------------------------

def _cell_to_str(val) -> str:
    """Render any Excel cell value as a clean display string."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    s = str(val).strip()
    # Excel non-breaking spaces
    s = s.replace("\xa0", " ")
    return s


def _load_worksheet(xlsx_path: Path, sheet_name: str) -> Dict[str, Any]:
    """Load a worksheet into a normalized dict.

    Returns: {name, columns: [str], rows: [[str, ...]], is_doc: bool}
    For doc sheets, columns is empty and rows is a list of single-string lines.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"name": sheet_name, "columns": [], "rows": [], "is_doc": False}
    ws = wb[sheet_name]

    raw_rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_to_str(c) for c in row]
        # Drop trailing empties
        while cells and not cells[-1]:
            cells.pop()
        raw_rows.append(cells)
    wb.close()

    # Strip leading/trailing fully-empty rows
    while raw_rows and not any(raw_rows[0]):
        raw_rows.pop(0)
    while raw_rows and not any(raw_rows[-1]):
        raw_rows.pop()

    if not raw_rows:
        return {"name": sheet_name, "columns": [], "rows": [], "is_doc": False}

    # Doc sheet: render as paragraph list (only column-0 has content for most rows)
    if sheet_name in REGIONAL_DOC_SHEETS:
        lines = []
        for r in raw_rows:
            text = " ".join(c for c in r if c).strip()
            if text:
                lines.append([text])
        return {"name": sheet_name, "columns": [], "rows": lines, "is_doc": True}

    # Treat first non-empty row as headers
    headers = raw_rows[0]
    # Drop trailing empty headers
    while headers and not headers[-1]:
        headers.pop()
    if not headers:
        return {"name": sheet_name, "columns": [], "rows": [], "is_doc": False}

    n_cols = len(headers)
    headers = [h or f"Col {i + 1}" for i, h in enumerate(headers)]

    body = []
    for r in raw_rows[1:]:
        # Pad / truncate to header width
        cells = list(r) + [""] * max(0, n_cols - len(r))
        cells = cells[:n_cols]
        if any(cells):
            body.append(cells)

    return {"name": sheet_name, "columns": headers, "rows": body, "is_doc": False}


def migrate_from_regional_excel(force: bool = False) -> int:
    """Load all worksheets from Regional Contact List.xlsx into sheet_tabs/sheet_rows.

    Returns: number of rows loaded across all sheets.
    Always replaces existing sheet data when called (idempotent).
    """
    if not REGIONAL_EXCEL_FILE.exists():
        logger.info("Regional Contact List Excel not found at %s — skipping", REGIONAL_EXCEL_FILE)
        return 0

    if not force:
        with _get_connection() as conn:
            existing = conn.execute("SELECT COUNT(*) FROM sheet_rows").fetchone()[0]
            if existing > 0:
                logger.info("sheet_rows already populated (%d rows), skipping migration", existing)
                return 0

    total_rows = 0
    with _get_connection() as conn:
        # Wipe all sheet data — we re-load from scratch
        conn.execute("DELETE FROM sheet_rows")
        conn.execute("DELETE FROM sheet_tabs")

        for order, sheet_name in enumerate(REGIONAL_SHEET_ORDER):
            try:
                ws_data = _load_worksheet(REGIONAL_EXCEL_FILE, sheet_name)
            except Exception as e:
                logger.error("Failed to load worksheet %s: %s", sheet_name, e, exc_info=True)
                continue

            # Skip empty sheets entirely (no rows)
            if not ws_data["rows"] and not ws_data["is_doc"]:
                logger.info("Skipping empty sheet: %s", sheet_name)
                continue

            cur = conn.execute(
                """INSERT INTO sheet_tabs (sheet_name, display_order, columns_json, is_doc, source_file)
                   VALUES (?, ?, ?, ?, ?)""",
                (
                    sheet_name,
                    order,
                    json.dumps(ws_data["columns"]),
                    1 if ws_data["is_doc"] else 0,
                    str(REGIONAL_EXCEL_FILE.name),
                ),
            )
            sheet_id = cur.lastrowid

            row_payload = [
                (sheet_id, idx, json.dumps(row))
                for idx, row in enumerate(ws_data["rows"])
            ]
            if row_payload:
                conn.executemany(
                    "INSERT INTO sheet_rows (sheet_id, row_index, values_json) VALUES (?, ?, ?)",
                    row_payload,
                )
                total_rows += len(row_payload)
            logger.info("Loaded %d rows for sheet '%s'", len(row_payload), sheet_name)

    logger.info("Regional Contact List migration complete: %d rows across %d sheets",
                total_rows, len(REGIONAL_SHEET_ORDER))
    return total_rows


def get_all_sheet_tabs() -> List[Dict[str, Any]]:
    """Return all sheet tabs in display order, each with columns and rows."""
    with _get_connection() as conn:
        tabs = conn.execute(
            "SELECT id, sheet_name, display_order, columns_json, is_doc FROM sheet_tabs ORDER BY display_order"
        ).fetchall()

        result = []
        for tab in tabs:
            t = dict(tab)
            rows = conn.execute(
                "SELECT values_json FROM sheet_rows WHERE sheet_id = ? ORDER BY row_index",
                (t["id"],),
            ).fetchall()
            result.append({
                "name": t["sheet_name"],
                "slug": _sheet_slug(t["sheet_name"]),
                "columns": json.loads(t["columns_json"]),
                "rows": [json.loads(r["values_json"]) for r in rows],
                "is_doc": bool(t["is_doc"]),
            })
        return result


def _sheet_slug(name: str) -> str:
    """Make a tab/anchor-safe slug from a sheet name."""
    out = []
    for ch in name.lower():
        if ch.isalnum():
            out.append(ch)
        elif ch in (" ", "-", "_"):
            out.append("-")
    slug = "".join(out)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-") or "sheet"


# ---------------------------------------------------------------------------
# Rebuild ChromaDB embeddings
# ---------------------------------------------------------------------------

def rebuild_embeddings() -> Dict[str, Any]:
    """Rebuild the contacts ChromaDB vector store from SQLite data."""
    try:
        from src.components.contacts_lookup import get_contacts_store
        store = get_contacts_store()
        store.rebuild()
        count = store.collection.count() if store.collection else 0
        return {"success": True, "message": f"Rebuilt embeddings for {count} entries"}
    except Exception as e:
        logger.error("Failed to rebuild embeddings: %s", e, exc_info=True)
        return {"success": False, "error": str(e)}


# Auto-init and migrate on import
init_db()
try:
    migrate_from_excel()
except Exception:
    logger.debug("Excel migration skipped (file missing or error)", exc_info=True)
try:
    migrate_from_regional_excel()
except Exception:
    logger.debug("Regional Excel migration skipped (file missing or error)", exc_info=True)
