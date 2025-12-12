"""
CS Host Invalid Ring Tag Analyzer

Identifies CrowdStrike hosts that have incorrect ring tags based on:

1. Environment validation (all host types):
   - Ring 1: dev, poc, lab, integration, development environments
   - Ring 2: qa, test environments
   - Ring 3: dr (disaster recovery) environments
   - Ring 4: production or unknown environments

2. Country/Region validation (all hosts):
   - Validates ring tag region matches expected region for host's current country
   - Example: A host in France (EMEA region) should not have USASRVRing tags

Creates two reports:
- Complete dataset of all hosts with ring tags
- Filtered report showing only hosts with invalid ring tags
"""

import json
import logging
import re
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle

from services import crowdstrike, service_now
from src.epp.cs_hosts_without_ring_tag import get_dated_path
from src.utils.excel_formatting import apply_professional_formatting

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

ROOT_DIRECTORY = Path(__file__).parent.parent.parent
DATA_DIR = ROOT_DIRECTORY / "data" / "transient" / "epp_device_tagging"

# Timezone constant for consistent usage
EASTERN_TZ = ZoneInfo("America/New_York")

# Environment to ring mappings (all values must be lowercase)
RING_1_ENVS = {"dev", "poc", "lab", "integration", "development"}
RING_2_ENVS = {"qa", "test"}
RING_3_ENVS = {"dr", "qa/dr"}

# Ring 4 is for production or unknown environments

# Load region mappings from JSON file for country-based validation
REGIONS_FILE = ROOT_DIRECTORY / "data" / "regions_by_country.json"
with open(REGIONS_FILE, 'r') as f:
    REGIONS_BY_COUNTRY = json.load(f)


def get_expected_ring(env):
    """Return expected ring number based on environment."""
    if env in RING_1_ENVS:
        return 1
    if env in RING_2_ENVS:
        return 2
    if env in RING_3_ENVS:
        return 3
    return 4


def extract_region_from_ring_tag(ring_tag):
    """Extract region code from a ring tag.

    Examples:
        'FalconGroupingTags/USSRVRing1' -> 'US'
        'FalconGroupingTags/EMEASRVRing2' -> 'EMEA'
        'FalconGroupingTags/JPSRVRing4' -> 'JP'

    Returns the region code or None if pattern doesn't match.
    """
    match = re.search(r'FalconGroupingTags/([A-Z]+)SRVRing\d+', ring_tag, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return None


def adjust_column_widths(file_path):
    """Adjust column widths, format headers, and add professional formatting in Excel file."""
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        # Column width mappings for common columns
        column_widths = {
            'hostname': 25,
            'host_id': 20,
            'current_tags': 80,
            'invalid_tags': 60,
            'last_seen': 20,
            'status': 15,
            'cs_host_category': 20,
            'SNOW_environment': 15,
            'SNOW_country': 20,
            'SNOW_lifecycleStatus': 20,
            'comment': 50,
            'environment': 15,
            'platform': 15,
        }

        # Text wrap columns (for long content)
        wrap_columns = {'current_tags', 'invalid_tags', 'comment'}

        # Date columns (for date formatting)
        date_columns = {'last_seen'}

        # Get header row to map column names to letters
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')

        # Zebra stripe colors
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Create date style
        date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY HH:MM')

        # Format headers
        for col_idx, header in enumerate(header_row, 1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            cell = worksheet.cell(row=1, column=col_idx)

            # Set column width
            if header and header.lower() in column_widths:
                worksheet.column_dimensions[col_letter].width = column_widths[header.lower()]
            else:
                worksheet.column_dimensions[col_letter].width = 15

            # Format header cell
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Format data rows
        for row_idx in range(2, worksheet.max_row + 1):
            # Zebra striping - every other row
            is_alternate_row = (row_idx % 2 == 0)

            for col_idx, header in enumerate(header_row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Add borders to all cells
                cell.border = thin_border

                # Zebra striping
                if is_alternate_row:
                    cell.fill = light_fill

                # Text wrapping for long content columns
                if header and header.lower() in wrap_columns:
                    cell.alignment = wrap_alignment

                # Date formatting
                if header and header.lower() in date_columns and cell.value:
                    cell.style = date_style

        # Freeze the header row
        worksheet.freeze_panes = 'A2'

        # Add auto filter to the data range
        if worksheet.max_row > 1:  # Only add filter if there's data beyond headers
            worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"

        workbook.save(file_path)
        logger.info(f"Applied professional formatting to {file_path}")
    except Exception as e:
        logger.warning(f"Could not format Excel file {file_path}: {e}")


def analyze_ring_tags(servers_df):
    """Analyze servers and mark those with invalid ring tags, completely ignoring Citrix rings."""
    servers_df['has_invalid_ring_tag'] = False
    servers_df['comment'] = ''

    total_servers = len(servers_df)
    logger.info(f"Starting ring tag analysis for {total_servers} servers")

    for idx, (index, server) in enumerate(servers_df.iterrows(), 1):
        # Log progress every 500 rows
        if idx % 500 == 0:
            percent_complete = (idx / total_servers) * 100
            logger.info(f"Progress: {idx}/{total_servers} servers analyzed ({percent_complete:.1f}%)")

        env = str(server.get('SNOW_environment', '')).lower()
        current_tags = server.get('current_tags', '')

        # Extract all ring numbers first
        all_ring_numbers = re.findall(r'FalconGroupingTags/[^,]*?SRVRing(\d+)', current_tags, re.IGNORECASE)

        # Find complete ring tag strings to check for Citrix
        complete_ring_tag_matches = re.findall(r'(FalconGroupingTags/[^,]*?SRVRing\d+)', current_tags, re.IGNORECASE)

        # Filter to only non-Citrix ring numbers by checking the complete tag
        ring_numbers = []
        for i, ring_num in enumerate(all_ring_numbers):
            if 'citrix' not in complete_ring_tag_matches[i].lower():
                ring_numbers.append(int(ring_num))

        if not ring_numbers:
            continue

        # SPECIAL CASE: Skip validation for Ring 0 hosts
        # Ring 0 hosts are exempt from environment-based ring validation rules.
        # Unlike other rings (1-4) which must match their environment type,
        # Ring 0 hosts are considered valid regardless of their environment.
        # This exception overrides the general validation logic of this file.
        if any(ring == 0 for ring in ring_numbers):
            continue

        if len(ring_numbers) > 1:
            servers_df.loc[index, 'has_invalid_ring_tag'] = True
            servers_df.loc[index, 'comment'] = 'multiple ring tags found'
            logger.info(f"Multiple non-Citrix ring tags found for host {server.get('hostname', 'Unknown')}: {ring_numbers}")
            continue

        expected_ring = get_expected_ring(env)
        if any(num != expected_ring for num in ring_numbers):
            servers_df.loc[index, 'has_invalid_ring_tag'] = True
            servers_df.loc[index, 'comment'] = f'{env} server should not be in Ring {ring_numbers}, expected Ring {expected_ring}'
            logger.info(f"Invalid ring tag found for host {server.get('hostname', 'Unknown')}: Ring {ring_numbers}, expected Ring {expected_ring}")

        # Country-based validation (additive to environment validation)
        # This checks if the region in the ring tag matches the expected region for the host's country
        country = str(server.get('SNOW_country', '')).strip()

        if country and country in REGIONS_BY_COUNTRY:
            expected_region = REGIONS_BY_COUNTRY[country]

            # Extract region from the current ring tag
            complete_ring_tag_matches = re.findall(r'(FalconGroupingTags/[^,]*?SRVRing\d+)', current_tags, re.IGNORECASE)

            for ring_tag in complete_ring_tag_matches:
                # Skip Citrix rings
                if 'citrix' in ring_tag.lower():
                    continue

                # Skip Ring 0 (exempt from all validation)
                ring_num_match = re.search(r'SRVRing(\d+)', ring_tag, re.IGNORECASE)
                if ring_num_match and int(ring_num_match.group(1)) == 0:
                    continue

                # Extract region from the ring tag
                actual_region = extract_region_from_ring_tag(ring_tag)

                if actual_region and actual_region != expected_region:
                    # Region mismatch detected
                    servers_df.loc[index, 'has_invalid_ring_tag'] = True

                    # Append to comment (may already have environment-based comment)
                    existing_comment = servers_df.loc[index, 'comment']
                    country_comment = f"host in country '{country}' (region {expected_region}) has ring tag for region {actual_region}"

                    if existing_comment:
                        servers_df.loc[index, 'comment'] = f"{existing_comment}; {country_comment}"
                    else:
                        servers_df.loc[index, 'comment'] = country_comment

                    logger.info(f"Country-based invalid ring tag for host {server.get('hostname', 'Unknown')}: "
                               f"Expected region {expected_region} (from country '{country}'), "
                               f"but has ring tag for region {actual_region}")

    # Log completion
    invalid_count = servers_df['has_invalid_ring_tag'].sum()
    logger.info(f"Completed ring tag analysis: {total_servers} servers processed, {invalid_count} with invalid tags")


def generate_report():
    """Generate the complete invalid ring tag analysis report."""
    today_date = datetime.now(EASTERN_TZ).strftime('%m-%d-%Y')
    output_dir = DATA_DIR / today_date
    output_dir.mkdir(parents=True, exist_ok=True)

    # Get input file
    input_file_path = get_dated_path(DATA_DIR, "unique_cs_hosts.xlsx")
    if not input_file_path.exists():
        logger.info("Unique hosts file not found. Generating it now...")
        crowdstrike.update_unique_hosts_from_cs()

    # Read data
    unique_cs_hosts_df = pd.read_excel(input_file_path, engine="openpyxl")
    logger.info(f"Read {len(unique_cs_hosts_df)} records from {input_file_path}")

    # Filter hosts with ring tags (all categories)
    servers_with_ring_tags = unique_cs_hosts_df[
        unique_cs_hosts_df["current_tags"].str.contains("FalconGroupingTags/.*SrvRing", regex=True, case=False, na=False)
    ]
    logger.info(f"Found {len(servers_with_ring_tags)} hosts with ring tags")

    # Save servers with ring tags
    servers_with_ring_tags_file_path = output_dir / "cs_hosts_with_ring_tags.xlsx"
    servers_with_ring_tags.to_excel(servers_with_ring_tags_file_path, index=False, engine="openpyxl")
    apply_professional_formatting(servers_with_ring_tags_file_path)

    # Enrich with ServiceNow data
    logger.info(f"Starting ServiceNow enrichment for {len(servers_with_ring_tags)} servers")
    enriched_file_path = service_now.enrich_host_report(servers_with_ring_tags_file_path)
    logger.info("ServiceNow enrichment completed")
    enriched_servers = pd.read_excel(enriched_file_path, engine="openpyxl")

    # Analyze ring tags
    analyze_ring_tags(enriched_servers)

    # Save complete report
    complete_report_path = output_dir / "cs_hosts_last_seen_with_invalid_ring_tags.xlsx"
    enriched_servers.to_excel(complete_report_path, index=False, engine="openpyxl")
    apply_professional_formatting(complete_report_path)

    # Save filtered report (invalid tags only)
    invalid_servers = enriched_servers[enriched_servers['has_invalid_ring_tag']].copy()
    if not invalid_servers.empty:
        # Add invalid_tags column
        def extract_invalid_tags(row):
            current_tags = row.get('current_tags', '')
            # Extract all non-Citrix ring tags
            all_ring_numbers = re.findall(r'FalconGroupingTags/[^,]*?SRVRing(\d+)', current_tags, re.IGNORECASE)
            complete_ring_tag_matches = re.findall(r'(FalconGroupingTags/[^,]*?SRVRing\d+)', current_tags, re.IGNORECASE)
            invalid = []
            for i, ring_num in enumerate(all_ring_numbers):
                if 'citrix' not in complete_ring_tag_matches[i].lower():
                    invalid.append(complete_ring_tag_matches[i])
            return ', '.join(invalid) if invalid else ''

        invalid_servers['invalid_tags'] = invalid_servers.apply(extract_invalid_tags, axis=1)

        # Only keep the requested columns
        columns_to_keep = [
            'hostname',
            'host_id',
            'current_tags',
            'invalid_tags',
            'last_seen',
            'status',
            'cs_host_category',
            'SNOW_environment',
            'SNOW_country',
            'SNOW_lifecycleStatus',
            'comment',
        ]
        filtered_report_path = output_dir / "cs_hosts_with_invalid_ring_tags_only.xlsx"
        invalid_servers[columns_to_keep].to_excel(filtered_report_path, index=False, engine="openpyxl")
        apply_professional_formatting(filtered_report_path)
        logger.info(f"Found {len(invalid_servers)} hosts with invalid ring tags")
    else:
        logger.info("No hosts with invalid ring tags found")


def main():
    """Main function to run the complete workflow."""
    try:
        logger.info("Starting CrowdStrike host invalid ring tag analysis")
        generate_report()
        logger.info("Completed CrowdStrike host invalid ring tag analysis")
    except FileNotFoundError as e:
        logger.error(f"Input file not found: {e}")
        raise
    except Exception as e:
        logger.error(f"Error in main workflow: {e}")
        raise


if __name__ in ('__main__', '__builtin__', 'builtins'):
    main()
