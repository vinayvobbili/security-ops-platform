"""
CS Host Invalid Ring Tag Analyzer

Identifies CrowdStrike servers that have incorrect ring tags based on their environment:
- Ring 1: dev, poc, lab, integration, development environments
- Ring 2: qa, test environments
- Ring 3: dr (disaster recovery) environments
- Ring 4: production or unknown environments

Creates two reports:
- Complete dataset of all servers with ring tags
- Filtered report showing only servers with invalid ring tags
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle

from services import crowdstrike, service_now
from src.epp.cs_hosts_without_ring_tag import get_dated_path

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

ROOT_DIRECTORY = Path(__file__).parent.parent.parent
DATA_DIR = ROOT_DIRECTORY / "data" / "transient" / "epp_device_tagging"

# Timezone constant for consistent usage
EASTERN_TZ = ZoneInfo("America/New_York")

# Environment to ring mappings (all values must be lowercase)
RING_1_ENVS = {"dev", "poc", "lab", "integration", "development"}
RING_2_ENVS = {"qa", "test"}
RING_3_ENVS = {"dr"}


# Ring 4 is for production or unknown environments


def get_expected_ring(env):
    """Return expected ring number based on environment."""
    if env in RING_1_ENVS:
        return 1
    if env in RING_2_ENVS:
        return 2
    if env in RING_3_ENVS:
        return 3
    return 4


def adjust_column_widths(file_path):
    """Adjust column widths, format headers, and add professional formatting in Excel file."""
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        # Column width mappings for common columns
        column_widths = {
            'hostname': 25,
            'host_id': 20,
            'current_tags': 80,
            'invalid_tags': 60,
            'last_seen': 20,
            'status': 15,
            'cs_host_category': 20,
            'SNOW_environment': 15,
            'SNOW_lifecycleStatus': 20,
            'comment': 50,
            'environment': 15,
            'platform': 15,
        }

        # Text wrap columns (for long content)
        wrap_columns = {'current_tags', 'invalid_tags', 'comment'}

        # Date columns (for date formatting)
        date_columns = {'last_seen'}

        # Get header row to map column names to letters
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')

        # Zebra stripe colors
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Create date style
        date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY HH:MM')

        # Format headers
        for col_idx, header in enumerate(header_row, 1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            cell = worksheet.cell(row=1, column=col_idx)

            # Set column width
            if header and header.lower() in column_widths:
                worksheet.column_dimensions[col_letter].width = column_widths[header.lower()]
            else:
                worksheet.column_dimensions[col_letter].width = 15

            # Format header cell
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Format data rows
        for row_idx in range(2, worksheet.max_row + 1):
            # Zebra striping - every other row
            is_alternate_row = (row_idx % 2 == 0)

            for col_idx, header in enumerate(header_row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Add borders to all cells
                cell.border = thin_border

                # Zebra striping
                if is_alternate_row:
                    cell.fill = light_fill

                # Text wrapping for long content columns
                if header and header.lower() in wrap_columns:
                    cell.alignment = wrap_alignment

                # Date formatting
                if header and header.lower() in date_columns and cell.value:
                    cell.style = date_style

        # Freeze the header row
        worksheet.freeze_panes = 'A2'

        # Add auto filter to the data range
        if worksheet.max_row > 1:  # Only add filter if there's data beyond headers
            worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"

        workbook.save(file_path)
        logger.info(f"Applied professional formatting to {file_path}")
    except Exception as e:
        logger.warning(f"Could not format Excel file {file_path}: {e}")


def analyze_ring_tags(servers_df):
    """Analyze servers and mark those with invalid ring tags, completely ignoring Citrix rings."""
    servers_df['has_invalid_ring_tag'] = False
    servers_df['comment'] = ''

    for index, server in servers_df.iterrows():
        env = str(server.get('SNOW_environment', '')).lower()
        current_tags = server.get('current_tags', '')

        # Extract all ring tags first
        all_ring_tags = re.findall(r'FalconGroupingTags/[^,]*?SRVRing(\d+)', current_tags, re.IGNORECASE)

        # Find complete ring tag strings to check for Citrix
        complete_ring_tag_matches = re.findall(r'(FalconGroupingTags/[^,]*?SRVRing\d+)', current_tags, re.IGNORECASE)

        # Filter to only non-Citrix rings by checking the complete tag
        non_citrix_rings = []
        for i, ring_num in enumerate(all_ring_tags):
            if 'citrix' not in complete_ring_tag_matches[i].lower():
                non_citrix_rings.append(int(ring_num))

        if not non_citrix_rings:
            continue

        if len(non_citrix_rings) > 1:
            servers_df.loc[index, 'has_invalid_ring_tag'] = True
            servers_df.loc[index, 'comment'] = 'multiple ring tags found'
            logger.info(f"Multiple non-Citrix ring tags found for host {server.get('hostname', 'Unknown')}: {non_citrix_rings}")
            continue

        expected_ring = get_expected_ring(env)
        if any(num != expected_ring for num in non_citrix_rings):
            servers_df.loc[index, 'has_invalid_ring_tag'] = True
            servers_df.loc[index, 'comment'] = f'{env} server should not be in Ring {non_citrix_rings}, expected Ring {expected_ring}'
            logger.info(f"Invalid ring tag found for host {server.get('hostname', 'Unknown')}: Ring {non_citrix_rings}, expected Ring {expected_ring}")


def generate_report():
    """Generate the complete invalid ring tag analysis report."""
    today_date = datetime.now(EASTERN_TZ).strftime('%m-%d-%Y')
    output_dir = DATA_DIR / today_date
    output_dir.mkdir(parents=True, exist_ok=True)

    # Get input file
    input_file_path = get_dated_path(DATA_DIR, "unique_cs_hosts.xlsx")
    if not input_file_path.exists():
        logger.info("Unique hosts file not found. Generating it now...")
        crowdstrike.update_unique_hosts_from_cs()

    # Read and filter data
    unique_cs_hosts_df = pd.read_excel(input_file_path, engine="openpyxl")
    logger.info(f"Read {len(unique_cs_hosts_df)} records from {input_file_path}")

    # Filter for servers only
    servers = unique_cs_hosts_df[
        unique_cs_hosts_df['cs_host_category'].str.lower().isin(['server', 'domain controller'])
    ]
    logger.info(f"Found {len(servers)} servers")

    # Filter servers with ring tags
    servers_with_ring_tags = servers[
        servers["current_tags"].str.contains("FalconGroupingTags/.*SrvRing", regex=True, case=False, na=False)
    ]
    logger.info(f"Found {len(servers_with_ring_tags)} servers with ring tags")

    # Save servers with ring tags
    servers_with_ring_tags_file_path = output_dir / "cs_servers_with_ring_tags.xlsx"
    servers_with_ring_tags.to_excel(servers_with_ring_tags_file_path, index=False, engine="openpyxl")
    adjust_column_widths(servers_with_ring_tags_file_path)

    # Enrich with ServiceNow data
    enriched_file_path = service_now.enrich_host_report(servers_with_ring_tags_file_path)
    enriched_servers = pd.read_excel(enriched_file_path, engine="openpyxl")

    # Analyze ring tags
    analyze_ring_tags(enriched_servers)

    # Save complete report
    complete_report_path = output_dir / "cs_servers_last_seen_with_invalid_ring_tags.xlsx"
    enriched_servers.to_excel(complete_report_path, index=False, engine="openpyxl")
    adjust_column_widths(complete_report_path)

    # Save filtered report (invalid tags only)
    invalid_servers = enriched_servers[enriched_servers['has_invalid_ring_tag']].copy()
    if not invalid_servers.empty:
        # Add invalid_tags column
        def extract_invalid_tags(row):
            current_tags = row.get('current_tags', '')
            # Extract all non-Citrix ring tags
            all_ring_tags = re.findall(r'FalconGroupingTags/[^,]*?SRVRing(\d+)', current_tags, re.IGNORECASE)
            complete_ring_tag_matches = re.findall(r'(FalconGroupingTags/[^,]*?SRVRing\d+)', current_tags, re.IGNORECASE)
            invalid = []
            for i, ring_num in enumerate(all_ring_tags):
                if 'citrix' not in complete_ring_tag_matches[i].lower():
                    invalid.append(complete_ring_tag_matches[i])
            return ', '.join(invalid) if invalid else ''

        invalid_servers['invalid_tags'] = invalid_servers.apply(extract_invalid_tags, axis=1)

        # Only keep the requested columns
        columns_to_keep = [
            'hostname',
            'host_id',
            'current_tags',
            'invalid_tags',
            'last_seen',
            'status',
            'cs_host_category',
            'SNOW_environment',
            'SNOW_lifecycleStatus',
            'comment',
        ]
        filtered_report_path = output_dir / "cs_servers_with_invalid_ring_tags_only.xlsx"
        invalid_servers[columns_to_keep].to_excel(filtered_report_path, index=False, engine="openpyxl")
        adjust_column_widths(filtered_report_path)
        logger.info(f"Found {len(invalid_servers)} hosts with invalid ring tags")
    else:
        logger.info("No hosts with invalid ring tags found")


def main():
    """Main function to run the complete workflow."""
    try:
        logger.info("Starting CrowdStrike host invalid ring tag analysis")
        generate_report()
        logger.info("Completed CrowdStrike host invalid ring tag analysis")
    except FileNotFoundError as e:
        logger.error(f"Input file not found: {e}")
        raise
    except Exception as e:
        logger.error(f"Error in main workflow: {e}")
        raise


if __name__ in ('__main__', '__builtin__', 'builtins'):
    main()
