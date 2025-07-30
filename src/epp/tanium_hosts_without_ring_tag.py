"""
Tanium Host Processing - Clean Architecture Refactor
Following SOLID principles and Clean Code practices
"""

import json
import logging
from collections import defaultdict
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Protocol

import openpyxl
import pandas as pd

from config import get_config
from services.service_now import enrich_host_report
from services.tanium import Computer, TaniumClient


# ============================================================================
# Domain Models (Clean, focused data structures)
# ============================================================================

@dataclass
class ProcessingConfig:
    """Configuration for ring processing"""
    ring_1_percent: float = 0.10
    ring_2_percent: float = 0.20
    ring_3_percent: float = 0.30
    ring_1_envs: set = None
    ring_2_envs: set = None
    ring_3_envs: set = None

    def __post_init__(self):
        if self.ring_1_envs is None:
            self.ring_1_envs = {'dev', 'development', 'sandbox', 'lab', 'poc', 'integration', 'int'}
        if self.ring_2_envs is None:
            self.ring_2_envs = {'test', 'testing', 'qa'}
        if self.ring_3_envs is None:
            self.ring_3_envs = {'stage', 'staging', 'uat', 'pre-prod', 'preprod', 'dr', 'qa/dr'}


@dataclass
class EnrichedComputer:
    """Computer with enrichment data - immutable where possible"""
    computer: Computer
    country: str = ""
    region: str = ""
    environment: str = ""
    category: str = ""
    was_country_guessed: bool = False
    status: str = ""
    ring_tag: Optional[str] = None

    @property
    def is_workstation(self) -> bool:
        return self.category.lower() == "workstation"

    @property
    def is_server(self) -> bool:
        return self.category.lower() in ("server", "srv")

    def add_status(self, message: str) -> 'EnrichedComputer':
        """Return new instance with appended status (immutable pattern)"""
        new_status = f"{self.status}; {message}" if self.status else message
        return replace(self, status=new_status)


# ============================================================================
# Protocols (Interface Segregation Principle)
# ============================================================================

class DataLoader(Protocol):
    """Interface for loading data"""

    def load_tanium_computers(self, test_limit: Optional[int] = None) -> List[Computer]:
        ...

    def load_country_mappings(self) -> Dict[str, str]:
        ...

    def load_region_mappings(self) -> Dict[str, str]:
        ...


class ComputerEnricher(Protocol):
    """Interface for enriching computer data"""

    def enrich_computers(self, computers: List[Computer]) -> List[EnrichedComputer]:
        ...


class CountryResolver(Protocol):
    """Interface for resolving countries"""

    def resolve_country(self, computer: Computer, snow_country: str) -> Tuple[str, bool]:
        ...


class RegionResolver(Protocol):
    """Interface for resolving regions"""

    def resolve_region(self, country: str) -> str:
        ...


class RingTagGenerator(Protocol):
    """Interface for generating ring tags"""

    def generate_tags(self, computers: List[EnrichedComputer]) -> List[EnrichedComputer]:
        ...


class ReportExporter(Protocol):
    """Interface for exporting reports"""

    def export_to_excel(self, computers: List[EnrichedComputer], output_path: Path) -> str:
        ...


# ============================================================================
# Concrete Implementations (Single Responsibility Principle)
# ============================================================================

class TaniumDataLoader:
    """Loads data from various sources"""

    def __init__(self, data_dir: Path):
        self.data_dir = data_dir
        self.logger = logging.getLogger(__name__)

    def load_tanium_computers(self, test_limit: Optional[int] = None) -> List[Computer]:
        """Load computers from Tanium, handling all the Excel parsing complexity"""
        today = datetime.now().strftime('%m-%d-%Y')
        output_dir = self.data_dir / "transient" / "epp_device_tagging" / today
        output_dir.mkdir(parents=True, exist_ok=True)
        all_hosts_file = output_dir / "All Tanium Hosts.xlsx"

        client = TaniumClient()

        if all_hosts_file.exists():
            all_hosts_filename = str(all_hosts_file)
            self.logger.info(f"Using existing hosts file: {all_hosts_filename}")
        else:
            all_hosts_filename = client.get_and_export_all_computers()

        if not all_hosts_filename:
            raise ValueError("No computers retrieved from any instance!")

        computers = self._parse_excel_file(all_hosts_filename)

        # Filter and limit
        filtered_computers = [c for c in computers if not c.has_epp_ring_tag()]
        if test_limit is not None and test_limit > 0:
            filtered_computers = filtered_computers[:test_limit]

        self.logger.info(f"Loaded {len(filtered_computers)} computers without ring tags")
        return filtered_computers

    def _parse_excel_file(self, filename: str) -> List[Computer]:
        """Parse Excel file into Computer objects"""
        self._validate_input_file(filename)
        computers = []
        wb = None

        try:
            wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
            ws = wb.active

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 6:
                    continue

                if not row[0]:  # name is required
                    continue

                try:
                    computers.append(
                        Computer(
                            name=str(row[0]).strip(),
                            id=str(row[1]).strip() if row[1] else "",
                            ip=str(row[2]).strip() if row[2] else "",
                            eidLastSeen=row[3],
                            source=str(row[4]).strip() if row[4] else "",
                            custom_tags=[tag.strip() for tag in str(row[5]).split(',') if tag.strip()] if row[5] else []
                        )
                    )
                except Exception as e:
                    self.logger.warning(f"Error processing row {row_num}: {e}")
        except Exception as e:
            self.logger.error(f"Error loading Excel file: {e}")
        finally:
            if wb is not None:
                wb.close()

        return computers

    @staticmethod
    def _validate_input_file(filepath: str):
        """Validate input file before processing"""
        file_path = Path(filepath)
        if not file_path.exists():
            raise FileNotFoundError(f"Input file not found: {filepath}")
        if not str(filepath).lower().endswith(('.xlsx', '.xls')):
            raise ValueError("Input file must be an Excel file")

    def load_country_mappings(self) -> Dict[str, str]:
        """Load country code to name mappings"""
        return self._load_json_file(self.data_dir / "countries_by_code.json")

    def load_region_mappings(self) -> Dict[str, str]:
        """Load country to region mappings"""
        return self._load_json_file(self.data_dir / "regions_by_country.json")

    def _load_json_file(self, filepath: Path) -> Dict[str, str]:
        """Generic JSON file loader with error handling"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.logger.info(f"Loaded {len(data)} entries from {filepath}")
            return data
        except FileNotFoundError:
            self.logger.error(f"File not found: {filepath}")
            return {}
        except json.JSONDecodeError as e:
            self.logger.error(f"Invalid JSON in {filepath}: {e}")
            return {}


class ServiceNowComputerEnricher:
    """Enriches computers with ServiceNow data"""

    def __init__(self, temp_dir: Path):
        self.temp_dir = temp_dir
        self.logger = logging.getLogger(__name__)

    def enrich_computers(self, computers: List[Computer]) -> List[EnrichedComputer]:
        """Enrich computers with ServiceNow data"""
        # Export computers to Excel for ServiceNow enrichment
        client = TaniumClient()
        temp_filename = self.temp_dir / "temp_computers_for_enrichment.xlsx"
        computers_file = client.export_to_excel(computers, str(temp_filename))

        # Enrich with ServiceNow
        enriched_file = enrich_host_report(computers_file)

        # Parse enriched data back
        df = pd.read_excel(enriched_file, dtype=str, engine='openpyxl', keep_default_na=False, na_values=[''])

        enriched_computers = []
        for _, row in df.iterrows():
            # Find original computer
            tanium_id = str(row.get('ID', '')).strip()
            original_computer = next((c for c in computers if c.id == tanium_id), None)

            if original_computer:
                enriched_computers.append(
                    EnrichedComputer(
                        computer=original_computer,
                        country="",
                        region="",
                        environment=self._clean_value(row.get('SNOW_environment', '')),
                        category=self._clean_value(row.get('SNOW_category', '')),
                        was_country_guessed=False,
                        status=self._clean_value(row.get('SNOW_comments', '')),
                        ring_tag=None
                    )
                )

        return enriched_computers

    @staticmethod
    def _clean_value(value) -> str:
        """Clean nan/null values"""
        if not value or str(value).lower() in ['nan', 'none', 'null']:
            return ""
        return str(value).strip()


class SmartCountryResolver:
    """Resolves countries using multiple strategies"""

    def __init__(self, country_mappings: Dict[str, str], config):
        self.country_mappings = country_mappings
        self.config = config
        self.logger = logging.getLogger(__name__)

    def resolve_country(self, computer: Computer, snow_country: str) -> Tuple[str, bool]:
        """Resolve country with fallback strategies"""
        # Priority 1: Valid ServiceNow country
        if self._is_valid_country(snow_country):
            return snow_country, False

        # Priority 2: Guess from hostname
        guessed_country, _ = self._guess_country_from_hostname(computer)
        if self._is_valid_country(guessed_country):
            return guessed_country, True

        return "", False

    @staticmethod
    def _is_valid_country(country: str) -> bool:
        """Check if country value is valid"""
        if not country or pd.isna(country):
            return False
        country_str = str(country).strip()
        return country_str and country_str.lower() not in ['nan', 'none', 'null', '']

    def _guess_country_from_hostname(self, computer: Computer) -> Tuple[str, str]:
        """Guess country from hostname using various strategies"""
        name = computer.name.strip().lower()
        if not name:
            return '', 'Empty hostname'

        # Strategy 1: Special prefixes
        if name.startswith('vmvdi'):
            return 'United States', "VMVDI prefix"

        if hasattr(self.config, 'team_name') and name.startswith(self.config.team_name.lower()):
            return 'United States', f"{self.config.team_name} prefix"

        # Strategy 2: Country code from first 2 characters
        if len(name) >= 2:
            country_code = name[:2].upper()
            country_name = self.country_mappings.get(country_code, '')
            if country_name:
                return country_name, f"Country code {country_code}"

        # Strategy 3: Leading digit suggests Korea
        if name[0].isdigit():
            return 'Korea', "Leading digit"

        # Strategy 4: Check tags for country indicators
        for tag in getattr(computer, 'custom_tags', []):
            tag_upper = str(tag).upper()
            for code, country_name in self.country_mappings.items():
                if code in tag_upper or country_name.upper() in tag_upper:
                    return country_name, f"Tag: {tag}"

        return '', 'No indicators found'


class SafeRegionResolver:
    """Resolves regions with proper error handling"""

    def __init__(self, region_mappings: Dict[str, str]):
        self.region_mappings = region_mappings
        self.logger = logging.getLogger(__name__)

    def resolve_region(self, country: str) -> str:
        """Resolve region from country with proper error handling"""
        if not country:
            return ""

        region = self.region_mappings.get(country, "")
        if not region or region is None or pd.isna(region):
            self.logger.warning(f"No region mapping found for country: {country}")
            return ""

        return str(region).strip()


class SmartRingTagGenerator:
    """Generates ring tags following business rules"""

    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)

    def generate_tags(self, computers: List[EnrichedComputer]) -> List[EnrichedComputer]:
        """Generate ring tags for all computers"""
        workstations = [c for c in computers if c.is_workstation]
        servers = [c for c in computers if c.is_server]
        other = [c for c in computers if not c.is_workstation and not c.is_server]

        # Process each category
        tagged_workstations = self._process_workstations(workstations)
        tagged_servers = self._process_servers(servers)
        tagged_other = [c.add_status("Category missing or unknown - skipping") for c in other]

        return tagged_workstations + tagged_servers + tagged_other

    def _process_workstations(self, workstations: List[EnrichedComputer]) -> List[EnrichedComputer]:
        """Process workstations by region/country groups"""
        if not workstations:
            return []

        # Group by region and country
        groups = defaultdict(list)
        for ws in workstations:
            if ws.region:  # Only process if region exists
                groups[(ws.region, ws.country)].append(ws)

        tagged_computers = []

        # Process each group
        for (region, country), group in groups.items():
            ring_sizes = self._calculate_ring_sizes(len(group))

            # Sort by last seen date
            sorted_group = sorted(group, key=lambda c: (c.computer.eidLastSeen is None, c.computer.eidLastSeen))

            # Assign ring tags
            current_index = 0
            for ring, size in enumerate(ring_sizes, start=1):
                for _ in range(size):
                    if current_index < len(sorted_group):
                        computer = sorted_group[current_index]
                        tagged_computer = EnrichedComputer(
                            computer=computer.computer,
                            country=computer.country,
                            region=computer.region,
                            environment=computer.environment,
                            category=computer.category,
                            was_country_guessed=computer.was_country_guessed,
                            status=computer.status,
                            ring_tag=f"EPP_ECMTag_{region}_Wks_Ring{ring}"
                        ).add_status("Ring tag generated successfully")

                        if computer.was_country_guessed:
                            tagged_computer = tagged_computer.add_status("Country was guessed")

                        tagged_computers.append(tagged_computer)
                        current_index += 1

        # Add computers without regions
        for ws in workstations:
            if not ws.region:
                tagged_computers.append(ws.add_status("Region missing. Ring tag couldn't be generated"))

        return tagged_computers

    def _process_servers(self, servers: List[EnrichedComputer]) -> List[EnrichedComputer]:
        """Process servers based on environment"""
        tagged_servers = []

        for server in servers:
            if not server.region:
                tagged_servers.append(server.add_status("Region missing. Ring tag couldn't be generated"))
                continue

            env = self._normalize_environment(server.environment)

            if env in self.config.ring_1_envs:
                ring = 1
            elif env in self.config.ring_2_envs:
                ring = 2
            elif env in self.config.ring_3_envs:
                ring = 3
            else:
                ring = 4

            tagged_server = EnrichedComputer(
                computer=server.computer,
                country=server.country,
                region=server.region,
                environment=server.environment,
                category=server.category,
                was_country_guessed=server.was_country_guessed,
                status=server.status,
                ring_tag=f"EPP_ECMTag_{server.region}_SRV_Ring{ring}"
            ).add_status("Ring tag generated successfully")

            if server.was_country_guessed:
                tagged_server = tagged_server.add_status("Country was guessed")

            tagged_servers.append(tagged_server)

        return tagged_servers

    def _calculate_ring_sizes(self, total: int) -> List[int]:
        """Calculate ring distribution sizes"""
        if total == 1:
            return [0, 0, 0, 1]
        elif total == 2:
            return [0, 0, 1, 1]
        elif total <= 5:
            return [0, 1, 1, total - 2]
        else:
            ring_1 = max(1, int(self.config.ring_1_percent * total))
            ring_2 = max(1, int(self.config.ring_2_percent * total))
            ring_3 = max(1, int(self.config.ring_3_percent * total))
            ring_4 = total - ring_1 - ring_2 - ring_3
            return [ring_1, ring_2, ring_3, ring_4]

    @staticmethod
    def _normalize_environment(environment: str) -> str:
        """Normalize environment string"""
        if not environment:
            return ""
        return str(environment).lower().strip()


class ExcelReportExporter:
    """Exports enriched computers to Excel reports"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def export_to_excel(self, computers: List[EnrichedComputer], output_path: Path) -> str:
        """Export computers to Excel with proper formatting"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ring Assignments"

        # Headers
        headers = [
            "Computer Name", "Tanium ID", "Category", "Environment",
            "Country", "Region", "Was Country Guessed", "Current Tags",
            "Generated Tag", "Comments"
        ]
        ws.append(headers)

        # Make headers bold
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14)

        # Add data
        for computer in computers:
            current_tags = ""
            if hasattr(computer.computer, "custom_tags") and computer.computer.custom_tags:
                filtered_tags = [tag for tag in computer.computer.custom_tags
                                 if tag and str(tag).lower() != "nan"]
                current_tags = ", ".join(filtered_tags)

            category_display = "Workstation" if computer.is_workstation else "Server" if computer.is_server else computer.category

            ws.append([
                computer.computer.name,
                computer.computer.id,
                category_display,
                computer.environment,
                computer.country,
                computer.region,
                "Yes" if computer.was_country_guessed else "No",
                current_tags,
                computer.ring_tag or "",
                computer.status
            ])

        # Format columns
        column_widths = {
            'A': 40, 'B': 25, 'C': 18, 'D': 22, 'E': 22,
            'F': 18, 'G': 14, 'H': 50, 'I': 28, 'J': 80
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add filters and freeze panes
        ws.auto_filter.ref = f"A1:J{len(computers) + 1}"
        ws.freeze_panes = "A2"

        # Save file
        wb.save(output_path)

        self.logger.info(f"Exported {len(computers)} computers to {output_path}")
        return str(output_path)


# ============================================================================
# Main Orchestrator (Dependency Inversion Principle)
# ============================================================================

class TaniumRingTagProcessor:
    """Main orchestrator following Clean Architecture principles"""

    def __init__(self,
                 data_loader: DataLoader,
                 enricher: ComputerEnricher,
                 country_resolver: CountryResolver,
                 region_resolver: RegionResolver,
                 tag_generator: RingTagGenerator,
                 report_exporter: ReportExporter):
        self.data_loader = data_loader
        self.enricher = enricher
        self.country_resolver = country_resolver
        self.region_resolver = region_resolver
        self.tag_generator = tag_generator
        self.report_exporter = report_exporter
        self.logger = logging.getLogger(__name__)

    def process_hosts_without_ring_tags(self, test_limit: Optional[int] = None) -> str:
        """Main processing pipeline - clean and readable"""
        try:
            # Step 1: Load raw data
            self.logger.info("Loading Tanium computers...")
            computers = self.data_loader.load_tanium_computers(test_limit)

            # Step 2: Enrich with ServiceNow data
            self.logger.info("Enriching with ServiceNow data...")
            enriched_computers = self.enricher.enrich_computers(computers)

            # Step 3: Resolve countries and regions
            self.logger.info("Resolving countries and regions...")
            final_computers = []
            for comp in enriched_computers:
                # Resolve country
                country, was_guessed = self.country_resolver.resolve_country(
                    comp.computer,
                    comp.country  # This would be from SNOW
                )

                # Resolve region
                region = self.region_resolver.resolve_region(country) if country else ""

                # Create final enriched computer
                final_comp = EnrichedComputer(
                    computer=comp.computer,
                    country=country,
                    region=region,
                    environment=comp.environment,
                    category=comp.category,
                    was_country_guessed=was_guessed,
                    status=comp.status,
                    ring_tag=None
                )
                final_computers.append(final_comp)

            # Step 4: Generate ring tags
            self.logger.info("Generating ring tags...")
            tagged_computers = self.tag_generator.generate_tags(final_computers)

            # Step 5: Export report
            self.logger.info("Exporting final report...")
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            output_path = Path(f"Tanium_Ring_Tags_Report_{timestamp}.xlsx")

            report_path = self.report_exporter.export_to_excel(tagged_computers, output_path)

            # Log summary
            generated_count = sum(1 for c in tagged_computers if c.ring_tag)
            self.logger.info(f"Generated ring tags for {generated_count}/{len(tagged_computers)} computers")

            return report_path

        except Exception as e:
            self.logger.error(f"Processing failed: {e}", exc_info=True)
            raise


# ============================================================================
# Factory/Builder (Dependency Injection)
# ============================================================================

def create_processor() -> TaniumRingTagProcessor:
    """Factory method to create fully configured processor"""
    config = get_config()
    data_dir = Path(__file__).parent.parent.parent / "data"
    temp_dir = data_dir / "transient" / "epp_device_tagging" / datetime.now().strftime('%m-%d-%Y')
    temp_dir.mkdir(parents=True, exist_ok=True)

    # Create all dependencies
    data_loader = TaniumDataLoader(data_dir)
    country_mappings = data_loader.load_country_mappings()
    region_mappings = data_loader.load_region_mappings()

    enricher = ServiceNowComputerEnricher(temp_dir)
    country_resolver = SmartCountryResolver(country_mappings, config)
    region_resolver = SafeRegionResolver(region_mappings)
    tag_generator = SmartRingTagGenerator(ProcessingConfig())
    report_exporter = ExcelReportExporter()

    return TaniumRingTagProcessor(
        data_loader=data_loader,
        enricher=enricher,
        country_resolver=country_resolver,
        region_resolver=region_resolver,
        tag_generator=tag_generator,
        report_exporter=report_exporter
    )


# ============================================================================
# Main Entry Point
# ============================================================================

def main():
    """Clean main entry point"""
    logging.basicConfig(level=logging.INFO)

    try:
        processor = create_processor()
        result = processor.process_hosts_without_ring_tags(test_limit=10)
        print(f"Processing completed successfully: {result}")
    except Exception as e:
        logging.error(f"Processing failed: {e}")
        raise


if __name__ == "__main__":
    main()
