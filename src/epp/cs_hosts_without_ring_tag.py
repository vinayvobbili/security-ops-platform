import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Callable, TypeVar

import pandas as pd
from webexteamssdk import WebexTeamsAPI

from my_config import get_config
from services import crowdstrike
from services.service_now import ServiceNowClient

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Type variable for decorator
F = TypeVar('F', bound=Callable[..., Any])

# Constants - consolidated
ROOT_DIRECTORY = Path(__file__).parent.parent.parent
DATA_DIR = ROOT_DIRECTORY / "data" / "transient" / "epp_device_tagging"

# Configuration
CONFIG = get_config()

DEFAULT_CHUNK_SIZE = 500


def get_dated_path(base_dir: Path, filename: str) -> Path:
    """
    Create a path with today's date directory.

    Args:
        base_dir: Base directory path
        filename: Name of the file

    Returns:
        Path with date directory
    """
    today_date = datetime.now().strftime('%m-%d-%Y')
    return base_dir / today_date / filename


def read_excel_file(file_path: Path) -> pd.DataFrame:
    """Read data from an Excel file."""
    try:
        return pd.read_excel(file_path, engine="openpyxl")
    except FileNotFoundError:
        logger.error(f"File not found: {file_path}")
        raise
    except PermissionError:
        logger.error(f"Permission denied for file: {file_path}")
        raise
    except Exception as e:
        logger.error(f"Error reading file {file_path}: {e}")
        raise


def write_excel_file(df: pd.DataFrame, file_path: Path) -> None:
    """Write DataFrame to an Excel file with professional formatting."""
    try:
        # Ensure directory exists
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Hosts Without Ring Tags")

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets["Hosts Without Ring Tags"]

            # Format the header row
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter

            # Bold the header row
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            # Freeze the top row
            worksheet.freeze_panes = "A2"

            # Add autofilter
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                # Set a reasonable minimum and maximum width
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Successfully wrote {len(df)} records to {file_path} with formatting")
    except PermissionError:
        logger.error(f"Permission denied when writing to {file_path}")
        raise
    except OSError as e:
        logger.error(f"OS error when writing to {file_path}: {e}")
        raise
    except Exception as e:
        logger.error(f"Error writing to {file_path}: {e}")
        raise


def process_unique_hosts(df: pd.DataFrame) -> pd.DataFrame:
    """
    Process dataframe to get unique hosts with latest last_seen.

    Args:
        df: DataFrame with host data

    Returns:
        DataFrame with unique hosts (latest entry per hostname)
    """
    # Convert last_seen to datetime for proper sorting - only once
    df["last_seen"] = pd.to_datetime(df["last_seen"], errors='coerce').dt.tz_localize(None)

    # Group by hostname and get the record with the latest last_seen
    return df.loc[df.groupby("hostname")["last_seen"].idxmax()]


def filter_hosts_without_ring_tag(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filter hosts that don't have a Ring tag.

    Args:
        df: DataFrame with host data including tags

    Returns:
        DataFrame with filtered hosts
    """
    # Convert tags to strings and handle NaN values
    df["current_tags"] = df["current_tags"].astype(str).replace('nan', '')

    # Vectorized filtering for hosts without ring tags
    has_ring_tag = df["current_tags"].str.contains("FalconGroupingTags/.*Ring", regex=True, case=False, na=False)
    return df[~has_ring_tag]


def list_cs_hosts_without_ring_tag() -> None:
    """List CrowdStrike hosts that don't have a FalconGroupingTags/*Ring* tag."""
    input_file = get_dated_path(DATA_DIR, "unique_cs_hosts.xlsx")
    if not input_file.exists():
        logger.info("Unique hosts file not found. Generating it now...")
        crowdstrike.update_unique_hosts_from_cs()
    try:
        df = read_excel_file(input_file)

        # Filter hosts without ring tags
        output_df = filter_hosts_without_ring_tag(df)

        hosts_without_tag_file = get_dated_path(DATA_DIR, "cs_hosts_last_seen_without_ring_tag.xlsx")
        write_excel_file(output_df, hosts_without_tag_file)
        logger.info(f"Found {len(output_df)} hosts without a Ring tag.")
    except FileNotFoundError:
        logger.error(f"Input file not found at {input_file}")
        raise
    except Exception as e:
        logger.error(f"Error listing hosts without ring tag: {e}")
        raise


def fetch_host_details(hostname: str, service_now_client: ServiceNowClient) -> Dict[str, Any]:
    """
    Fetch host details from ServiceNow.

    Args:
        hostname: The hostname to look up
        service_now_client: Initialized ServiceNowClient

    Returns:
        Dictionary with host details
    """
    try:
        return service_now_client.get_host_details(hostname)
    except ConnectionError as e:
        logger.error(f"Connection error for {hostname}: {e}")
        return {"name": hostname, "error": "Connection error", "error_details": str(e)}
    except TimeoutError as e:
        logger.error(f"Timeout error for {hostname}: {e}")
        return {"name": hostname, "error": "Timeout error", "error_details": str(e)}
    except Exception as e:
        logger.error(f"Error fetching details for {hostname}: {e}")
        return {"name": hostname, "error": str(e)}


def main() -> None:
    """Main function to run the complete workflow."""
    try:
        logger.info("Starting CrowdStrike host ring tag analysis")
        generate_report()
        logger.info("Completed CrowdStrike host ring tag analysis")
    except Exception as e:
        logger.error(f"Error in main workflow: {e}")


def generate_report():
    """
    Run the complete workflow without profiling.

    Args:
    """
    list_cs_hosts_without_ring_tag()


if __name__ in ('__main__', '__builtin__', 'builtins'):
    main()
