import urllib3
from urllib3.exceptions import InsecureRequestWarning

urllib3.disable_warnings(InsecureRequestWarning)

import concurrent.futures
import logging
import sys
import time
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Dict, Optional, Any, List

import pandas as pd
import requests
from rich.progress import track

from falconpy import (
    Hosts, OAuth2, Incidents, Alerts, IOC, Iocs, Intel,
    SpotlightVulnerabilities, Quarantine, IdentityProtection,
)
from my_config import get_config
from src.utils.http_utils import get_session

# Setup logger
logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent.parent / "data" / "transient" / "epp_device_tagging"
CS_FETCH_MAX_WORKERS = 10

# A 401/403 means the credentials themselves are bad — retrying won't help and
# risks account lockout. Every other failure (429, 5xx, connection/timeout) is
# treated as transient and is safe to retry with backoff.
CS_AUTH_PERMANENT_FAILURE_CODES = {401, 403}

# Get robust HTTP session instance
http_session = get_session()


class CSCredentialProfile(Enum):
    """CrowdStrike API credential profiles for different permission levels."""
    READ = "read"
    WRITE = "write"
    RTR = "rtr"


class CrowdStrikeAPIError(Exception):
    """Raised when a CrowdStrike API call returns a non-200 status.

    Distinct from a successful query that simply matched no device — callers
    that opt in can tell a transient API/auth failure apart from a host that
    genuinely isn't in CrowdStrike (both otherwise collapse to a None result).
    """


class CrowdStrikeClient:
    """Client for interacting with the CrowdStrike Falcon API."""

    def __init__(self, credential_profile: CSCredentialProfile = CSCredentialProfile.READ, max_workers: Optional[int] = None):
        self.config = get_config()
        self.base_url = "api.us-2.crowdstrike.com"
        self.proxies = self._setup_proxy()
        self.last_error: str | None = None  # Stores last API/auth error for better error reporting
        self.last_status_code: Optional[int] = None  # HTTP status from the most recent validate_auth() (None on network/timeout)
        if self.proxies:
            logger.info(f"[CrowdStrikeClient] Proxy enabled: {self.proxies}")
        else:
            logger.info("[CrowdStrikeClient] Proxy not enabled.")
        self.credential_profile = credential_profile
        self.auth = self._create_auth()
        self.hosts_client = Hosts(auth_object=self.auth, timeout=30)
        self.alerts_client = Alerts(auth_object=self.auth, timeout=30)
        self.incidents_client = Incidents(auth_object=self.auth, timeout=30)
        # Allow thread pool size to be set via env or parameter
        self.max_workers = max_workers or CS_FETCH_MAX_WORKERS

    def _get_client_id_secret(self):
        match self.credential_profile:
            case CSCredentialProfile.READ:
                return self.config.cs_ro_client_id, self.config.cs_ro_client_secret
            case CSCredentialProfile.WRITE:
                return self.config.cs_host_write_client_id, self.config.cs_host_write_client_secret
            case CSCredentialProfile.RTR:
                return self.config.cs_rtr_client_id, self.config.cs_rtr_client_secret

    def _setup_proxy(self):
        """Setup proxy configuration if jump server is enabled"""
        if not self.config.should_use_jump_server:
            return None

        proxy_url = f"http://{self.config.jump_server_host}:8081"
        return {"http": proxy_url, "https": proxy_url}

    def _create_auth(self):
        client_id, client_secret = self._get_client_id_secret()
        """Create OAuth2 authentication object"""
        return OAuth2(
            client_id=client_id,
            client_secret=client_secret,
            base_url=self.base_url,
            ssl_verify=False,
            proxy=self.proxies,
            timeout=30  # 30 second timeout to prevent indefinite hangs
        )

    def get_access_token(self) -> str:
        """Get CrowdStrike access token using direct API call"""
        url = f'https://{self.base_url}/oauth2/token'
        client_id, client_secret = self._get_client_id_secret()
        body = {
            'client_id': client_id,
            'client_secret': client_secret
        }

        response = http_session.post(url, data=body, verify=False, proxies=self.proxies)
        if response is None:
            raise requests.exceptions.ConnectionError("Failed to connect after multiple retries")
        response.raise_for_status()
        return response.json()['access_token']

    def validate_auth(self) -> bool:
        """Validate that CrowdStrike API authentication is working.

        Makes a simple API call to verify credentials are valid.
        Stores any error in self.last_error for better error reporting.

        Returns:
            True if authentication is valid, False otherwise.
        """
        try:
            # Use a minimal query to validate auth - limit=1 for efficiency
            response = self.hosts_client.query_devices_by_filter_scroll(limit=1)
            self.last_status_code = response.get("status_code")

            if self.last_status_code == 200:
                self.last_error = None
                logger.info("CrowdStrike API authentication validated successfully")
                return True

            # Extract error details from response
            status_code = response.get("status_code", "Unknown")
            errors = response.get("body", {}).get("errors", [])
            if errors:
                error_msg = "; ".join(e.get("message", str(e)) for e in errors)
                self.last_error = f"HTTP {status_code} - {error_msg}"
            else:
                self.last_error = f"HTTP {status_code} - {response.get('body', {})}"

            logger.warning(f"CrowdStrike API authentication failed: {self.last_error}")
            return False

        except Exception as e:
            self.last_status_code = None  # network/timeout — no HTTP status; treat as transient
            self.last_error = str(e)
            logger.warning(f"CrowdStrike API authentication failed: {self.last_error}")
            return False

    def validate_auth_with_retry(self, max_attempts: int = 3, backoff_base: float = 2.0) -> bool:
        """Validate auth, retrying transient failures with exponential backoff.

        A 401/403 is a permanent credential problem and is not retried (it won't
        recover and repeated attempts risk account lockout). Transient failures
        (429, 5xx, connection/timeout) are retried up to max_attempts with
        exponential backoff. Returns True only once auth succeeds.
        """
        for attempt in range(1, max_attempts + 1):
            if self.validate_auth():
                return True

            if self.last_status_code in CS_AUTH_PERMANENT_FAILURE_CODES:
                logger.error(
                    f"CrowdStrike auth failed with permanent error HTTP {self.last_status_code}; "
                    f"not retrying: {self.last_error}"
                )
                return False

            if attempt < max_attempts:
                backoff = backoff_base ** attempt
                logger.warning(
                    f"CrowdStrike auth attempt {attempt}/{max_attempts} failed ({self.last_error}); "
                    f"retrying in {backoff:.1f}s"
                )
                time.sleep(backoff)

        logger.error(f"CrowdStrike auth failed after {max_attempts} attempts: {self.last_error}")
        return False

    def get_device_ids_batch(self, hostnames, batch_size=100):
        """Get device IDs for multiple hostnames in batches"""
        results = {}
        for i in range(0, len(hostnames), batch_size):
            batch = hostnames[i:i + batch_size]
            host_filter = f"hostname:['{'', ''.join(batch)}']"

            response = self.hosts_client.query_devices_by_filter(
                filter=host_filter,
                limit=len(batch)
            )

            if response.get("status_code") == 200:
                device_ids = response["body"].get("resources", [])
                if device_ids:
                    details = self.hosts_client.get_device_details(ids=device_ids)
                    if details.get("status_code") == 200:
                        for device in details["body"].get("resources", []):
                            hostname = device.get("hostname")
                            device_id = device.get("device_id")
                            if hostname and device_id:
                                results[hostname] = device_id

        return results

    def get_device_id(self, hostname: str, raise_on_error: bool = False) -> Optional[str]:
        """Retrieve the device ID for a given hostname.

        Returns None when the query succeeds but no device matches. When
        raise_on_error is True, a non-200 response raises CrowdStrikeAPIError
        instead of returning None, so callers can distinguish a transient
        API/auth failure from a host that genuinely isn't in CrowdStrike.
        """
        host_filter = f"hostname:'{hostname}'"
        response = self.hosts_client.query_devices_by_filter(
            filter=host_filter,
            sort='last_seen.desc',
            limit=1
        )

        status_code = response.get("status_code")
        if status_code == 200:
            devices = response["body"].get("resources", [])
            return devices[0] if devices else None

        if raise_on_error:
            errors = response.get("body", {}).get("errors", [])
            detail = "; ".join(e.get("message", str(e)) for e in errors) if errors else response.get("body", {})
            raise CrowdStrikeAPIError(f"query_devices_by_filter returned HTTP {status_code}: {detail}")

        return None

    def get_device_details(self, device_id: str) -> Dict[str, Any]:
        """Retrieve details for a specific device"""
        response = self.hosts_client.get_device_details_v2(ids=device_id)
        if response.get("status_code") == 200:
            resources = response["body"].get("resources", [])
            return resources[0] if resources else {}

        return {}

    def get_device_containment_status(self, hostname: str) -> Optional[str]:
        """Get containment status for a device using hostname"""
        device_id = self.get_device_id(hostname)
        if not device_id:
            return 'Host not found in CS console or an error occurred.'

        device_details = self.get_device_details(device_id)
        return device_details.get("status")

    def fetch_all_hosts_and_write_to_xlsx(self, xlsx_filename: str = "all_cs_hosts.xlsx") -> None:
        """Fetch all hosts from CrowdStrike Falcon and stream them directly to xlsx.

        Rows are appended to a write-only openpyxl workbook as each batch finishes,
        so only the active batch's dicts and the dedup set of device_ids live in
        memory. An earlier implementation accumulated every host in a list and
        materialized a single DataFrame at the end, which OOMed on lab-vm2.

        Raises:
            ConnectionError: If authentication fails or no hosts could be retrieved,
                           includes the actual error message from the API.
        """
        import logging
        logger = logging.getLogger(__name__)
        from openpyxl import Workbook

        # Validate authentication first
        if not self.validate_auth():
            raise ConnectionError(f"CrowdStrike API authentication failed: {self.last_error}")

        HEADERS = ["hostname", "host_id", "current_tags", "last_seen", "status", "cs_host_category", "chassis_type_desc"]

        today_date = datetime.now().strftime('%m-%d-%Y')
        output_path = DATA_DIR / today_date
        output_path.mkdir(parents=True, exist_ok=True)
        excel_file_path = output_path / xlsx_filename

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Hosts")
        ws.append(HEADERS)

        unique_device_ids: set = set()
        offset = None
        limit = 5000
        batch_count = 0
        start_time = time.time()
        api_error = None  # Track API errors during fetch
        total_written = 0

        def process_host_details(host_ids_batch: List[str]) -> List[Dict[str, Any]]:
            """Thread worker — return rows for this batch instead of mutating shared state."""
            details_response = self.hosts_client.get_device_details(ids=host_ids_batch)
            if details_response["status_code"] != 200:
                return []
            rows = []
            for host in details_response["body"].get("resources", []):
                device_id = host.get("device_id")
                if not device_id:
                    continue
                rows.append({
                    "hostname": host.get("hostname"),
                    "host_id": device_id,
                    "current_tags": ", ".join(host.get("tags", [])),
                    "last_seen": host.get("last_seen"),
                    "status": host.get("status"),
                    "cs_host_category": host.get("product_type_desc"),
                    "chassis_type_desc": host.get("chassis_type_desc"),
                })
            return rows

        logger.info(f"Starting fetch_all_hosts_and_write_to_xlsx (streaming) with max_workers={self.max_workers}")
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            while True:
                # Refresh auth token every 10 batches
                if batch_count > 0 and batch_count % 10 == 0:
                    self.auth = self._create_auth()
                    self.hosts_client = Hosts(auth_object=self.auth, timeout=30)

                response = self.hosts_client.query_devices_by_filter_scroll(
                    limit=limit, offset=offset
                )

                if response["status_code"] != 200:
                    # Capture the error details for better reporting
                    status_code = response.get("status_code", "Unknown")
                    errors = response.get("body", {}).get("errors", [])
                    if errors:
                        error_msg = "; ".join(e.get("message", str(e)) for e in errors)
                        api_error = f"HTTP {status_code} - {error_msg}"
                    else:
                        api_error = f"HTTP {status_code} - {response.get('body', {})}"
                    logger.error(f"CrowdStrike API error during host fetch: {api_error}")
                    self.last_error = api_error
                    break

                host_ids = response["body"].get("resources", [])
                if not host_ids:
                    break

                # Process in batches of 1000
                host_id_batches = [host_ids[i:i + 1000] for i in range(0, len(host_ids), 1000)]

                # Log for VM/non-interactive sessions
                logger.info(f"Processing batch {batch_count + 1}: {len(host_ids)} host IDs in {len(host_id_batches)} sub-batches")

                # Process with rich progress (shows progress bar locally, silent on VM).
                # Drain futures as they complete and write rows immediately so memory
                # never grows past one batch worth of host dicts.
                futures = [
                    executor.submit(process_host_details, id_batch)
                    for id_batch in track(host_id_batches, description=f"Batch {batch_count + 1}", disable=not sys.stdout.isatty())
                ]
                for future in concurrent.futures.as_completed(futures):
                    for row in future.result():
                        dev_id = row["host_id"]
                        if dev_id in unique_device_ids:
                            continue
                        unique_device_ids.add(dev_id)
                        ws.append([row[h] for h in HEADERS])
                        total_written += 1

                batch_count += 1
                logger.info(f"Completed batch {batch_count}, total hosts written so far: {total_written}")
                offset = response["body"].get("meta", {}).get("pagination", {}).get("offset")
                if not offset:
                    break

                time.sleep(0.5)

        elapsed = time.time() - start_time
        logger.info(f"Completed fetch_all_hosts_and_write_to_xlsx in {elapsed:.2f} seconds. Total hosts: {total_written}")

        # If no hosts were retrieved, and we had an API error, raise with details
        if total_written == 0 and api_error:
            raise ConnectionError(f"No hosts retrieved from CrowdStrike. {api_error}")

        wb.save(excel_file_path)
        # Intermediate file consumed by downstream code — skip apply_professional_formatting
        # since it would load the entire workbook back into memory and defeat streaming.

    def update_device_tags(self, action_name: str, ids: list, tags: list) -> dict:
        """Update device tags (add/remove) for a list of device IDs."""
        return self.hosts_client.update_device_tags(
            action_name=action_name,
            ids=ids,
            tags=tags
        )

    def get_device_online_state(self, hostname: str) -> Optional[str]:
        """Get the online state for a single hostname."""
        device_id = self.get_device_id(hostname)
        if not device_id:
            return None
        response = self.hosts_client.get_online_state(ids=[device_id])
        if response.get("status_code") == 200:
            resources = response['body'].get('resources', [])
            if resources:
                return resources[0].get('state')
        return None

    def get_detections(
        self,
        limit: int = 20,
        filter_query: Optional[str] = None,
        sort: str = "created_timestamp|desc"
    ) -> Dict[str, Any]:
        """Query CrowdStrike alerts/detections using the new Alerts API.

        Args:
            limit: Maximum number of alerts to return
            filter_query: FQL filter string (e.g., "status:'new'" or "device.hostname:'WORKSTATION01'")
            sort: Sort order (default: most recent first)

        Returns:
            Dict containing alert details or error info
        """
        try:
            # Query alert IDs using the new Alerts API
            query_params = {"limit": limit, "sort": sort}
            if filter_query:
                query_params["filter"] = filter_query

            response = self.alerts_client.query_alerts_v2(**query_params)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [{}])
                return {"error": f"Failed to query alerts: {error_msg}"}

            alert_ids = response.get("body", {}).get("resources", [])

            if not alert_ids:
                return {"results": [], "total": 0}

            # Get alert details using composite IDs
            details_response = self.alerts_client.get_alerts_v2(composite_ids=alert_ids)

            if details_response.get("status_code") != 200:
                return {"error": "Failed to retrieve alert details"}

            alerts = details_response.get("body", {}).get("resources", [])
            return {"results": alerts, "total": len(alerts)}

        except Exception as e:
            logger.error(f"Error querying CrowdStrike alerts: {e}")
            return {"error": str(e)}

    def get_detection_by_id(self, detection_id: str) -> Dict[str, Any]:
        """Get detailed information for a specific alert/detection.

        Args:
            detection_id: The alert composite ID

        Returns:
            Dict containing alert details or error info
        """
        try:
            response = self.alerts_client.get_alerts_v2(composite_ids=[detection_id])

            if response.get("status_code") != 200:
                return {"error": f"Failed to get alert {detection_id}"}

            resources = response.get("body", {}).get("resources", [])
            if not resources:
                return {"error": f"Alert {detection_id} not found"}

            return resources[0]

        except Exception as e:
            logger.error(f"Error getting CrowdStrike alert {detection_id}: {e}")
            return {"error": str(e)}

    def get_detections_by_hostname(self, hostname: str, limit: int = 20) -> Dict[str, Any]:
        """Get alerts/detections for a specific hostname.

        Args:
            hostname: The hostname to search for
            limit: Maximum number of alerts to return

        Returns:
            Dict containing alert details or error info
        """
        filter_query = f"device.hostname:'{hostname}'"
        return self.get_detections(limit=limit, filter_query=filter_query)

    def get_incidents(
        self,
        limit: int = 20,
        filter_query: Optional[str] = None,
        sort: str = "start|desc"
    ) -> Dict[str, Any]:
        """Query CrowdStrike incidents.

        Args:
            limit: Maximum number of incidents to return
            filter_query: FQL filter string (e.g., "status:'20'" for new incidents)
            sort: Sort order (default: most recent first)

        Returns:
            Dict containing incident details or error info
        """
        try:
            # Query incident IDs
            query_params = {"limit": limit, "sort": sort}
            if filter_query:
                query_params["filter"] = filter_query

            response = self.incidents_client.query_incidents(**query_params)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [{}])
                return {"error": f"Failed to query incidents: {error_msg}"}

            incident_ids = response.get("body", {}).get("resources", [])

            if not incident_ids:
                return {"results": [], "total": 0}

            # Get incident details
            details_response = self.incidents_client.get_incidents(ids=incident_ids)

            if details_response.get("status_code") != 200:
                return {"error": "Failed to retrieve incident details"}

            incidents = details_response.get("body", {}).get("resources", [])
            return {"results": incidents, "total": len(incidents)}

        except Exception as e:
            logger.error(f"Error querying CrowdStrike incidents: {e}")
            return {"error": str(e)}

    def get_incident_by_id(self, incident_id: str) -> Dict[str, Any]:
        """Get detailed information for a specific incident.

        Args:
            incident_id: The incident ID

        Returns:
            Dict containing incident details or error info
        """
        try:
            response = self.incidents_client.get_incidents(ids=[incident_id])

            if response.get("status_code") != 200:
                return {"error": f"Failed to get incident {incident_id}"}

            resources = response.get("body", {}).get("resources", [])
            if not resources:
                return {"error": f"Incident {incident_id} not found"}

            return resources[0]

        except Exception as e:
            logger.error(f"Error getting CrowdStrike incident {incident_id}: {e}")
            return {"error": str(e)}

    # ==================== IOC Search Methods ====================

    def search_ioc_by_value(self, ioc_value: str, ioc_type: str = None) -> Dict[str, Any]:
        """Search for a custom IOC in CrowdStrike.

        Args:
            ioc_value: The IOC value (IP, domain, hash, etc.)
            ioc_type: Optional IOC type filter (ipv4, domain, md5, sha256)

        Returns:
            Dict with IOC details or error
        """
        try:
            ioc_client = IOC(auth_object=self.auth, timeout=30)

            # Build filter
            filter_str = f"value:'{ioc_value}'"
            if ioc_type:
                filter_str += f"+type:'{ioc_type}'"

            response = ioc_client.indicator_combined_v1(filter=filter_str, limit=100)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"IOC search failed: {error_msg}"}

            resources = response.get("body", {}).get("resources", [])
            return {
                "count": len(resources),
                "indicators": resources
            }

        except Exception as e:
            logger.error(f"Error searching IOC {ioc_value}: {e}")
            return {"error": str(e)}

    @staticmethod
    def _refang(value: str) -> str:
        """Normalize a defanged indicator back to its live value.

        Analysts routinely paste defanged IOCs (``yowgames[.]com``,
        ``hxxps://...``, ``1[.]2[.]3[.]4``, ``user[at]evil[.]com``).
        CrowdStrike's IOC index matches the real value, so undo the common
        defang conventions before looking it up — otherwise a perfectly valid
        indicator silently resolves to zero hosts.
        """
        import re
        v = (value or "").strip().strip("\"'")
        # protocol obfuscation: hXXp(s) -> http(s)
        v = re.sub(r"h[xX]{2}p", "http", v)
        # bracketed/parenthesized/spelled-out separators -> real chars
        v = re.sub(r"[\[\(\{]\s*://\s*[\]\)\}]", "://", v)
        v = re.sub(r"[\[\(\{]\s*(?:\.|dot)\s*[\]\)\}]", ".", v, flags=re.IGNORECASE)
        v = re.sub(r"[\[\(\{]\s*(?::|colon)\s*[\]\)\}]", ":", v, flags=re.IGNORECASE)
        v = re.sub(r"[\[\(\{]\s*(?:@|at)\s*[\]\)\}]", "@", v, flags=re.IGNORECASE)
        v = re.sub(r"\s+dot\s+", ".", v, flags=re.IGNORECASE)
        # an IOC->host pivot wants the bare host: drop any leading scheme + trailing path
        v = re.sub(r"^[a-zA-Z][a-zA-Z0-9+.\-]*://", "", v)
        v = v.split("/", 1)[0]
        return v.strip().strip(".")

    @staticmethod
    def _detect_ioc_type(value: str) -> Optional[str]:
        """Infer the CrowdStrike IOC type from the raw indicator value.

        Returns one of the types `Iocs.devices_ran_on` accepts
        (sha256/sha1/md5/ipv4/ipv6/domain), or None if it can't tell.
        """
        import re
        v = (value or "").strip()
        if re.fullmatch(r"[a-fA-F0-9]{64}", v):
            return "sha256"
        if re.fullmatch(r"[a-fA-F0-9]{40}", v):
            return "sha1"
        if re.fullmatch(r"[a-fA-F0-9]{32}", v):
            return "md5"
        if re.fullmatch(r"\d{1,3}(\.\d{1,3}){3}", v):
            return "ipv4"
        if ":" in v and not v.endswith(":"):  # IPv6 (domains never contain ':')
            return "ipv6"
        if "." in v and re.search(r"[a-zA-Z]", v):
            return "domain"
        return None

    def _resolve_device_ids(self, device_ids: List[str]) -> List[Dict[str, Any]]:
        """Resolve CrowdStrike device IDs to hostname + key host facts.

        Tolerates partial resolution: IDs that no longer map to a managed host in
        this tenant are skipped. CrowdStrike returns HTTP 200 with the resolvable
        hosts in `resources` and a per-ID 404 in `errors` for the rest; when none
        resolve the whole call comes back 404 with an empty `resources`. Either
        way we parse whatever `resources` carries and return only real hosts —
        callers compare the resolved count against the requested count to detect a
        wholesale miss (e.g. globally-common hashes whose references aren't ours).
        """
        hosts: List[Dict[str, Any]] = []
        resp = self.hosts_client.get_device_details_v2(ids=device_ids[:100])
        status = resp.get("status_code")
        if status not in (200, 404):
            logger.warning(f"[_resolve_device_ids] get_device_details_v2 returned HTTP {status}: "
                           f"{resp.get('body', {}).get('errors')}")
        for d in resp.get("body", {}).get("resources", []) or []:
            hosts.append({
                "device_id": d.get("device_id"),
                "hostname": d.get("hostname"),
                "last_seen": d.get("last_seen"),
                "platform": d.get("platform_name"),
                "local_ip": d.get("local_ip"),
                "machine_domain": d.get("machine_domain"),
            })
        return hosts

    # ------------------------------------------------------------------ #
    # Spotlight — vulnerability management (host exposure + CVE pivot)    #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _summarize_vuln(v: Dict[str, Any]) -> Dict[str, Any]:
        """Trim a Spotlight vulnerability entity down to the fields analysts need.

        The raw entity carries large nested cve/host_info/remediation blocks;
        this keeps the host, CVE rating/score, exploit status, affected product
        and remediation actions — what a SOC analyst actually reads.
        """
        cve = v.get("cve", {}) or {}
        host = v.get("host_info", {}) or {}
        apps = v.get("apps", []) or []
        product = apps[0].get("product_name_version") if apps else None

        # Remediation entities arrive under the "remediation" facet as
        # {"ids": [...], "entities": [{"action": "..."}]}.
        rem = v.get("remediation", {}) or {}
        rem_entities = rem.get("entities", []) if isinstance(rem, dict) else []
        actions = [r.get("action") for r in rem_entities if r.get("action")]

        return {
            "vulnerability_id": v.get("id"),
            "status": v.get("status"),
            "cve_id": cve.get("id"),
            "cve_severity": cve.get("severity"),
            "exprt_rating": cve.get("exprt_rating"),          # CrowdStrike's ExPRT.AI rating
            "cvss_base_score": cve.get("base_score"),
            "exploit_status": cve.get("exploit_status"),
            "exploitability_score": cve.get("exploitability_score"),
            "hostname": host.get("hostname"),
            "local_ip": host.get("local_ip"),
            "platform": host.get("platform_name"),
            "os_version": host.get("os_version"),
            "product": product,
            "first_seen": v.get("created_timestamp"),
            "updated": v.get("updated_timestamp"),
            "remediations": actions,
        }

    def get_host_vulnerabilities(
        self, hostname: str, status: str = "open,reopen", limit: int = 100
    ) -> Dict[str, Any]:
        """List Spotlight vulnerabilities for a single host.

        Answers "what is exposed on this endpoint?" — the per-host view a SOC
        analyst needs when triaging a box. Returns vulnerabilities sorted by
        CVSS base score (most severe first).

        Args:
            hostname: The device hostname (Spotlight host_info.hostname).
            status: Comma-separated Spotlight statuses to include. Defaults to
                open vulns plus reopened ones; pass "" for every status.
            limit: Max vulnerabilities to return (Spotlight caps at 5000).

        Returns:
            Dict with 'hostname', 'count', 'vulnerabilities' (summarized) and the
            'fql_query' used, or 'error'.
        """
        hostname = (hostname or "").strip()
        if not hostname:
            return {"error": "Empty hostname"}

        clauses = [f"host_info.hostname:'{hostname}'"]
        if status:
            statuses = ",".join(f"'{s.strip()}'" for s in status.split(",") if s.strip())
            if statuses:
                clauses.append(f"status:[{statuses}]")
        filter_str = "+".join(clauses)

        try:
            spotlight = SpotlightVulnerabilities(auth_object=self.auth, timeout=30)
            resp = spotlight.query_vulnerabilities_combined(
                filter=filter_str,
                limit=min(max(1, limit), 5000),
                sort="updated_timestamp|desc",
                facet=["cve", "host_info", "remediation"],
            )
            if resp.get("status_code") != 200:
                errors = resp.get("body", {}).get("errors", [])
                return {"error": f"Spotlight query failed: {errors}", "fql_query": filter_str}

            resources = resp.get("body", {}).get("resources", []) or []
            vulns = [self._summarize_vuln(v) for v in resources]
            # Most severe first — Spotlight can't sort by score server-side.
            vulns.sort(key=lambda x: (x.get("cvss_base_score") or 0), reverse=True)
            return {
                "hostname": hostname,
                "count": len(vulns),
                "vulnerabilities": vulns,
                "fql_query": filter_str,
            }
        except Exception as e:
            logger.error(f"Error in get_host_vulnerabilities({hostname}): {e}")
            return {"error": str(e), "fql_query": filter_str}

    def search_vulnerabilities_by_cve(
        self, cve_id: str, status: str = "open,reopen", limit: int = 500
    ) -> Dict[str, Any]:
        """Find which hosts are exposed to a given CVE (the CVE -> hosts pivot).

        Answers "are we vulnerable to CVE-XXXX, and on which boxes?" — the
        exposure question for advisory/vuln triage. Complements the IOC->hosts
        pivot but for vulnerabilities rather than indicators.

        Args:
            cve_id: The CVE identifier, e.g. "CVE-2024-3094". Case-insensitive.
            status: Comma-separated Spotlight statuses to include (default open
                + reopen). Pass "" for every status.
            limit: Max vulnerability records to return.

        Returns:
            Dict with 'cve_id', 'host_count', 'hosts' (each summarized vuln) and
            the 'fql_query' used, or 'error'.
        """
        cve_id = (cve_id or "").strip().upper()
        if not cve_id:
            return {"error": "Empty CVE id"}

        clauses = [f"cve.id:'{cve_id}'"]
        if status:
            statuses = ",".join(f"'{s.strip()}'" for s in status.split(",") if s.strip())
            if statuses:
                clauses.append(f"status:[{statuses}]")
        filter_str = "+".join(clauses)

        try:
            spotlight = SpotlightVulnerabilities(auth_object=self.auth, timeout=30)
            resp = spotlight.query_vulnerabilities_combined(
                filter=filter_str,
                limit=min(max(1, limit), 5000),
                sort="updated_timestamp|desc",
                facet=["cve", "host_info", "remediation"],
            )
            if resp.get("status_code") != 200:
                errors = resp.get("body", {}).get("errors", [])
                return {"error": f"Spotlight query failed: {errors}", "fql_query": filter_str}

            resources = resp.get("body", {}).get("resources", []) or []
            hosts = [self._summarize_vuln(v) for v in resources]
            # De-dup to one row per host (a host can carry the CVE on >1 product).
            seen, unique = set(), []
            for h in hosts:
                key = h.get("hostname")
                if key and key in seen:
                    continue
                seen.add(key)
                unique.append(h)
            return {
                "cve_id": cve_id,
                "host_count": len(unique),
                "hosts": unique,
                "fql_query": filter_str,
            }
        except Exception as e:
            logger.error(f"Error in search_vulnerabilities_by_cve({cve_id}): {e}")
            return {"error": str(e), "fql_query": filter_str}

    # ------------------------------------------------------------------ #
    # Quarantine — quarantined-file triage + release/unrelease/delete    #
    # ------------------------------------------------------------------ #

    QUARANTINE_ACTIONS = ("release", "unrelease", "delete")

    @staticmethod
    def _summarize_quarantine_file(f: Dict[str, Any]) -> Dict[str, Any]:
        """Trim a quarantined-file record to the fields an analyst triages on."""
        paths = f.get("paths", []) or []
        return {
            "id": f.get("id"),
            "sha256": f.get("sha256"),
            "state": f.get("state"),                       # quarantined / released / deleted
            "hostname": f.get("hostname"),
            "username": f.get("username"),
            "paths": [p.get("path") for p in paths if p.get("path")],
            "detect_ids": f.get("detect_ids", []),
            "date_created": f.get("date_created"),
            "date_updated": f.get("date_updated"),
        }

    def query_quarantine_files(
        self,
        hostname: Optional[str] = None,
        sha256: Optional[str] = None,
        status: Optional[str] = None,
        limit: int = 100,
    ) -> Dict[str, Any]:
        """List quarantined files, optionally scoped to a host / hash / status.

        Answers "what has CrowdStrike quarantined on this box?" — the read view a
        SOC analyst needs before deciding to release a false positive or delete a
        confirmed-malicious file. Returns summarized file metadata (resolved from
        IDs in one follow-up call).

        Args:
            hostname: Filter to a single device hostname.
            sha256: Filter to a specific file hash (matched via the q phrase search).
            status: Filter by state (e.g. 'quarantined', 'released', 'deleted').
            limit: Max files to return.

        Returns:
            Dict with 'count', 'files' (summarized) and the 'fql_query' used, or 'error'.
        """
        clauses = []
        if hostname:
            clauses.append(f"device.hostname:'{hostname.strip()}'")
        if status:
            clauses.append(f"status:'{status.strip()}'")
        filter_str = "+".join(clauses) if clauses else "*"

        try:
            q_client = Quarantine(auth_object=self.auth, timeout=30)
            kwargs: Dict[str, Any] = {
                "filter": filter_str,
                "limit": min(max(1, limit), 5000),
                "sort": "date_created|desc",
            }
            if sha256:
                kwargs["q"] = sha256.strip()
            resp = q_client.query_quarantine_files(**kwargs)
            if resp.get("status_code") != 200:
                errors = resp.get("body", {}).get("errors", [])
                return {"error": f"Quarantine query failed: {errors}", "fql_query": filter_str}

            file_ids = resp.get("body", {}).get("resources", []) or []
            if not file_ids:
                return {"count": 0, "files": [], "fql_query": filter_str}

            details = q_client.get_quarantine_files(ids=file_ids[:100])
            if details.get("status_code") != 200:
                errors = details.get("body", {}).get("errors", [])
                return {"error": f"Quarantine detail fetch failed: {errors}", "fql_query": filter_str}

            files = [self._summarize_quarantine_file(f)
                     for f in details.get("body", {}).get("resources", []) or []]
            return {"count": len(file_ids), "files": files, "fql_query": filter_str}
        except Exception as e:
            logger.error(f"Error in query_quarantine_files: {e}")
            return {"error": str(e), "fql_query": filter_str}

    def get_quarantine_file_details(self, ids: List[str]) -> Dict[str, Any]:
        """Fetch full metadata for specific quarantined-file IDs."""
        if not ids:
            return {"error": "No quarantine file IDs provided"}
        try:
            q_client = Quarantine(auth_object=self.auth, timeout=30)
            resp = q_client.get_quarantine_files(ids=ids[:100])
            if resp.get("status_code") != 200:
                errors = resp.get("body", {}).get("errors", [])
                return {"error": f"Quarantine detail fetch failed: {errors}"}
            files = [self._summarize_quarantine_file(f)
                     for f in resp.get("body", {}).get("resources", []) or []]
            return {"count": len(files), "files": files}
        except Exception as e:
            logger.error(f"Error in get_quarantine_file_details: {e}")
            return {"error": str(e)}

    def update_quarantine_files(
        self, action: str, ids: List[str], comment: str = ""
    ) -> Dict[str, Any]:
        """Apply a containment action to quarantined files — the response action.

        release   -> restore the file to its original location (false positive).
        unrelease -> re-quarantine a previously released file.
        delete    -> permanently remove the quarantined file (irreversible).

        This MUTATES endpoint state. The client must be built with a credential
        profile that carries the 'Quarantine: Write' scope.

        Args:
            action: One of 'release', 'unrelease', 'delete'.
            ids: Quarantine file IDs to act on.
            comment: Audit comment recorded alongside the action.

        Returns:
            Dict with 'action', 'ids', 'status_code', 'ok', plus 'errors' on failure.
        """
        action = (action or "").strip().lower()
        if action not in self.QUARANTINE_ACTIONS:
            return {"error": f"Invalid action '{action}'. "
                             f"Must be one of {', '.join(self.QUARANTINE_ACTIONS)}."}
        if not ids:
            return {"error": "No quarantine file IDs provided"}

        try:
            q_client = Quarantine(auth_object=self.auth, timeout=30)
            resp = q_client.update_quarantined_detects_by_id(
                action=action, ids=ids, comment=comment or f"{action} via IR/Sleuth"
            )
            status_code = resp.get("status_code")
            if status_code == 200:
                return {"action": action, "ids": ids, "status_code": 200, "ok": True}
            errors = resp.get("body", {}).get("errors", [])
            return {"action": action, "ids": ids, "status_code": status_code,
                    "ok": False, "errors": errors}
        except Exception as e:
            logger.error(f"Error in update_quarantine_files({action}): {e}")
            return {"action": action, "ids": ids, "ok": False, "error": str(e)}

    # ------------------------------------------------------------------ #
    # Identity Protection — entity risk investigation (GraphQL)          #
    # ------------------------------------------------------------------ #

    # Falcon Identity Protection risk severities, weakest -> strongest.
    IDP_SEVERITIES = ["LOW", "MEDIUM", "HIGH"]

    def _run_idp_graphql(self, query: str):
        """Execute an Identity Protection GraphQL query.

        Returns (data, None) on success or (None, error_message) on failure —
        the IDP API is GraphQL, so a 200 can still carry an 'errors' array.
        """
        idp = IdentityProtection(auth_object=self.auth, timeout=30)
        resp = idp.graphql(body={"query": query})
        status = resp.get("status_code")
        body = resp.get("body", {}) or {}
        if status != 200:
            return None, f"Identity Protection query failed (HTTP {status}): {body.get('errors')}"
        if body.get("errors"):
            return None, f"Identity Protection GraphQL errors: {body.get('errors')}"
        return body.get("data", {}) or {}, None

    @staticmethod
    def _summarize_idp_entity(node: Dict[str, Any]) -> Dict[str, Any]:
        """Trim an Identity Protection entity node to the risk fields analysts read."""
        factors = node.get("riskFactors", []) or []
        return {
            "entity_id": node.get("entityId"),
            "name": node.get("primaryDisplayName"),
            "secondary_name": node.get("secondaryDisplayName"),
            "type": node.get("type"),
            "risk_score": node.get("riskScore"),
            "risk_severity": node.get("riskScoreSeverity"),
            "emails": node.get("emailAddresses", []) or [],
            "risk_factors": [
                {"type": f.get("type"), "severity": f.get("severity")}
                for f in factors
            ],
        }

    # Shared node selection — kept identical across queries so the summarizer
    # always sees the same fields. emailAddresses lives on UserEntity only.
    _IDP_NODE_FIELDS = (
        "entityId primaryDisplayName secondaryDisplayName type "
        "riskScore riskScoreSeverity riskFactors { type severity } "
        "... on UserEntity { emailAddresses }"
    )

    def get_identity_entity_risk(
        self, name: str, limit: int = 10
    ) -> Dict[str, Any]:
        """Look up Falcon Identity Protection risk for an entity by display name.

        Answers "what's the identity risk on this user/account?" — risk score,
        severity and the contributing risk factors (e.g. stale account, weak
        password, attack-path exposure). Searches by primary display name and
        returns the riskiest matches first.

        Args:
            name: The entity's display name (user or endpoint), e.g. "Jane Doe".
            limit: Max matching entities to return.

        Returns:
            Dict with 'query', 'count', 'entities' (summarized), or 'error'.
        """
        name = (name or "").strip()
        if not name:
            return {"error": "Empty entity name"}

        import json
        gql = f"""
        query {{
          entities(primaryDisplayNames: [{json.dumps(name)}],
                   first: {min(max(1, limit), 100)},
                   sortKey: RISK_SCORE, sortOrder: DESCENDING, archived: false) {{
            nodes {{ {self._IDP_NODE_FIELDS} }}
          }}
        }}
        """
        data, err = self._run_idp_graphql(gql)
        if err:
            return {"error": err, "query": name}
        nodes = (data.get("entities", {}) or {}).get("nodes", []) or []
        entities = [self._summarize_idp_entity(n) for n in nodes]
        return {"query": name, "count": len(entities), "entities": entities}

    def get_high_risk_identities(
        self, min_severity: str = "HIGH", limit: int = 20
    ) -> Dict[str, Any]:
        """List the highest-risk identity entities in the tenant right now.

        Answers "who are our riskiest identities?" — the prioritized identity
        watchlist for a threat hunter / identity-focused analyst.

        Args:
            min_severity: Lowest severity to include ('LOW','MEDIUM','HIGH').
                Defaults to HIGH (most urgent only).
            limit: Max entities to return.

        Returns:
            Dict with 'min_severity', 'count', 'entities' (summarized), or 'error'.
        """
        min_severity = (min_severity or "HIGH").strip().upper()
        if min_severity not in self.IDP_SEVERITIES:
            return {"error": f"Invalid severity '{min_severity}'. "
                             f"Must be one of {', '.join(self.IDP_SEVERITIES)}."}
        # Include the requested severity and everything above it.
        idx = self.IDP_SEVERITIES.index(min_severity)
        severities = ", ".join(self.IDP_SEVERITIES[idx:])

        gql = f"""
        query {{
          entities(riskScoreSeverities: [{severities}],
                   first: {min(max(1, limit), 100)},
                   sortKey: RISK_SCORE, sortOrder: DESCENDING, archived: false) {{
            nodes {{ {self._IDP_NODE_FIELDS} }}
          }}
        }}
        """
        data, err = self._run_idp_graphql(gql)
        if err:
            return {"error": err, "min_severity": min_severity}
        nodes = (data.get("entities", {}) or {}).get("nodes", []) or []
        entities = [self._summarize_idp_entity(n) for n in nodes]
        return {"min_severity": min_severity, "count": len(entities), "entities": entities}

    def get_devices_by_ioc(self, ioc_value: str, ioc_type: Optional[str] = None) -> Dict[str, Any]:
        """Find the endpoints that observed an IOC (domain / IP / file hash).

        Uses CrowdStrike's IOC "devices ran on" index — given an indicator
        value it returns the device IDs that have seen it — then resolves those
        IDs to hostnames. This is the IOC -> affected-hosts pivot, e.g. "which
        hosts connected to yowgames.com?" or "what machines executed this hash?".

        Args:
            ioc_value: The indicator value — a domain, IPv4/IPv6 address, or
                MD5/SHA1/SHA256 hash.
            ioc_type: Optional explicit type ('domain','ipv4','ipv6','md5',
                'sha1','sha256'). Auto-detected from the value when omitted.

        Returns:
            Dict with 'ioc', 'ioc_type', 'device_count' (managed hosts actually
            resolved), 'references_returned' (raw device references the IOC index
            returned), 'hosts' (each: hostname, device_id, last_seen, platform,
            local_ip, machine_domain), and a 'note' when references came back but
            none map to a managed host — or 'error'.
        """
        ioc_value = self._refang(ioc_value)
        if not ioc_value:
            return {"error": "Empty IOC value"}
        ioc_type = ioc_type or self._detect_ioc_type(ioc_value)
        if not ioc_type:
            return {"error": f"Could not determine IOC type for '{ioc_value}'. "
                             "Pass ioc_type explicitly (domain/ipv4/ipv6/md5/sha1/sha256)."}
        try:
            iocs_client = Iocs(auth_object=self.auth, timeout=30)
            resp = iocs_client.devices_ran_on(type=ioc_type, value=ioc_value, limit=100)
            if resp.get("status_code") != 200:
                return {"error": f"IOC device search failed: {resp.get('body', {}).get('errors')}",
                        "ioc": ioc_value, "ioc_type": ioc_type}
            device_ids = resp.get("body", {}).get("resources", [])
            if not device_ids:
                return {"ioc": ioc_value, "ioc_type": ioc_type, "device_count": 0,
                        "references_returned": 0, "hosts": []}
            hosts = self._resolve_device_ids(device_ids)
            result = {
                "ioc": ioc_value,
                "ioc_type": ioc_type,
                "device_count": len(hosts),                 # managed hosts actually resolved
                "references_returned": len(device_ids),     # raw refs the IOC index returned
                "hosts": hosts,
            }
            # Guard against the misleading "N references, 0 hostnames" result: the
            # IOC index can return device references that don't map to any managed
            # host in this tenant (globally-common/benign hashes, stale refs).
            # Report that plainly rather than implying N affected hosts.
            if device_ids and not hosts:
                result["note"] = (
                    f"CrowdStrike's IOC index returned {len(device_ids)} device "
                    "reference(s) for this indicator, but none resolve to a "
                    "currently-managed host in this tenant. This is expected for "
                    "globally-common or benign hashes and for stale references, "
                    "and means there is no confirmed managed-host hit — not that "
                    "those hosts are affected."
                )
            return result
        except Exception as e:
            logger.error(f"Error in get_devices_by_ioc({ioc_value}): {e}")
            return {"error": str(e), "ioc": ioc_value, "ioc_type": ioc_type}

    def search_detections_by_ip(self, ip: str, hours: int = 168) -> Dict[str, Any]:
        """Search for detections/alerts involving an IP address.

        Note: Uses Alerts API (Detects API was decommissioned by CrowdStrike).

        Args:
            ip: The IP address to search for
            hours: Hours to look back

        Returns:
            Dict with detection results or error
        """
        from datetime import timedelta, timezone

        # Build filter string outside try block (always returned for transparency)
        start_date = (datetime.now(timezone.utc) - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
        filter_str = f"(device.local_ip:'{ip}',device.external_ip:'{ip}')+created_timestamp:>='{start_date}'"

        try:
            response = self.alerts_client.query_alerts_v1(filter=filter_str, limit=100)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"Detection search failed: {error_msg}", "fql_query": filter_str}

            alert_ids = response.get("body", {}).get("resources", [])

            if not alert_ids:
                return {"count": 0, "detections": [], "fql_query": filter_str}

            # Get alert details
            details_resp = self.alerts_client.get_alerts_v1(ids=alert_ids[:20])
            if details_resp.get("status_code") == 200:
                detections = details_resp.get("body", {}).get("resources", [])
                return {
                    "count": len(alert_ids),
                    "detections": detections,
                    "fql_query": filter_str
                }

            return {"count": len(alert_ids), "detections": [], "fql_query": filter_str}

        except Exception as e:
            logger.error(f"Error searching detections by IP {ip}: {e}")
            return {"error": str(e), "fql_query": filter_str}

    def batch_search_detections_by_ips(self, ips: list, hours: int = 168) -> Dict[str, Any]:
        """Search for detections/alerts involving multiple IP addresses in a single query.

        Note: Uses Alerts API (Detects API was decommissioned by CrowdStrike).

        Args:
            ips: List of IP addresses to search for
            hours: Hours to look back

        Returns:
            Dict with detection results grouped by IP, or error
        """
        from datetime import timedelta, timezone

        if not ips:
            return {"count": 0, "detections": [], "by_ip": {}}

        # Build filter string outside try block (always returned for transparency)
        start_date = (datetime.now(timezone.utc) - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
        ip_list = ",".join([f"'{ip}'" for ip in ips])
        filter_str = f"(device.local_ip:[{ip_list}],device.external_ip:[{ip_list}])+created_timestamp:>='{start_date}'"

        try:
            response = self.alerts_client.query_alerts_v1(filter=filter_str, limit=500)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"Detection search failed: {error_msg}", "fql_query": filter_str}

            alert_ids = response.get("body", {}).get("resources", [])

            if not alert_ids:
                return {"count": 0, "detections": [], "by_ip": {}, "fql_query": filter_str}

            # Get alert details
            details_resp = self.alerts_client.get_alerts_v1(ids=alert_ids[:100])
            detections = []
            if details_resp.get("status_code") == 200:
                detections = details_resp.get("body", {}).get("resources", [])

            # Group detections by IP
            by_ip = {ip: {"count": 0, "detections": [], "hostnames": set()} for ip in ips}
            for det in detections:
                device = det.get("device", {})
                local_ip = device.get("local_ip", "")
                external_ip = device.get("external_ip", "")
                hostname = device.get("hostname", "")

                for ip in ips:
                    if ip == local_ip or ip == external_ip:
                        by_ip[ip]["count"] += 1
                        by_ip[ip]["detections"].append(det)
                        if hostname:
                            by_ip[ip]["hostnames"].add(hostname)

            # Convert sets to lists
            for ip in by_ip:
                by_ip[ip]["hostnames"] = list(by_ip[ip]["hostnames"])

            return {
                "count": len(alert_ids),
                "detections": detections,
                "by_ip": by_ip,
                "fql_query": filter_str
            }

        except Exception as e:
            logger.error(f"Error batch searching detections by IPs: {e}")
            return {"error": str(e), "fql_query": filter_str}

    def search_detections_by_hash(self, file_hash: str, hours: int = 168) -> Dict[str, Any]:
        """Search for detections/alerts involving a file hash.

        Note: Uses Alerts API (Detects API was decommissioned by CrowdStrike).

        Args:
            file_hash: The MD5 or SHA256 hash to search for
            hours: Hours to look back

        Returns:
            Dict with detection results or error
        """
        from datetime import timedelta, timezone

        # Build filter string outside try block (always returned for transparency)
        start_date = (datetime.now(timezone.utc) - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
        if len(file_hash) == 32:
            filter_str = f"behaviors.md5:'{file_hash}'"
        else:
            filter_str = f"behaviors.sha256:'{file_hash}'"
        filter_str += f"+created_timestamp:>='{start_date}'"

        try:
            response = self.alerts_client.query_alerts_v1(filter=filter_str, limit=100)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"Detection search failed: {error_msg}", "fql_query": filter_str}

            alert_ids = response.get("body", {}).get("resources", [])

            if not alert_ids:
                return {"count": 0, "detections": [], "fql_query": filter_str}

            # Get alert details
            details_resp = self.alerts_client.get_alerts_v1(ids=alert_ids[:20])
            if details_resp.get("status_code") == 200:
                detections = details_resp.get("body", {}).get("resources", [])
                return {
                    "count": len(alert_ids),
                    "detections": detections,
                    "fql_query": filter_str
                }

            return {"count": len(alert_ids), "detections": [], "fql_query": filter_str}

        except Exception as e:
            logger.error(f"Error searching detections by hash {file_hash[:16]}...: {e}")
            return {"error": str(e), "fql_query": filter_str}

    def search_detections_by_filename(self, filename: str, hours: int = 168) -> Dict[str, Any]:
        """Search for detections/alerts involving a specific filename.

        Note: Uses Alerts API (Detects API was decommissioned by CrowdStrike).

        Useful for hunting malicious scripts like install.ps1, install.sh, etc.

        Args:
            filename: The filename to search for (e.g., "install.ps1")
            hours: Hours to look back

        Returns:
            Dict with detection results or error
        """
        from datetime import timedelta, timezone

        # Build filter string outside try block (always returned for transparency)
        start_date = (datetime.now(timezone.utc) - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
        filter_str = f"behaviors.filename:*'{filename}'+created_timestamp:>='{start_date}'"

        try:
            response = self.alerts_client.query_alerts_v1(filter=filter_str, limit=100)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"Detection search failed: {error_msg}", "fql_query": filter_str}

            alert_ids = response.get("body", {}).get("resources", [])

            if not alert_ids:
                return {"count": 0, "detections": [], "fql_query": filter_str}

            # Get alert details
            details_resp = self.alerts_client.get_alerts_v1(ids=alert_ids[:20])
            if details_resp.get("status_code") == 200:
                detections = details_resp.get("body", {}).get("resources", [])
                return {
                    "count": len(alert_ids),
                    "detections": detections,
                    "fql_query": filter_str
                }

            return {"count": len(alert_ids), "detections": [], "fql_query": filter_str}

        except Exception as e:
            logger.error(f"Error searching detections by filename {filename}: {e}")
            return {"error": str(e), "fql_query": filter_str}

    def lookup_intel_indicator(self, indicator: str) -> Dict[str, Any]:
        """Lookup threat intelligence for an indicator (Falcon X).

        Args:
            indicator: The indicator value (IP, domain, hash, URL)

        Returns:
            Dict with intel results or error
        """
        # Build the filter string (always returned for transparency)
        filter_str = f"indicator:'{indicator}'"

        try:
            intel_client = Intel(auth_object=self.auth, timeout=30)

            response = intel_client.query_indicator_entities(
                filter=filter_str,
                limit=10
            )

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"Intel lookup failed: {error_msg}", "fql_query": filter_str}

            resources = response.get("body", {}).get("resources", [])
            return {
                "count": len(resources),
                "indicators": resources,
                "fql_query": filter_str
            }

        except Exception as e:
            logger.error(f"Error looking up intel for {indicator}: {e}")
            return {"error": str(e), "fql_query": filter_str}

    def search_threatgraph_domain(self, domain: str) -> Dict[str, Any]:
        """Search ThreatGraph for hosts that connected to a domain.

        Uses CrowdStrike's ThreatGraph API to find hosts that have
        DNS requests or network connections to the specified domain.

        Args:
            domain: The domain to search for

        Returns:
            Dict with vertex/edge results or error
        """
        # Build the query description for analysts (always returned for transparency)
        query_str = f"ThreatGraph.combined_summary_get(ids=['{domain}'], vertex_types=['domain', 'host'], edge_types=['dns_request', 'communicates_with'])"

        try:
            from falconpy import ThreatGraph

            tg_client = ThreatGraph(auth_object=self.auth, timeout=30)

            # Search for domain indicator and get related hosts
            response = tg_client.combined_summary_get(
                ids=[domain],
                vertex_types=["domain", "host"],
                edge_types=["dns_request", "communicates_with"]
            )

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"ThreatGraph search failed: {error_msg}", "api_call": query_str}

            resources = response.get("body", {}).get("resources", [])
            # Extract affected hosts from the graph
            hosts = []
            for resource in resources:
                vertices = resource.get("vertices", {})
                for vertex_id, vertex in vertices.items():
                    if vertex.get("vertex_type") == "host":
                        hostname = vertex.get("properties", {}).get("hostname", "")
                        if hostname:
                            hosts.append(hostname)

            return {
                "count": len(hosts),
                "hosts": list(set(hosts))[:20],
                "raw": resources,
                "api_call": query_str
            }

        except Exception as e:
            logger.error(f"Error searching ThreatGraph for domain {domain}: {e}")
            return {"error": str(e), "api_call": query_str}

    def batch_search_threatgraph_ips(self, ips: list) -> Dict[str, Any]:
        """Search ThreatGraph for hosts that communicated with IP addresses.

        Uses CrowdStrike's ThreatGraph API to find hosts that have
        network connections to the specified IPs. This shows network
        activity even without a detection being triggered.

        Args:
            ips: List of IP addresses to search for

        Returns:
            Dict with results grouped by IP, or error
        """
        if not ips:
            return {"count": 0, "by_ip": {}}

        # Build the query description for analysts (always returned for transparency)
        ip_list_str = ", ".join([f"'{ip}'" for ip in ips])
        query_str = f"ThreatGraph.combined_summary_get(ids=[{ip_list_str}], vertex_types=['ip_address', 'host'], edge_types=['communicates_with'])"

        try:
            from falconpy import ThreatGraph

            tg_client = ThreatGraph(auth_object=self.auth, timeout=60)

            # Search for all IPs at once
            response = tg_client.combined_summary_get(
                ids=ips,
                vertex_types=["ip_address", "host"],
                edge_types=["communicates_with"]
            )

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"ThreatGraph search failed: {error_msg}", "api_call": query_str}

            resources = response.get("body", {}).get("resources", [])

            # Group results by IP
            by_ip = {ip: {"hosts": set(), "count": 0} for ip in ips}
            total_hosts = set()

            for resource in resources:
                vertices = resource.get("vertices", {})
                edges = resource.get("edges", {})

                # Map vertex IDs to their data
                vertex_map = {}
                for vertex_id, vertex in vertices.items():
                    vertex_map[vertex_id] = vertex

                # Find connections: IP -> Host via edges
                for edge_id, edge in edges.items():
                    source_id = edge.get("source_vertex_id", "")
                    dest_id = edge.get("destination_vertex_id", "")

                    source_vertex = vertex_map.get(source_id, {})
                    dest_vertex = vertex_map.get(dest_id, {})

                    # Check if one is IP and other is host
                    if source_vertex.get("vertex_type") == "ip_address":
                        ip_props = source_vertex.get("properties", {})
                        ip_value = ip_props.get("ip_address", "")
                        if ip_value in ips and dest_vertex.get("vertex_type") == "host":
                            hostname = dest_vertex.get("properties", {}).get("hostname", "")
                            if hostname:
                                by_ip[ip_value]["hosts"].add(hostname)
                                total_hosts.add(hostname)

                    elif dest_vertex.get("vertex_type") == "ip_address":
                        ip_props = dest_vertex.get("properties", {})
                        ip_value = ip_props.get("ip_address", "")
                        if ip_value in ips and source_vertex.get("vertex_type") == "host":
                            hostname = source_vertex.get("properties", {}).get("hostname", "")
                            if hostname:
                                by_ip[ip_value]["hosts"].add(hostname)
                                total_hosts.add(hostname)

            # Convert sets to lists and add counts
            for ip in by_ip:
                by_ip[ip]["hosts"] = list(by_ip[ip]["hosts"])[:20]
                by_ip[ip]["count"] = len(by_ip[ip]["hosts"])

            return {
                "count": len(total_hosts),
                "total_hosts": list(total_hosts)[:50],
                "by_ip": by_ip,
                "api_call": query_str
            }

        except Exception as e:
            logger.error(f"Error searching ThreatGraph for IPs: {e}")
            return {"error": str(e), "api_call": query_str}

    def search_dns_requests(self, domain: str, hours: int = 168) -> Dict[str, Any]:
        """Search for DNS requests to a domain using alerts.

        Note: Uses Alerts API (Detects API was decommissioned by CrowdStrike).

        Searches for alerts where the domain appears in network/DNS fields.

        Args:
            domain: The domain to search for
            hours: Hours to look back

        Returns:
            Dict with detection results or error
        """
        from datetime import timedelta, timezone

        try:
            start_date = (datetime.now(timezone.utc) - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")

            # Search alerts for domain in network fields
            # behaviors.dns_requests contains domain info
            filter_str = f"behaviors.dns_requests.domain:*'{domain}'"
            filter_str += f"+created_timestamp:>='{start_date}'"

            response = self.alerts_client.query_alerts_v1(filter=filter_str, limit=100)

            if response.get("status_code") != 200:
                # Fallback: try a broader time-based filter
                filter_str = f"created_timestamp:>='{start_date}'"
                response = self.alerts_client.query_alerts_v1(filter=filter_str, limit=100)

            alert_ids = response.get("body", {}).get("resources", [])

            if not alert_ids:
                return {"count": 0, "detections": [], "hosts": []}

            # Get alert details and filter for domain
            details_resp = self.alerts_client.get_alerts_v1(ids=alert_ids[:50])
            matching_detections = []
            hosts = set()

            if details_resp.get("status_code") == 200:
                for det in details_resp.get("body", {}).get("resources", []):
                    # Check behaviors for domain references
                    for behavior in det.get("behaviors", []):
                        dns_requests = behavior.get("dns_requests", [])
                        network = behavior.get("network_accesses", [])

                        domain_found = False
                        for dns in dns_requests:
                            if domain.lower() in str(dns).lower():
                                domain_found = True
                                break
                        for net in network:
                            if domain.lower() in str(net).lower():
                                domain_found = True
                                break

                        if domain_found:
                            matching_detections.append(det)
                            hostname = det.get("device", {}).get("hostname")
                            if hostname:
                                hosts.add(hostname)
                            break

            return {
                "count": len(matching_detections),
                "detections": matching_detections[:10],
                "hosts": list(hosts)[:20]
            }

        except Exception as e:
            logger.error(f"Error searching DNS requests for {domain}: {e}")
            return {"error": str(e)}

    def search_alerts_by_ip(self, ip: str, hours: int = 168) -> Dict[str, Any]:
        """Search for alerts involving an IP address.

        Args:
            ip: The IP address to search for
            hours: Hours to look back

        Returns:
            Dict with alert results or error
        """
        from datetime import timedelta, timezone

        # Build filter string outside try block (always returned for transparency)
        start_date = (datetime.now(timezone.utc) - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
        filter_str = f"(device.local_ip:'{ip}',device.external_ip:'{ip}')+created_timestamp:>='{start_date}'"

        try:
            response = self.alerts_client.query_alerts_v1(filter=filter_str, limit=100)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"Alert search failed: {error_msg}", "fql_query": filter_str}

            alert_ids = response.get("body", {}).get("resources", [])

            if not alert_ids:
                return {"count": 0, "alerts": [], "fql_query": filter_str}

            # Get alert details
            details_resp = self.alerts_client.get_alerts_v1(ids=alert_ids[:20])
            if details_resp.get("status_code") == 200:
                alerts = details_resp.get("body", {}).get("resources", [])
                return {
                    "count": len(alert_ids),
                    "alerts": alerts,
                    "fql_query": filter_str
                }

            return {"count": len(alert_ids), "alerts": [], "fql_query": filter_str}

        except Exception as e:
            logger.error(f"Error searching alerts by IP {ip}: {e}")
            return {"error": str(e), "fql_query": filter_str}

    def batch_search_alerts_by_ips(self, ips: list, hours: int = 168) -> Dict[str, Any]:
        """Search for alerts involving multiple IP addresses in a single query.

        Args:
            ips: List of IP addresses to search for
            hours: Hours to look back

        Returns:
            Dict with alert results grouped by IP, or error
        """
        from datetime import timedelta, timezone

        if not ips:
            return {"count": 0, "alerts": [], "by_ip": {}}

        # Build filter string outside try block (always returned for transparency)
        start_date = (datetime.now(timezone.utc) - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
        ip_list = ",".join([f"'{ip}'" for ip in ips])
        filter_str = f"(device.local_ip:[{ip_list}],device.external_ip:[{ip_list}])+created_timestamp:>='{start_date}'"

        try:
            response = self.alerts_client.query_alerts_v1(filter=filter_str, limit=500)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"Alert search failed: {error_msg}", "fql_query": filter_str}

            alert_ids = response.get("body", {}).get("resources", [])

            if not alert_ids:
                return {"count": 0, "alerts": [], "by_ip": {}, "fql_query": filter_str}

            # Get alert details
            details_resp = self.alerts_client.get_alerts_v1(ids=alert_ids[:100])
            alerts = []
            if details_resp.get("status_code") == 200:
                alerts = details_resp.get("body", {}).get("resources", [])

            # Group alerts by IP
            by_ip = {ip: {"count": 0, "alerts": [], "hostnames": set()} for ip in ips}
            for alert in alerts:
                device = alert.get("device", {})
                local_ip = device.get("local_ip", "")
                external_ip = device.get("external_ip", "")
                hostname = device.get("hostname", "")

                for ip in ips:
                    if ip == local_ip or ip == external_ip:
                        by_ip[ip]["count"] += 1
                        by_ip[ip]["alerts"].append(alert)
                        if hostname:
                            by_ip[ip]["hostnames"].add(hostname)

            # Convert sets to lists
            for ip in by_ip:
                by_ip[ip]["hostnames"] = list(by_ip[ip]["hostnames"])

            return {
                "count": len(alert_ids),
                "alerts": alerts,
                "by_ip": by_ip,
                "fql_query": filter_str
            }

        except Exception as e:
            logger.error(f"Error batch searching alerts by IPs: {e}")
            return {"error": str(e), "fql_query": filter_str}

    # ==================== Detection Rules Catalog Methods ====================

    def list_custom_ioa_rule_groups(self) -> Dict[str, Any]:
        """List all custom IOA rule groups with their rules.

        Returns:
            Dict with rule_groups list or error
        """
        try:
            from falconpy import CustomIOA
            ioa_client = CustomIOA(auth_object=self.auth, timeout=30)

            # Query all rule group IDs
            query_resp = ioa_client.query_rule_groupsMixin0(limit=500)
            if query_resp.get("status_code") != 200:
                error_msg = query_resp.get("body", {}).get("errors", [])
                return {"error": f"IOA rule group query failed: {error_msg}"}

            group_ids = query_resp.get("body", {}).get("resources", [])
            if not group_ids:
                return {"rule_groups": [], "count": 0}

            # Get full details for each group (includes rules within)
            details_resp = ioa_client.get_rule_groupsMixin0(ids=group_ids)
            if details_resp.get("status_code") != 200:
                error_msg = details_resp.get("body", {}).get("errors", [])
                return {"error": f"IOA rule group details failed: {error_msg}"}

            groups = details_resp.get("body", {}).get("resources", [])
            return {"rule_groups": groups, "count": len(groups)}

        except Exception as e:
            logger.error(f"Error listing custom IOA rule groups: {e}")
            return {"error": str(e)}

    def list_ioc_indicators(self, limit: int = 500) -> Dict[str, Any]:
        """List custom IOC indicators from CrowdStrike.

        Args:
            limit: Maximum number of indicators to return

        Returns:
            Dict with indicators list or error
        """
        try:
            ioc_client = IOC(auth_object=self.auth, timeout=30)
            response = ioc_client.indicator_combined_v1(limit=limit)

            if response.get("status_code") != 200:
                error_msg = response.get("body", {}).get("errors", [])
                return {"error": f"IOC indicator list failed: {error_msg}"}

            resources = response.get("body", {}).get("resources", [])
            return {"indicators": resources, "count": len(resources)}

        except Exception as e:
            logger.error(f"Error listing IOC indicators: {e}")
            return {"error": str(e)}

    def list_intel_yara_rules(self, limit: int = 500) -> Dict[str, Any]:
        """List YARA rule metadata from CrowdStrike Intel API.

        Queries for yara-master rule IDs, then fetches full rule entities.
        Handles 403 (no Intel API access) gracefully.

        Args:
            limit: Maximum number of rule IDs per page

        Returns:
            Dict with rules list and count, or error
        """
        try:
            intel_client = Intel(auth_object=self.auth, timeout=30)

            # Step 1: Paginate rule IDs
            all_ids = []
            offset = 0
            while True:
                id_response = intel_client.query_rule_ids(
                    type="yara-master", limit=limit, offset=offset
                )

                status_code = id_response.get("status_code", 0)

                if status_code == 403:
                    msg = "Intel YARA rules not permitted (403) — API scope not yet granted"
                    logger.warning(msg)
                    return {"error": msg}

                if status_code != 200:
                    error_msg = id_response.get("body", {}).get("errors", [])
                    return {"error": f"Intel YARA rule ID query failed: {error_msg}"}

                ids = id_response.get("body", {}).get("resources", [])
                if not ids:
                    break

                all_ids.extend(ids)
                offset += len(ids)

                if len(ids) < limit:
                    break

            if not all_ids:
                return {"rules": [], "count": 0}

            # Step 2: Batch-fetch rule entities
            all_rules = []
            for i in range(0, len(all_ids), 500):
                batch_ids = all_ids[i : i + 500]
                entity_response = intel_client.get_rule_entities(ids=batch_ids)

                if entity_response.get("status_code") != 200:
                    error_msg = entity_response.get("body", {}).get("errors", [])
                    return {"error": f"Intel YARA rule entity fetch failed: {error_msg}"}

                resources = entity_response.get("body", {}).get("resources", [])
                all_rules.extend(resources)

            return {"rules": all_rules, "count": len(all_rules)}

        except Exception as e:
            logger.error(f"Error listing Intel YARA rules: {e}")
            return {"error": str(e)}

    # ==================== LogScale/Event Search Methods ====================

    def run_logscale_query(
        self,
        query: str,
        start: str = "7d",
        end: str = "now",
        limit: int = 100,
        timeout: int = 60,
        repo: str = "base_sensor"
    ) -> Dict[str, Any]:
        """Run a LogScale query via the Foundry API.

        Args:
            query: LogScale query string (e.g., "#event_simpleName=ProcessRollup2 | head(10)")
            start: Start time (e.g., "7d", "24h", "2024-01-01T00:00:00Z")
            end: End time (e.g., "now", "1h", "2024-01-02T00:00:00Z")
            limit: Maximum results to return
            timeout: Query timeout in seconds
            repo: LogScale repository to query (default: base_sensor for EDR telemetry)

        Returns:
            Dict with events list, count, or error
        """
        try:
            from falconpy import FoundryLogScale

            # Get Foundry app_id from config - required for LogScale API access
            app_id = getattr(self.config, 'cs_foundry_app_id', None)
            if not app_id:
                return {
                    "error": "LogScale API not configured (set cs_foundry_app_id in config)",
                    "access_denied": True,
                    "not_configured": True
                }

            logscale = FoundryLogScale(auth_object=self.auth, timeout=timeout)

            # Execute dynamic query
            result = logscale.execute_dynamic(
                app_id=app_id,
                search_query=query,
                start=start,
                end=end,
                repo_or_view=repo,
                search_query_args={}
            )

            status_code = result.get('status_code', 0)

            if status_code == 401 or status_code == 403:
                # Access denied - API scope not available
                return {
                    "error": "LogScale API access not available (missing app-logs:write scope)",
                    "access_denied": True,
                    "status_code": status_code
                }

            if status_code == 400:
                error_msg = result.get('body', {}).get('errors', [])
                # Check for invalid app_id error
                if any('App ID is invalid' in str(e) for e in error_msg):
                    return {
                        "error": f"Invalid Foundry app_id in CS_FOUNDRY_APP_ID (check your CrowdStrike Foundry app registration)",
                        "access_denied": True,
                        "status_code": status_code
                    }
                return {"error": f"LogScale query failed: {error_msg}", "status_code": status_code}

            if status_code != 200:
                error_msg = result.get('body', {}).get('errors', [])
                return {"error": f"LogScale query failed: {error_msg}", "status_code": status_code}

            # Parse response - results may come back synchronously
            body = result.get('body', {})
            resources = body.get('resources', [])
            if not resources:
                return {"events": [], "count": 0, "query": query}

            resource = resources[0]
            job_status = resource.get('job_status', {})
            job_id = job_status.get('job_id')

            # Check if results already complete (synchronous response)
            if job_status.get('status') == 'complete':
                events = resource.get('events', [])
                return {
                    "events": events[:limit],
                    "count": len(events),
                    "query": query,
                    "job_id": job_id
                }

            # If not complete, poll for results (async)
            if not job_id:
                return {"error": "No job ID returned from LogScale"}

            max_attempts = timeout // 2
            for attempt in range(max_attempts):
                results = logscale.get_search_results(job_id=job_id)

                if results.get('status_code') != 200:
                    if attempt < max_attempts - 1:
                        time.sleep(2)
                        continue
                    return {"error": f"Failed to get results: {results.get('body', {})}"}

                result_body = results.get('body', {})
                res = result_body.get('resources', [{}])[0] if result_body.get('resources') else {}
                events = res.get('events', [])
                status = res.get('job_status', {}).get('status', 'running')

                if status == 'complete' or events:
                    return {
                        "events": events[:limit],
                        "count": len(events),
                        "query": query,
                        "job_id": job_id
                    }

                time.sleep(2)

            return {"error": "Query timed out", "query": query}

        except ImportError:
            return {"error": "FoundryLogScale not available in falconpy", "access_denied": True}
        except Exception as e:
            logger.error(f"Error running LogScale query: {e}")
            return {"error": str(e), "query": query}

    def run_logscale_queries_batch(
        self,
        queries: List[Dict[str, str]],
        hours: int = 168
    ) -> Dict[str, Any]:
        """Run multiple LogScale queries and aggregate results.

        Args:
            queries: List of query dicts with 'type' and 'query' keys
            hours: Hours to search back

        Returns:
            Dict with results per query, total events found, and any errors
        """
        results = {
            "total_events": 0,
            "queries_run": 0,
            "queries_failed": 0,
            "access_denied": False,
            "query_results": [],
            "errors": []
        }

        # Convert hours to LogScale time format
        days = hours // 24
        start = f"{days}d"

        for q in queries:
            query_type = q.get('type', 'Unknown')
            query_text = q.get('query', '')

            if not query_text:
                continue

            logger.info(f"[LogScale] Running: {query_type}")
            result = self.run_logscale_query(query_text, start=start, end="now", limit=100)

            if result.get('access_denied'):
                results["access_denied"] = True
                if result.get('not_configured'):
                    results["errors"].append("LogScale API not configured (set cs_foundry_app_id in config)")
                    logger.info("[LogScale] API not configured - set cs_foundry_app_id in config")
                else:
                    results["errors"].append(f"LogScale API access denied: {result.get('error', 'unknown')}")
                    logger.warning(f"[LogScale] API access denied: {result.get('error', 'unknown')}")
                break  # No point running more queries if access is denied

            if 'error' in result:
                results["queries_failed"] += 1
                results["errors"].append(f"{query_type}: {result['error']}")
                logger.warning(f"[LogScale] Query failed: {result['error']}")
            else:
                results["queries_run"] += 1
                event_count = result.get('count', 0)
                results["total_events"] += event_count

                results["query_results"].append({
                    "type": query_type,
                    "query": query_text,
                    "count": event_count,
                    "events": result.get('events', [])[:20]  # Limit stored events
                })

                if event_count > 0:
                    logger.info(f"[LogScale] {query_type}: {event_count} events found")

        return results


def process_unique_hosts(df: pd.DataFrame) -> pd.DataFrame:
    """Process dataframe to get unique hosts with latest last_seen.

    Retained for callers that already hold the full DataFrame in memory.
    For the disk → disk path, prefer update_unique_hosts_from_cs which
    avoids materializing the all-hosts DataFrame entirely.
    """
    df["last_seen"] = pd.to_datetime(df["last_seen"], errors='coerce', utc=True).dt.tz_convert(None)  # type: ignore[union-attr]
    return df.loc[df.groupby("hostname")["last_seen"].idxmax()]


def update_unique_hosts_from_cs() -> None:
    """Fetch CS hosts and write unique_cs_hosts.xlsx (latest last_seen per hostname).

    Streams all_cs_hosts.xlsx via openpyxl read-only and dedupes hostname → row in
    a dict; only the unique-hostname dict (bounded by fleet size, not raw row
    count) stays in memory. Previous implementation pd.read_excel'd the entire
    sheet, which OOMed on lab-vm2 at fleet scale.
    """
    from openpyxl import load_workbook, Workbook
    cs_client = CrowdStrikeClient()
    cs_client.fetch_all_hosts_and_write_to_xlsx()

    today_date = datetime.now().strftime('%m-%d-%Y')
    hosts_file = DATA_DIR / today_date / "all_cs_hosts.xlsx"
    unique_hosts_file = DATA_DIR / today_date / "unique_cs_hosts.xlsx"
    unique_hosts_file.parent.mkdir(parents=True, exist_ok=True)

    src_wb = load_workbook(hosts_file, read_only=True)
    try:
        src_ws = src_wb.active
        rows_iter = src_ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows_iter))
        except StopIteration:
            headers = []

        # last_seen comes from CS as ISO-8601 strings; lexicographic compare is
        # correct for that format. If a hostname-less or comparable-less row
        # shows up, keep whichever record was seen first.
        hostname_idx = headers.index("hostname") if "hostname" in headers else 0
        last_seen_idx = headers.index("last_seen") if "last_seen" in headers else None
        latest_by_hostname: Dict[str, tuple] = {}
        for row in rows_iter:
            hostname = row[hostname_idx]
            if not hostname:
                continue
            existing = latest_by_hostname.get(hostname)
            if existing is None:
                latest_by_hostname[hostname] = row
                continue
            if last_seen_idx is None:
                continue
            new_ls = row[last_seen_idx]
            old_ls = existing[last_seen_idx]
            if new_ls is None:
                continue
            if old_ls is None or new_ls > old_ls:
                latest_by_hostname[hostname] = row
    finally:
        src_wb.close()

    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet("Hosts")
    out_ws.append(headers)
    for row in latest_by_hostname.values():
        out_ws.append(list(row))
    out_wb.save(unique_hosts_file)
    # Intermediate file consumed by downstream code — skip apply_professional_formatting
    # since it would load the entire workbook back into memory and defeat streaming.


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description="CrowdStrike API test utility")
    parser.add_argument("--test", type=int, choices=[1, 2, 3, 4, 5, 6], default=1,
                        help="Test to run: 1=basic, 2=detection_by_ip, 3=detection_by_hash, "
                             "4=ioc_search, 5=intel_lookup, 6=alerts_by_ip")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO)
    client = CrowdStrikeClient()

    # Test token
    token = client.get_access_token()
    if not token:
        logger.error("Failed to obtain access token")
        return
    logger.info("Access token obtained successfully")

    if args.test == 1:
        # Test 1: Basic device queries
        logger.info("=== Test 1: Basic device queries ===")
        host_name_cs = 'uscku1metu03c7l'
        device_id = client.get_device_id(host_name_cs)
        if device_id:
            logger.info(f"Device ID: {device_id}")
            logger.info(client.get_device_details(device_id))

        containment_status = client.get_device_containment_status(host_name_cs)
        logger.info(f"Containment status: {containment_status}")

        online_status = client.get_device_online_state(host_name_cs)
        logger.info(f"Online status: {online_status}")

    elif args.test == 2:
        # Test 2: Search detections by IP
        logger.info("=== Test 2: Search detections by IP ===")
        test_ip = "<internal-host>"  # Example internal IP
        result = client.search_detections_by_ip(test_ip, hours=168)
        if "error" in result:
            logger.error(f"Error: {result['error']}")
        else:
            logger.info(f"Found {result['count']} detections for IP {test_ip}")
            for d in result.get('detections', [])[:3]:
                logger.info(f"  - {d.get('detection_id')}: {d.get('detect_description', 'N/A')}")

    elif args.test == 3:
        # Test 3: Search detections by hash
        logger.info("=== Test 3: Search detections by hash ===")
        # Example SHA256 (Mimikatz)
        test_hash = "e930b05efe23891d19bc354a4209be3e113e2b8a1a5c19c5d1c5a5e5a5e5a5e5"
        result = client.search_detections_by_hash(test_hash, hours=720)
        if "error" in result:
            logger.error(f"Error: {result['error']}")
        else:
            logger.info(f"Found {result['count']} detections for hash {test_hash[:16]}...")
            for d in result.get('detections', [])[:3]:
                hostname = d.get('device', {}).get('hostname', 'Unknown')
                logger.info(f"  - Host: {hostname}, Detection: {d.get('detect_description', 'N/A')}")

    elif args.test == 4:
        # Test 4: Search custom IOCs
        logger.info("=== Test 4: Search custom IOCs ===")
        test_ioc = "8.8.8.8"  # Example IOC
        result = client.search_ioc_by_value(test_ioc)
        if "error" in result:
            logger.error(f"Error: {result['error']}")
        else:
            logger.info(f"Found {result['count']} custom IOCs matching {test_ioc}")
            for ioc in result.get('indicators', [])[:3]:
                logger.info(f"  - Type: {ioc.get('type')}, Value: {ioc.get('value')}")

    elif args.test == 5:
        # Test 5: Falcon X intel lookup
        logger.info("=== Test 5: Falcon X intel lookup ===")
        test_indicator = "evil.com"  # Example domain
        result = client.lookup_intel_indicator(test_indicator)
        if "error" in result:
            logger.error(f"Error: {result['error']}")
        else:
            logger.info(f"Found {result['count']} intel records for {test_indicator}")
            for ind in result.get('indicators', [])[:3]:
                logger.info(f"  - Type: {ind.get('type')}, Malicious: {ind.get('malicious_confidence')}")

    elif args.test == 6:
        # Test 6: Search alerts by IP
        logger.info("=== Test 6: Search alerts by IP ===")
        test_ip = "<internal-host>"
        result = client.search_alerts_by_ip(test_ip, hours=168)
        if "error" in result:
            logger.error(f"Error: {result['error']}")
        else:
            logger.info(f"Found {result['count']} alerts for IP {test_ip}")
            for a in result.get('alerts', [])[:3]:
                logger.info(f"  - Alert: {a.get('name', 'N/A')}")


if __name__ == "__main__":
    main()
