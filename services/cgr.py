"""Prisma Cloud container-image (CGR) vulnerability export ingestion + triage reconcile.

Prisma Cloud exports a very large spreadsheet of container-image vulnerability
findings (grain = one row per CVE x image x package occurrence). This module
stream-parses that export memory-safely (the file is ~175MB / ~959k rows, so we
never materialize the rows -- only bounded per-CVE / per-cluster aggregates),
writes three rolled-up SQLite tables, and reconciles the scanner's CVE coverage
against our own CVE-triage results so the triage app can act as a
vendor-independent backup that surfaces CVEs Prisma flagged but we have not yet
triaged (the coverage gap).

This module also overlays an EAI-enriched composite risk score on top of the
ingested CGR findings. Each CGR CVE is joined to the cached EAI application
inventory (``eai_app_info.db``, a snapshot of EAI prod ``dbo.V_APP_INFO``) and
to our CVE-triage enrichment, then scored with the war-room composite model:

    composite = (KEV_listed      ? 5.0 : 0)
              + (external_facing  ? 2.5 : 0)
              + (pre_auth         ? 1.5 : 0)
              + (cvss_mean / 10.0) * 0.3

CVEs whose description/package signal contains TLS or CURL are flagged exempt
(not deleted) -- the war room exempted those en masse.

Entry points:
    latest_export()            -> Path | None
    ingest_export(path=None)   -> dict
    cgr_cves()                 -> set[str]
    reconcile(triage_db_path)  -> dict
    summary()                  -> dict
    cluster_rollup(limit=80)   -> list[dict]
    cve_gaps(limit=200)        -> list[dict]
    eai_info(eai_code)         -> dict | None
    load_eai_map()             -> dict[str, dict]
    refresh_eai_cache(...)     -> dict          (optional live re-pull, on-tunnel)
    score_all(...)             -> dict          (rebuild cgr_scored)
    top_composite(...)         -> list[dict]
    gaps(limit=200)            -> list[dict]
    scoring_summary()          -> dict
"""

import argparse
import json
import logging
import os
import re
import sqlite3
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

logger = logging.getLogger(__name__)

# Both worktrees share the transient export drop; scan each (prod first).
_REPO_ROOT = Path(__file__).resolve().parent.parent
_TRANSIENT_DIRS = [
    Path("/home/vinay/security-ops-platform/data/transient"),
    Path("/home/vinay/security-ops-platform-dev/data/transient"),
    _REPO_ROOT / "data" / "transient",
]
_EXPORT_GLOB = "cgr-by-vulnerabilities-cluster_*.xlsx"
_DATA_SHEET = "cgr-by-vulnerabilities-cluster_"

# Write the rolled-up store to the DEV worktree (data-isolated).
VULN_DB_PATH = _REPO_ROOT / "data" / "transient" / "cgr_findings.db"
DEFAULT_TRIAGE_DB = "/home/vinay/security-ops-platform-dev/data/transient/cve_triage_results.db"
# Cached EAI application inventory (snapshot of EAI prod dbo.V_APP_INFO).
EAI_DB_PATH = _REPO_ROOT / "data" / "transient" / "eai_app_info.db"

# Worst-first severity ranking. Unknown severities rank 0.
_SEVERITY_RANK = {"low": 1, "medium": 2, "moderate": 2, "high": 3, "critical": 4}
_RANK_SEVERITY = {1: "low", 2: "medium", 3: "high", 4: "critical"}
_HIGH_RANK = 3  # high or critical

_LIST_CAP = 40  # cap distinct packageName / eaiCode lists per CVE

# --- composite risk model (war-room weights) ------------------------------- #
_W_KEV = 5.0
_W_EXTERNAL = 2.5
_W_PREAUTH = 1.5
_W_CVSS = 0.3  # applied to (cvss_mean / 10.0)

# Reachability multiplier on the composite (war-room: "reachability is a key
# factor in the final score", 2026-06-03). This is the *code-path* axis — is the
# vulnerable code actually exercisable, or is it a base-image OS package the
# scanner sees but no app path reaches (the glibc-scanf case). The complementary
# *internet-exposure* axis is already the additive _W_EXTERNAL term, so it is NOT
# repeated here (no double-count). Derived from Prisma's base-image attribution:
#   all_from_base  -> entirely a base-image OS pkg, code path unconfirmed (pending-image)
#   any_from_base  -> mixed base + app layers (present-conditional)
#   neither        -> purely application-layer, scanner-visible in the app (present-confirmed)
_REACH_FACTOR = {
    "present-confirmed": 1.0,
    "present-conditional": 0.7,
    "pending-image": 0.4,
    "unknown": 0.4,
}


def _reach_level_from_base(any_from_base: bool, all_from_base: bool) -> str:
    """Map Prisma base-image attribution to a code-path reachability label."""
    if all_from_base:
        return "pending-image"
    if any_from_base:
        return "present-conditional"
    return "present-confirmed"


# TLS/CURL exemption: word-ish, case-insensitive (the war room exempted these).
_EXEMPT_RE = re.compile(r"\b(tls|curl|libcurl|openssl)\b", re.IGNORECASE)

# Cap on how many top candidates get a live NVD/CVE.org CVSS-vector fetch for
# the pre_auth term (only the +1.5 term; bound so we don't hammer NVD).
_PREAUTH_FETCH_CAP = 150

# How many EAI apps to attach to each scored CVE row (for the table).
_EAI_ATTACH_CAP = 6


# --------------------------- value parsing helpers ------------------------- #
def _sev_rank(sev) -> int:
    if sev is None:
        return 0
    return _SEVERITY_RANK.get(str(sev).strip().lower(), 0)


def _num(val) -> Optional[float]:
    """Parse a numeric cell that may arrive as int, float, str, or blank."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _as_int(val) -> Optional[int]:
    f = _num(val)
    return int(f) if f is not None else None


def _as_bool(val) -> bool:
    """Truthiness for isFromBase (bool, 1/0, 'true'/'false')."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    return str(val).strip().lower() in ("true", "1", "yes", "y")


def _epoch_to_iso(val) -> Optional[str]:
    """Convert a Unix epoch-seconds cell to an ISO date; None for blank/0."""
    i = _as_int(val)
    if not i or i <= 0:
        return None
    try:
        return datetime.fromtimestamp(i, tz=timezone.utc).strftime("%Y-%m-%d")
    except (OverflowError, OSError, ValueError):
        return None


def _is_open_status(status: Optional[str]) -> bool:
    """A status that indicates the vuln is still open/affected (not fixed)."""
    if not status:
        return False
    s = status.strip().lower()
    if s.startswith("fixed"):
        return False
    return any(k in s for k in ("open", "affected", "needed", "vulnerable")) or s == ""


# ------------------------------ export discovery --------------------------- #
def latest_export() -> Optional[Path]:
    """Newest cgr-by-vulnerabilities-cluster_*.xlsx across both worktrees."""
    candidates: Dict[Path, Path] = {}
    for d in _TRANSIENT_DIRS:
        try:
            if not d.is_dir():
                continue
            for p in d.glob(_EXPORT_GLOB):
                candidates[p.resolve()] = p
        except OSError as e:
            logger.debug("scan failed for %s: %s", d, e)
    if not candidates:
        logger.warning("no CGR export found in %s", _TRANSIENT_DIRS)
        return None
    return max(candidates.values(), key=lambda p: p.stat().st_mtime)


# --------------------------------- storage --------------------------------- #
def _connect() -> sqlite3.Connection:
    VULN_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(VULN_DB_PATH))
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS cgr_cve (
            cve_id          TEXT PRIMARY KEY,
            deployments     INTEGER,
            distinct_images INTEGER,
            distinct_eai    INTEGER,
            max_cvss        REAL,
            severity        TEXT,
            any_from_base   INTEGER,
            all_from_base   INTEGER,
            min_grace       INTEGER,
            worst_status    TEXT,
            packages        TEXT,
            eai_codes       TEXT,
            published       TEXT,
            fix_date        TEXT,
            source_file     TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_cgr_cve_sev ON cgr_cve(severity);

        CREATE TABLE IF NOT EXISTS cgr_cluster (
            eai_code         TEXT PRIMARY KEY,
            distinct_cve     INTEGER,
            distinct_high_cve INTEGER,
            total_findings   INTEGER,
            distinct_images  INTEGER,
            min_grace        INTEGER,
            source_file      TEXT
        );

        CREATE TABLE IF NOT EXISTS cgr_meta (
            source_file       TEXT PRIMARY KEY,
            ingested_at       TEXT,
            total_rows        INTEGER,
            distinct_cves     INTEGER,
            distinct_clusters INTEGER,
            distinct_images   INTEGER
        );
        """
    )
    return conn


def _none_eai(code: Optional[str]) -> bool:
    """True when an eaiCode should not count as a real cluster."""
    if not code:
        return True
    return str(code).strip().lower() in ("unknown-eaicode", "none", "", "unknown")


def ingest_export(path: Optional[Path] = None) -> dict:
    """Stream-parse a CGR export, aggregate, and rebuild all three tables.

    Memory-safe: rows are streamed via openpyxl read_only + iter_rows and folded
    into bounded per-CVE / per-cluster aggregate dicts as we go (no row lists are
    retained; distinct package / eai-code sets are capped at _LIST_CAP). Image
    counts use per-key sets of _id hashes -- the only unbounded structures, which
    is unavoidable for an exact distinct count but is just strings.

    Idempotent: the file is a single snapshot, so we DELETE this source_file's
    rows and re-insert. Returns {rows, distinct_cves, distinct_clusters,
    distinct_images}.
    """
    if path is None:
        path = latest_export()
    if path is None:
        logger.error("ingest_export: no export available")
        return {"rows": 0, "distinct_cves": 0, "distinct_clusters": 0, "distinct_images": 0}
    path = Path(path)
    if not path.exists():
        logger.error("ingest_export: file does not exist: %s", path)
        return {"rows": 0, "distinct_cves": 0, "distinct_clusters": 0, "distinct_images": 0}

    source_file = path.name
    try:
        ingested_at = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%dT%H:%M:%S")
    except OSError:
        ingested_at = time.strftime("%Y-%m-%dT%H:%M:%S")

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        logger.error("ingest_export: failed to open %s: %s", path, e)
        return {"rows": 0, "distinct_cves": 0, "distinct_clusters": 0, "distinct_images": 0}

    # Per-CVE aggregate accumulators.
    cve_agg: Dict[str, dict] = {}
    # Per-cluster aggregate accumulators.
    clu_agg: Dict[str, dict] = {}
    all_images: set = set()
    total_rows = 0

    try:
        ws = wb[_DATA_SHEET] if _DATA_SHEET in wb.sheetnames else wb.worksheets[-1]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            logger.warning("ingest_export: empty sheet in %s", path)
            wb.close()
            return {"rows": 0, "distinct_cves": 0, "distinct_clusters": 0, "distinct_images": 0}
        col = {h: i for i, h in enumerate(header) if h is not None}

        def cell(row, name):
            i = col.get(name)
            return row[i] if i is not None and i < len(row) else None

        for row in rows:
            if row is None or all(v is None for v in row):
                continue
            cve = cell(row, "cve")
            if cve is None:
                continue
            cve = str(cve).strip().upper()
            if not cve:
                continue

            total_rows += 1
            image = cell(row, "_id")
            image = str(image).strip() if image is not None else None
            if image:
                all_images.add(image)
            eai_raw = cell(row, "eaiCode")
            eai = str(eai_raw).strip() if eai_raw is not None else None
            real_eai = eai if not _none_eai(eai) else None

            cvss = _num(cell(row, "cvss"))
            srank = _sev_rank(cell(row, "severity"))
            from_base = _as_bool(cell(row, "isFromBase"))
            grace = _as_int(cell(row, "grace"))
            status = cell(row, "status")
            status = str(status).strip() if status is not None else None
            pkg = cell(row, "packageName")
            pkg = str(pkg).strip() if pkg is not None else None
            published = _epoch_to_iso(cell(row, "published"))
            fix_date = _epoch_to_iso(cell(row, "fixDate"))

            # ---- per-CVE fold ----
            a = cve_agg.get(cve)
            if a is None:
                a = {
                    "deployments": 0,
                    "images": set(),
                    "eais": set(),
                    "max_cvss": None,
                    "sev_rank": 0,
                    "any_from_base": False,
                    "all_from_base": True,
                    "min_grace": None,
                    "open_status": None,
                    "any_status": None,
                    "packages": set(),
                    "published": None,
                    "fix_date": None,
                }
                cve_agg[cve] = a
            a["deployments"] += 1
            if image:
                a["images"].add(image)
            if real_eai and (real_eai in a["eais"] or len(a["eais"]) < _LIST_CAP):
                a["eais"].add(real_eai)
            if cvss is not None and (a["max_cvss"] is None or cvss > a["max_cvss"]):
                a["max_cvss"] = cvss
            if srank > a["sev_rank"]:
                a["sev_rank"] = srank
            a["any_from_base"] = a["any_from_base"] or from_base
            a["all_from_base"] = a["all_from_base"] and from_base
            if grace is not None and (a["min_grace"] is None or grace < a["min_grace"]):
                a["min_grace"] = grace
            if status:
                if a["any_status"] is None:
                    a["any_status"] = status
                if a["open_status"] is None and _is_open_status(status):
                    a["open_status"] = status
            if pkg and len(a["packages"]) < _LIST_CAP:
                a["packages"].add(pkg)
            if published and a["published"] is None:
                a["published"] = published
            if fix_date and a["fix_date"] is None:
                a["fix_date"] = fix_date

            # ---- per-cluster fold ----
            ckey = eai if eai else "unknown-eaiCode"
            c = clu_agg.get(ckey)
            if c is None:
                c = {
                    "cves": set(),
                    "high_cves": set(),
                    "total": 0,
                    "images": set(),
                    "min_grace": None,
                }
                clu_agg[ckey] = c
            c["total"] += 1
            c["cves"].add(cve)
            if srank >= _HIGH_RANK:
                c["high_cves"].add(cve)
            if image:
                c["images"].add(image)
            if grace is not None and (c["min_grace"] is None or grace < c["min_grace"]):
                c["min_grace"] = grace
    finally:
        wb.close()

    # --------------------------- write rolled-up tables --------------------- #
    conn = _connect()
    try:
        for tbl in ("cgr_cve", "cgr_cluster", "cgr_meta"):
            conn.execute(f"DELETE FROM {tbl} WHERE source_file = ?", (source_file,))

        cve_rows = []
        for cve, a in cve_agg.items():
            sev = _RANK_SEVERITY.get(a["sev_rank"])
            worst_status = a["open_status"] or a["any_status"]
            cve_rows.append((
                cve,
                a["deployments"],
                len(a["images"]),
                len(a["eais"]),
                a["max_cvss"],
                sev,
                1 if a["any_from_base"] else 0,
                1 if (a["all_from_base"] and a["deployments"] > 0) else 0,
                a["min_grace"],
                worst_status,
                json.dumps(sorted(a["packages"])[:_LIST_CAP]),
                json.dumps(sorted(a["eais"])[:_LIST_CAP]),
                a["published"],
                a["fix_date"],
                source_file,
            ))
        conn.executemany(
            "INSERT OR REPLACE INTO cgr_cve (cve_id, deployments, distinct_images, "
            "distinct_eai, max_cvss, severity, any_from_base, all_from_base, "
            "min_grace, worst_status, packages, eai_codes, published, fix_date, "
            "source_file) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            cve_rows,
        )

        clu_rows = []
        for eai, c in clu_agg.items():
            clu_rows.append((
                eai,
                len(c["cves"]),
                len(c["high_cves"]),
                c["total"],
                len(c["images"]),
                c["min_grace"],
                source_file,
            ))
        conn.executemany(
            "INSERT OR REPLACE INTO cgr_cluster (eai_code, distinct_cve, "
            "distinct_high_cve, total_findings, distinct_images, min_grace, "
            "source_file) VALUES (?,?,?,?,?,?,?)",
            clu_rows,
        )

        distinct_cves = len(cve_agg)
        distinct_clusters = len(clu_agg)
        distinct_images = len(all_images)
        conn.execute(
            "INSERT OR REPLACE INTO cgr_meta (source_file, ingested_at, total_rows, "
            "distinct_cves, distinct_clusters, distinct_images) VALUES (?,?,?,?,?,?)",
            (source_file, ingested_at, total_rows, distinct_cves, distinct_clusters,
             distinct_images),
        )
        conn.commit()
    finally:
        conn.close()

    logger.info(
        "ingested %d rows / %d CVEs / %d clusters / %d images from %s",
        total_rows, distinct_cves, distinct_clusters, distinct_images, source_file,
    )
    return {
        "rows": total_rows,
        "distinct_cves": distinct_cves,
        "distinct_clusters": distinct_clusters,
        "distinct_images": distinct_images,
    }


# ----------------------------- query / aggregate --------------------------- #
def _ro_connect() -> Optional[sqlite3.Connection]:
    if not VULN_DB_PATH.exists():
        return None
    conn = sqlite3.connect(f"file:{VULN_DB_PATH}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def cgr_cves() -> set:
    """Distinct CVE IDs in cgr_cve."""
    conn = _ro_connect()
    if conn is None:
        return set()
    try:
        return {r[0] for r in conn.execute("SELECT cve_id FROM cgr_cve")}
    except sqlite3.OperationalError:
        return set()
    finally:
        conn.close()


def _ranked_cgr_only(triage_db_path: str, limit: int) -> List[dict]:
    """cgr-only CVEs ranked by (severity desc, max_cvss desc, deployments desc)."""
    conn = _ro_connect()
    if conn is None:
        return []
    try:
        cgr_rows = {
            r["cve_id"]: r
            for r in conn.execute("SELECT * FROM cgr_cve")
        }
    except sqlite3.OperationalError:
        conn.close()
        return []
    conn.close()

    triaged = _triage_cves(triage_db_path)
    cgr_only = [r for cid, r in cgr_rows.items() if cid not in triaged]

    def keyfn(r):
        return (
            _sev_rank(r["severity"]),
            r["max_cvss"] if r["max_cvss"] is not None else -1.0,
            r["deployments"] or 0,
        )

    cgr_only.sort(key=keyfn, reverse=True)
    out = []
    for r in cgr_only[:limit]:
        try:
            packages = json.loads(r["packages"]) if r["packages"] else []
        except (TypeError, ValueError):
            packages = []
        out.append({
            "cve_id": r["cve_id"],
            "severity": r["severity"],
            "max_cvss": r["max_cvss"],
            "deployments": r["deployments"],
            "distinct_eai": r["distinct_eai"],
            "any_from_base": bool(r["any_from_base"]),
            "min_grace": r["min_grace"],
            "packages": packages,
        })
    return out


def _triage_cves(triage_db_path: str) -> set:
    p = Path(triage_db_path)
    if not p.exists():
        logger.warning("triage DB absent at %s", triage_db_path)
        return set()
    conn = sqlite3.connect(f"file:{p}?mode=ro", uri=True)
    try:
        return {
            str(r[0]).upper()
            for r in conn.execute("SELECT cve_id FROM triage")
            if r[0]
        }
    except sqlite3.OperationalError as e:
        logger.warning("triage DB unreadable: %s", e)
        return set()
    finally:
        conn.close()


# ---------------------- reconciliation vs our triage ----------------------- #
def reconcile(triage_db_path: str = DEFAULT_TRIAGE_DB, top_limit: int = 200) -> dict:
    """Compare the CGR CVE set against our triaged CVE set.

    cgr_only = CVEs Prisma flagged that aren't in our triage set -- candidate
    coverage gaps. Guards gracefully when the triage DB is absent.
    """
    cgr = cgr_cves()
    triaged = _triage_cves(triage_db_path)

    overlap = cgr & triaged
    cgr_only = cgr - triaged
    triage_only = triaged - cgr

    # cgr_only high/critical count.
    cgr_only_high = 0
    conn = _ro_connect()
    if conn is not None:
        try:
            for r in conn.execute("SELECT cve_id, severity FROM cgr_cve"):
                if r[0] in cgr_only and _sev_rank(r[1]) >= _HIGH_RANK:
                    cgr_only_high += 1
        except sqlite3.OperationalError:
            pass
        finally:
            conn.close()

    return {
        "overlap": len(overlap),
        "cgr_only": len(cgr_only),
        "triage_only": len(triage_only),
        "cgr_only_high": cgr_only_high,
        "cgr_only_top": _ranked_cgr_only(triage_db_path, top_limit),
    }


# ----------------------------- web read accessors -------------------------- #
def summary() -> dict:
    """Totals from cgr_meta plus severity distribution and from-base split."""
    conn = _ro_connect()
    if conn is None:
        return {"available": False}
    try:
        meta = conn.execute(
            "SELECT source_file, ingested_at, total_rows, distinct_cves, "
            "distinct_clusters, distinct_images FROM cgr_meta "
            "ORDER BY ingested_at DESC LIMIT 1"
        ).fetchone()
        if meta is None:
            return {"available": False}

        sev_dist = {}
        for r in conn.execute(
            "SELECT severity, COUNT(*) FROM cgr_cve GROUP BY severity"
        ):
            sev_dist[r[0] or "unknown"] = r[1]

        any_base = conn.execute(
            "SELECT COUNT(*) FROM cgr_cve WHERE any_from_base = 1"
        ).fetchone()[0]
        not_base = conn.execute(
            "SELECT COUNT(*) FROM cgr_cve WHERE any_from_base = 0"
        ).fetchone()[0]
        overdue = conn.execute(
            "SELECT COUNT(*) FROM cgr_cve WHERE min_grace IS NOT NULL AND min_grace < 0"
        ).fetchone()[0]

        return {
            "available": True,
            "source_file": meta["source_file"],
            "ingested_at": meta["ingested_at"],
            "total_rows": meta["total_rows"],
            "distinct_cves": meta["distinct_cves"],
            "distinct_clusters": meta["distinct_clusters"],
            "distinct_images": meta["distinct_images"],
            "severity_distribution": sev_dist,
            "cves_from_base": any_base,
            "cves_not_from_base": not_base,
            "cves_overdue": overdue,
        }
    except sqlite3.OperationalError as e:
        logger.warning("summary: %s", e)
        return {"available": False}
    finally:
        conn.close()


def cluster_rollup(limit: int = 80) -> List[dict]:
    """Top eaiCodes by distinct_high_cve then distinct_cve.

    For the EAI-joined (internet-facing-first) variant used by the leadership
    view, see :func:`cgr_cluster_eai`.
    """
    conn = _ro_connect()
    if conn is None:
        return []
    try:
        rows = conn.execute(
            "SELECT eai_code, distinct_cve, distinct_high_cve, total_findings, "
            "distinct_images, min_grace FROM cgr_cluster "
            "ORDER BY distinct_high_cve DESC, distinct_cve DESC LIMIT ?",
            (limit,),
        ).fetchall()
        return [dict(r) for r in rows]
    except sqlite3.OperationalError:
        return []
    finally:
        conn.close()


def cve_gaps(limit: int = 200, triage_db_path: str = DEFAULT_TRIAGE_DB) -> List[dict]:
    """CVEs the scanner flagged that our triage hasn't processed (coverage gap)."""
    return _ranked_cgr_only(triage_db_path, limit)


# ============================== EAI inventory ============================== #
_EAI_COLUMNS = (
    "eai_id", "app_name", "app_long", "status", "lob", "critical_ind",
    "app_class", "internet_facing", "accessibility", "cloud", "pii", "phi",
    "crit_metal", "crit_score", "prod_urls", "cio", "officer",
)


def _eai_ro_connect() -> Optional[sqlite3.Connection]:
    """Read-only connection to the cached EAI inventory (None if absent)."""
    if not EAI_DB_PATH.exists():
        return None
    conn = sqlite3.connect(f"file:{EAI_DB_PATH}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def load_eai_map() -> Dict[str, dict]:
    """Return ``{eai_id: {col: val, ...}}`` for the whole cached inventory.

    eai_id is normalized to a stripped string so it joins cleanly against the
    eaiCode strings stored in ``cgr_cve.eai_codes``.
    """
    conn = _eai_ro_connect()
    if conn is None:
        logger.warning("EAI inventory cache absent at %s", EAI_DB_PATH)
        return {}
    try:
        out: Dict[str, dict] = {}
        for r in conn.execute("SELECT * FROM eai_app"):
            d = dict(r)
            key = str(d.get("eai_id") or "").strip()
            if key:
                out[key] = d
        return out
    except sqlite3.OperationalError as e:
        logger.warning("load_eai_map: %s", e)
        return {}
    finally:
        conn.close()


def eai_info(eai_code) -> Optional[dict]:
    """Look up a single EAI application by its code (string-normalized)."""
    if eai_code is None:
        return None
    key = str(eai_code).strip()
    if not key:
        return None
    conn = _eai_ro_connect()
    if conn is None:
        return None
    try:
        r = conn.execute(
            "SELECT * FROM eai_app WHERE eai_id = ?", (key,)
        ).fetchone()
        return dict(r) if r is not None else None
    except sqlite3.OperationalError:
        return None
    finally:
        conn.close()


def _is_external(app: Optional[dict]) -> bool:
    return bool(app) and str(app.get("internet_facing") or "").strip().lower() == "yes"


def external_facing_map() -> Dict[str, bool]:
    """{CVE -> is any carrying app internet-facing per EAI}.

    Joins each ``cgr_cve``'s eaiCodes to the EAI inventory and returns the real
    ``V_APP_INFO.Internet_Facing_Indicator`` signal. A CVE is **absent** from the
    map when none of its eaiCodes resolve to an EAI record (so callers treat it as
    "exposure unknown", not "internal"). Returns ``{}`` if the cgr cache is absent
    — lets consumers (e.g. cve_triage.enrich_priorities) degrade gracefully to
    code-path-only reachability.
    """
    conn = _ro_connect()
    if conn is None:
        return {}
    try:
        rows = conn.execute("SELECT cve_id, eai_codes FROM cgr_cve").fetchall()
    except sqlite3.OperationalError:
        conn.close()
        return {}
    conn.close()
    eai_map = load_eai_map()
    out: Dict[str, bool] = {}
    for r in rows:
        try:
            codes = json.loads(r["eai_codes"]) if r["eai_codes"] else []
        except (TypeError, ValueError):
            codes = []
        apps = [eai_map.get(str(c).strip()) for c in codes]
        apps = [a for a in apps if a is not None]
        if not apps:
            continue  # no EAI linkage -> leave as exposure-unknown
        out[str(r["cve_id"]).upper()] = any(_is_external(a) for a in apps)
    return out


def _crit_score_num(app: Optional[dict]) -> float:
    """Numeric criticality score for ordering 'top app' picks (0 if unknown)."""
    if not app:
        return 0.0
    v = _num(app.get("crit_score"))
    if v is not None:
        return v
    # Fall back to critical_ind Y/N if no numeric score.
    return 5.0 if str(app.get("critical_ind") or "").strip().upper() == "Y" else 0.0


# ----------------------- optional live EAI re-pull ------------------------- #
# Map snapshot columns -> EAI prod dbo.V_APP_INFO source columns. The view is
# wide (~190 cols); names below are best-effort and only used on-tunnel. Any
# missing column degrades to '' rather than failing the whole pull.
_V_APP_INFO_MAP = {
    "eai_id": "EAI_Application_ID",
    "app_name": "Application_Short_Name",
    "app_long": "Application_Long_Name",
    "status": "Application_Status",
    "lob": "Line_of_Business_Name",
    "critical_ind": "Critical_Application_Indicator",
    "app_class": "Application_Class",
    "internet_facing": "Internet_Facing_Indicator",
    "accessibility": "Application_Accessibility",
    "cloud": "Cloud_Hosted_Indicator",
    "pii": "PII_Indicator",
    "phi": "PHI_Indicator",
    "crit_metal": "Criticality_Metal",
    "crit_score": "Criticality_Score",
    "prod_urls": "Production_URLs",
    "cio": "CIO_Full_Name",
    "officer": "Officer_Full_Name",
}


def refresh_eai_cache(eai_codes: Optional[List[str]] = None) -> dict:
    """OPTIONAL: re-pull the EAI inventory from EAI prod into the local cache.

    Reads ``dbo.V_APP_INFO`` via :func:`src.utils.eai_db.connect` (env='prod'),
    which requires the studio1 reverse tunnel + ``EAI_PROD_HOST/PORT`` env on
    the isolated lab net. Fails gracefully (returns ``{"refreshed": 0,
    "error": ...}``) when off-tunnel or creds are missing. Default flows READ
    the cache; this only exists to make the snapshot reproducible.

    ``eai_codes`` optionally restricts the refresh to a subset (used to top up
    holes); ``None`` re-pulls the whole view.
    """
    try:
        from src.utils.eai_db import connect as eai_connect
    except Exception as e:  # noqa: BLE001 - import may pull config that's absent
        logger.warning("refresh_eai_cache: eai_db import failed: %s", e)
        return {"refreshed": 0, "error": f"import: {e}"}

    # Project only the columns we mirror; alias to our snapshot column names.
    select_cols = ", ".join(
        f"{src} AS {dst}" for dst, src in _V_APP_INFO_MAP.items()
    )
    sql = f"SELECT {select_cols} FROM dbo.V_APP_INFO"
    want = {str(c).strip() for c in eai_codes} if eai_codes else None

    rows: List[tuple] = []
    try:
        with eai_connect(env="prod", timeout=120) as cn:
            cur = cn.cursor()
            cur.execute(sql)
            colnames = [d[0] for d in cur.description]
            idx = {name: i for i, name in enumerate(colnames)}
            for raw in cur.fetchall():
                rec = {}
                for dst in _EAI_COLUMNS:
                    i = idx.get(dst)
                    val = raw[i] if i is not None and i < len(raw) else None
                    rec[dst] = "" if val is None else str(val).strip()
                eid = rec.get("eai_id", "")
                if not eid:
                    continue
                if want is not None and eid not in want:
                    continue
                rows.append(tuple(rec[c] for c in _EAI_COLUMNS))
    except Exception as e:  # noqa: BLE001 - off-tunnel / auth / driver failures
        logger.warning("refresh_eai_cache: EAI prod pull failed: %s", e)
        return {"refreshed": 0, "error": str(e)}

    EAI_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(EAI_DB_PATH))
    try:
        coldefs = ", ".join(
            f"{c} TEXT PRIMARY KEY" if c == "eai_id" else f"{c} TEXT"
            for c in _EAI_COLUMNS
        )
        conn.execute(f"CREATE TABLE IF NOT EXISTS eai_app ({coldefs})")
        placeholders = ",".join("?" for _ in _EAI_COLUMNS)
        conn.executemany(
            f"INSERT OR REPLACE INTO eai_app ({', '.join(_EAI_COLUMNS)}) "
            f"VALUES ({placeholders})",
            rows,
        )
        conn.commit()
    finally:
        conn.close()
    logger.info("refresh_eai_cache: refreshed %d EAI rows", len(rows))
    return {"refreshed": len(rows), "error": None}


# ============================ composite scoring =========================== #
def _enrichment_map(triage_db_path: str) -> Dict[str, dict]:
    """``{cve_id: {pre_auth, cvss_vector, kev}}`` from the triage enrichment table."""
    p = Path(triage_db_path)
    if not p.exists():
        return {}
    conn = sqlite3.connect(f"file:{p}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    try:
        out: Dict[str, dict] = {}
        for r in conn.execute(
            "SELECT cve_id, pre_auth, cvss_vector, kev FROM enrichment"
        ):
            cid = str(r["cve_id"] or "").upper()
            if cid:
                out[cid] = dict(r)
        return out
    except sqlite3.OperationalError as e:
        logger.warning("_enrichment_map: %s", e)
        return {}
    finally:
        conn.close()


def _cached_cvss_vector(cve_id: str) -> Optional[str]:
    """CVSS vector from the NVD / CVE.org on-disk caches WITHOUT a network call."""
    try:
        from services import nvd
        cached = nvd._load_cached(cve_id)  # noqa: SLF001 - intentional cache-only read
        if cached and cached.get("severity") and cached["severity"].get("vector"):
            return cached["severity"]["vector"]
    except Exception:  # noqa: BLE001
        pass
    try:
        from services import cve_org
        cached = cve_org._load_cached(cve_id)  # noqa: SLF001
        if cached and cached.get("found") and cached.get("severity"):
            v = cached["severity"].get("vector")
            if v:
                return v
    except Exception:  # noqa: BLE001
        pass
    return None


def _live_cvss_vector(cve_id: str) -> Optional[str]:
    """Fetch the CVSS vector via NVD then CVE.org (both cache to disk)."""
    try:
        from services import nvd
        rec = nvd.get_cve(cve_id)
        if rec and rec.get("severity") and rec["severity"].get("vector"):
            return rec["severity"]["vector"]
    except Exception as e:  # noqa: BLE001
        logger.debug("nvd vector fetch failed for %s: %s", cve_id, e)
    try:
        from services import cve_org
        rec = cve_org.get_cve_org(cve_id)
        if rec and rec.get("severity") and rec["severity"].get("vector"):
            return rec["severity"]["vector"]
    except Exception as e:  # noqa: BLE001
        logger.debug("cve.org vector fetch failed for %s: %s", cve_id, e)
    return None


def _cached_description(cve_id: str) -> Optional[str]:
    """CVE description from NVD / CVE.org caches WITHOUT a network call."""
    try:
        from services import nvd
        cached = nvd._load_cached(cve_id)  # noqa: SLF001
        if cached and cached.get("description"):
            return cached["description"]
    except Exception:  # noqa: BLE001
        pass
    try:
        from services import cve_org
        cached = cve_org._load_cached(cve_id)  # noqa: SLF001
        if cached and cached.get("found") and cached.get("description"):
            return cached["description"]
    except Exception:  # noqa: BLE001
        pass
    return None


def _exempt_signal(cve_id: str, packages: List[str]) -> bool:
    """TLS/CURL exemption.

    Signal precedence (bounded -- NO network for 5.5k CVEs):
      1. CVE description from the NVD/CVE.org on-disk cache, if present.
      2. Otherwise the CGR ``packages`` field (the only always-on-disk signal):
         a package name matching tls/curl/openssl flags the CVE exempt.
    """
    desc = _cached_description(cve_id)
    if desc and _EXEMPT_RE.search(desc):
        return True
    for pkg in packages or []:
        if _EXEMPT_RE.search(str(pkg)):
            return True
    return False


def _connect_scored() -> sqlite3.Connection:
    VULN_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(VULN_DB_PATH))
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS cgr_scored (
            cve_id          TEXT PRIMARY KEY,
            composite       REAL,
            kev             INTEGER,
            external_facing INTEGER,
            pre_auth        INTEGER,
            pre_auth_known  INTEGER,
            cvss            REAL,
            exempt          INTEGER,
            top_eai         TEXT,
            eai_apps        TEXT,
            deployments     INTEGER,
            severity        TEXT,
            min_grace       INTEGER,
            in_triage       INTEGER,
            reach_factor    REAL,
            reach_level     TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_cgr_scored_comp
            ON cgr_scored(composite DESC);
        CREATE INDEX IF NOT EXISTS idx_cgr_scored_exempt
            ON cgr_scored(exempt);
        """
    )
    # Lightweight migration: add reachability columns to a pre-existing table
    # (the composite became reachability-adjusted on 2026-06-03).
    cols = {r[1] for r in conn.execute("PRAGMA table_info(cgr_scored)")}
    if "reach_factor" not in cols:
        conn.execute("ALTER TABLE cgr_scored ADD COLUMN reach_factor REAL")
    if "reach_level" not in cols:
        conn.execute("ALTER TABLE cgr_scored ADD COLUMN reach_level TEXT")
    conn.commit()
    return conn


def score_all(
    fetch_preauth_cap: int = _PREAUTH_FETCH_CAP,
    triage_db_path: str = DEFAULT_TRIAGE_DB,
) -> dict:
    """Compute the EAI-enriched composite score for every ``cgr_cve`` row.

    Idempotently rebuilds the ``cgr_scored`` table. Returns tier/exemption
    stats plus how many pre_auth vectors were live-fetched vs left unknown.
    """
    conn = _ro_connect()
    if conn is None:
        logger.error("score_all: cgr_findings.db absent -- ingest first")
        return {"scored": 0, "error": "no cgr_findings.db"}
    try:
        cgr_rows = [dict(r) for r in conn.execute("SELECT * FROM cgr_cve")]
    except sqlite3.OperationalError as e:
        conn.close()
        return {"scored": 0, "error": str(e)}
    conn.close()

    eai_map = load_eai_map()
    enrich = _enrichment_map(triage_db_path)
    triaged = _triage_cves(triage_db_path)

    try:
        from services.cve_priority import is_kev
    except Exception as e:  # noqa: BLE001
        logger.warning("score_all: is_kev import failed (%s); KEV=0 for all", e)
        is_kev = lambda _cid: False  # noqa: E731

    # ---- per-CVE base computation (no network yet) ----
    records: Dict[str, dict] = {}
    preauth_candidates: List[tuple] = []  # (base_no_preauth, cve_id)

    for r in cgr_rows:
        cve = str(r["cve_id"]).upper()
        try:
            eai_codes = json.loads(r["eai_codes"]) if r["eai_codes"] else []
        except (TypeError, ValueError):
            eai_codes = []
        try:
            packages = json.loads(r["packages"]) if r["packages"] else []
        except (TypeError, ValueError):
            packages = []

        # Join eaiCodes -> EAI inventory.
        apps = []
        for code in eai_codes:
            app = eai_map.get(str(code).strip())
            if app is not None:
                apps.append(app)
        external = any(_is_external(a) for a in apps)

        # KEV.
        try:
            kev = bool(is_kev(cve))
        except Exception:  # noqa: BLE001
            kev = False

        cvss = r["max_cvss"] if r["max_cvss"] is not None else 0.0

        # pre_auth: prefer the enrichment table.
        pre_auth = 0
        pre_auth_known = 0
        en = enrich.get(cve)
        if en is not None and en.get("pre_auth") is not None:
            pre_auth = 1 if en["pre_auth"] else 0
            pre_auth_known = 1
        else:
            # Try cache-only vector now (free); live fetch deferred to the cap.
            vec = (en or {}).get("cvss_vector") or _cached_cvss_vector(cve)
            if vec:
                from services.cve_priority import pre_auth_from_cvss_vector
                pre_auth = 1 if pre_auth_from_cvss_vector(vec) else 0
                pre_auth_known = 1

        exempt = 1 if _exempt_signal(cve, packages) else 0

        # Code-path reachability multiplier from Prisma base-image attribution.
        reach_level = _reach_level_from_base(
            _as_bool(r.get("any_from_base")), _as_bool(r.get("all_from_base"))
        )
        reach_factor = _REACH_FACTOR.get(reach_level, 1.0)

        # Pick the representative app: prefer an internet-facing one, else the
        # highest-criticality app.
        top_eai = None
        ext_apps = [a for a in apps if _is_external(a)]
        pool = ext_apps if ext_apps else apps
        if pool:
            best = max(pool, key=_crit_score_num)
            top_eai = (best.get("app_name") or "").strip() or best.get("eai_id")

        # Attach a bounded, external-first set of apps for the table.
        apps_sorted = sorted(
            apps,
            key=lambda a: (1 if _is_external(a) else 0, _crit_score_num(a)),
            reverse=True,
        )
        eai_apps_json = json.dumps([
            {
                "eai": a.get("eai_id"),
                "app_name": (a.get("app_name") or "").strip(),
                "internet_facing": (a.get("internet_facing") or "").strip(),
                "critical_ind": (a.get("critical_ind") or "").strip(),
            }
            for a in apps_sorted[:_EAI_ATTACH_CAP]
        ])

        rec = {
            "cve_id": cve,
            "kev": 1 if kev else 0,
            "external_facing": 1 if external else 0,
            "pre_auth": pre_auth,
            "pre_auth_known": pre_auth_known,
            "cvss": float(cvss),
            "exempt": exempt,
            "top_eai": top_eai,
            "eai_apps": eai_apps_json,
            "deployments": r["deployments"] or 0,
            "severity": r["severity"],
            "min_grace": r["min_grace"],
            "in_triage": 1 if cve in triaged else 0,
            "reach_factor": reach_factor,
            "reach_level": reach_level,
        }
        records[cve] = rec

        if not pre_auth_known:
            # Rank live-fetch candidates by their reach-adjusted base so the
            # bounded NVD budget is spent on the more-reachable CVEs first.
            base_no_pa = (
                (_W_KEV if kev else 0.0)
                + (_W_EXTERNAL if external else 0.0)
                + (float(cvss) / 10.0) * _W_CVSS
            ) * reach_factor
            preauth_candidates.append((base_no_pa, cve))

    # ---- bounded live pre_auth fetch for the top unknown candidates ----
    preauth_candidates.sort(reverse=True)
    fetched = 0
    pre_auth_found = 0
    cap = max(0, int(fetch_preauth_cap))
    for _base, cve in preauth_candidates[:cap]:
        try:
            from services.nvd import CVE_RE
            if not CVE_RE.match(cve):
                continue
        except Exception:  # noqa: BLE001
            pass
        vec = _live_cvss_vector(cve)
        fetched += 1
        if vec:
            from services.cve_priority import pre_auth_from_cvss_vector
            pa = pre_auth_from_cvss_vector(vec)
            records[cve]["pre_auth"] = 1 if pa else 0
            records[cve]["pre_auth_known"] = 1
            if pa:
                pre_auth_found += 1
    left_unknown = max(0, len(preauth_candidates) - fetched)

    # ---- finalize composite + write table ----
    rows_out = []
    tiers = {"5+": 0, "2.5-5": 0, "0-2.5": 0}
    for cve, rec in records.items():
        composite = (
            (_W_KEV if rec["kev"] else 0.0)
            + (_W_EXTERNAL if rec["external_facing"] else 0.0)
            + (_W_PREAUTH if rec["pre_auth"] else 0.0)
            + (rec["cvss"] / 10.0) * _W_CVSS
        ) * rec["reach_factor"]
        rec["composite"] = round(composite, 4)
        if not rec["exempt"]:
            if composite >= 5.0:
                tiers["5+"] += 1
            elif composite >= 2.5:
                tiers["2.5-5"] += 1
            else:
                tiers["0-2.5"] += 1
        rows_out.append((
            rec["cve_id"], rec["composite"], rec["kev"], rec["external_facing"],
            rec["pre_auth"], rec["pre_auth_known"], rec["cvss"], rec["exempt"],
            rec["top_eai"], rec["eai_apps"], rec["deployments"], rec["severity"],
            rec["min_grace"], rec["in_triage"], rec["reach_factor"], rec["reach_level"],
        ))

    conn = _connect_scored()
    try:
        conn.execute("DELETE FROM cgr_scored")
        conn.executemany(
            "INSERT OR REPLACE INTO cgr_scored (cve_id, composite, kev, "
            "external_facing, pre_auth, pre_auth_known, cvss, exempt, top_eai, "
            "eai_apps, deployments, severity, min_grace, in_triage, "
            "reach_factor, reach_level) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            rows_out,
        )
        conn.commit()
    finally:
        conn.close()

    exempt_n = sum(1 for r in records.values() if r["exempt"])
    external_n = sum(1 for r in records.values() if r["external_facing"])
    kev_n = sum(1 for r in records.values() if r["kev"])
    result = {
        "scored": len(records),
        "exempt": exempt_n,
        "external_facing": external_n,
        "kev": kev_n,
        "tiers": tiers,
        "preauth_fetched": fetched,
        "preauth_found": pre_auth_found,
        "preauth_left_unknown": left_unknown,
        "eai_inventory_loaded": len(eai_map),
    }
    logger.info(
        "score_all: scored=%d exempt=%d external=%d kev=%d | pre_auth "
        "fetched=%d (found=%d) left_unknown=%d | tiers=%s",
        result["scored"], exempt_n, external_n, kev_n, fetched,
        pre_auth_found, left_unknown, tiers,
    )
    return result


# --------------------- scored read accessors (read-only) ------------------- #
def _scored_row_dict(r: sqlite3.Row) -> dict:
    try:
        eai_apps = json.loads(r["eai_apps"]) if r["eai_apps"] else []
    except (TypeError, ValueError):
        eai_apps = []
    return {
        "cve_id": r["cve_id"],
        "composite": r["composite"],
        "kev": bool(r["kev"]),
        "external_facing": bool(r["external_facing"]),
        "pre_auth": bool(r["pre_auth"]),
        "pre_auth_known": bool(r["pre_auth_known"]),
        "cvss": r["cvss"],
        "exempt": bool(r["exempt"]),
        "top_eai": r["top_eai"],
        "eai_apps": eai_apps,
        "deployments": r["deployments"],
        "severity": r["severity"],
        "min_grace": r["min_grace"],
        "in_triage": bool(r["in_triage"]),
        "reach_factor": (r["reach_factor"] if "reach_factor" in r.keys() else None),
        "reach_level": (r["reach_level"] if "reach_level" in r.keys() else None),
    }


def top_composite(limit: int = 100, include_exempt: bool = False) -> List[dict]:
    """Scored CVEs ordered by composite desc, then cvss desc, then deployments desc."""
    conn = _ro_connect()
    if conn is None:
        return []
    try:
        where = "" if include_exempt else "WHERE exempt = 0"
        rows = conn.execute(
            f"SELECT * FROM cgr_scored {where} "
            "ORDER BY composite DESC, cvss DESC, deployments DESC LIMIT ?",
            (limit,),
        ).fetchall()
        return [_scored_row_dict(r) for r in rows]
    except sqlite3.OperationalError as e:
        logger.warning("top_composite: %s (run score_all first)", e)
        return []
    finally:
        conn.close()


def scoring_summary() -> dict:
    """War-room SUMMARY-STATISTICS block: counts + composite/cvss ranges."""
    conn = _ro_connect()
    if conn is None:
        return {"available": False}
    try:
        total = conn.execute("SELECT COUNT(*) FROM cgr_scored").fetchone()[0]
        if not total:
            return {"available": False}
        exempt = conn.execute(
            "SELECT COUNT(*) FROM cgr_scored WHERE exempt = 1"
        ).fetchone()[0]
        external = conn.execute(
            "SELECT COUNT(*) FROM cgr_scored WHERE external_facing = 1"
        ).fetchone()[0]
        kev = conn.execute(
            "SELECT COUNT(*) FROM cgr_scored WHERE kev = 1"
        ).fetchone()[0]
        pre_auth = conn.execute(
            "SELECT COUNT(*) FROM cgr_scored WHERE pre_auth = 1"
        ).fetchone()[0]
        rng = conn.execute(
            "SELECT MIN(composite), MAX(composite) FROM cgr_scored WHERE exempt = 0"
        ).fetchone()
        top10_cvss = conn.execute(
            "SELECT AVG(cvss) FROM (SELECT cvss FROM cgr_scored WHERE exempt = 0 "
            "ORDER BY composite DESC, cvss DESC LIMIT 10)"
        ).fetchone()[0]
        return {
            "available": True,
            "total_scored": total,
            "exempt": exempt,
            "scored_non_exempt": total - exempt,
            "external_facing": external,
            "kev": kev,
            "pre_auth": pre_auth,
            "composite_min": rng[0],
            "composite_max": rng[1],
            "top10_avg_cvss": round(top10_cvss, 2) if top10_cvss is not None else None,
        }
    except sqlite3.OperationalError as e:
        logger.warning("scoring_summary: %s", e)
        return {"available": False}
    finally:
        conn.close()


def cgr_cluster_eai(limit: int = 80) -> List[dict]:
    """EAI-joined cluster rollup: each cluster annotated with its EAI app facts.

    Ordered by internet_facing desc, then distinct_high_cve desc. This is the
    EAI-enriched extension of :func:`cluster_rollup`.
    """
    conn = _ro_connect()
    if conn is None:
        return []
    try:
        rows = conn.execute(
            "SELECT eai_code, distinct_cve, distinct_high_cve, total_findings, "
            "distinct_images, min_grace FROM cgr_cluster"
        ).fetchall()
    except sqlite3.OperationalError:
        conn.close()
        return []
    conn.close()

    eai_map = load_eai_map()
    out = []
    for r in rows:
        app = eai_map.get(str(r["eai_code"]).strip())
        out.append({
            "eai_code": r["eai_code"],
            "app_name": (app.get("app_name") or "").strip() if app else None,
            "internet_facing": (app.get("internet_facing") or "").strip() if app else "",
            "critical_ind": (app.get("critical_ind") or "").strip() if app else "",
            "lob": (app.get("lob") or "").strip() if app else "",
            "distinct_cve": r["distinct_cve"],
            "distinct_high_cve": r["distinct_high_cve"],
            "total_findings": r["total_findings"],
            "min_grace": r["min_grace"],
        })
    out.sort(
        key=lambda d: (
            1 if str(d["internet_facing"]).lower() == "yes" else 0,
            d["distinct_high_cve"] or 0,
            d["distinct_cve"] or 0,
        ),
        reverse=True,
    )
    return out[:limit]


def gaps(limit: int = 200, triage_db_path: str = DEFAULT_TRIAGE_DB) -> List[dict]:
    """Risk-ranked coverage gap: cgr-only CVEs annotated with composite/EAI.

    Same coverage-gap set as :func:`cve_gaps` (CVEs Prisma flagged but we have
    not triaged), but each row is annotated with its composite score,
    external_facing flag, and representative app_name, and the list is ordered
    by composite desc so the highest-risk untriaged CVEs surface first.
    """
    triaged = _triage_cves(triage_db_path)

    conn = _ro_connect()
    if conn is None:
        return []
    try:
        cgr_rows = {r["cve_id"]: dict(r) for r in conn.execute("SELECT * FROM cgr_cve")}
    except sqlite3.OperationalError:
        conn.close()
        return []
    # scored overlay (may be absent if score_all not yet run).
    scored: Dict[str, dict] = {}
    try:
        for r in conn.execute("SELECT * FROM cgr_scored"):
            scored[r["cve_id"]] = _scored_row_dict(r)
    except sqlite3.OperationalError:
        pass
    conn.close()

    out = []
    for cid, r in cgr_rows.items():
        if cid in triaged:
            continue
        try:
            packages = json.loads(r["packages"]) if r["packages"] else []
        except (TypeError, ValueError):
            packages = []
        s = scored.get(cid)
        out.append({
            "cve_id": cid,
            "composite": s["composite"] if s else None,
            "external_facing": bool(s["external_facing"]) if s else False,
            "kev": bool(s["kev"]) if s else False,
            "app_name": (s["top_eai"] if s else None),
            "exempt": bool(s["exempt"]) if s else False,
            "severity": r["severity"],
            "max_cvss": r["max_cvss"],
            "deployments": r["deployments"],
            "distinct_eai": r["distinct_eai"],
            "any_from_base": bool(r["any_from_base"]),
            "min_grace": r["min_grace"],
            "packages": packages,
        })

    def keyfn(d):
        return (
            d["composite"] if d["composite"] is not None else -1.0,
            _sev_rank(d["severity"]),
            d["max_cvss"] if d["max_cvss"] is not None else -1.0,
            d["deployments"] or 0,
        )

    out.sort(key=keyfn, reverse=True)
    return out[:limit]


# --------------------------------- CLI ------------------------------------- #
def _main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    ap = argparse.ArgumentParser(description="CGR (Prisma Cloud) vuln export tooling")
    ap.add_argument("--ingest", action="store_true", help="ingest the latest CGR export")
    ap.add_argument("--reconcile", action="store_true", help="reconcile vs CVE triage")
    ap.add_argument("--summary", action="store_true", help="print store summary")
    ap.add_argument("--score", action="store_true",
                    help="compute EAI-enriched composite scores (rebuild cgr_scored)")
    ap.add_argument("--top", nargs="?", type=int, const=15, default=None,
                    metavar="N", help="print top N composite-scored CVEs (default 15)")
    ap.add_argument("--clusters", action="store_true",
                    help="print EAI-joined cluster rollup (internet-facing first)")
    ap.add_argument("--refresh-eai", action="store_true",
                    help="re-pull EAI inventory from EAI prod into the cache (on-tunnel)")
    ap.add_argument("--include-exempt", action="store_true",
                    help="include TLS/CURL-exempt CVEs in --top")
    ap.add_argument("--preauth-cap", type=int, default=_PREAUTH_FETCH_CAP,
                    help="max live pre_auth CVSS-vector fetches in --score")
    ap.add_argument("--limit", type=int, default=10, help="rows for top lists")
    args = ap.parse_args()

    if not (args.ingest or args.reconcile or args.summary or args.score
            or args.top is not None or args.clusters or args.refresh_eai):
        ap.print_help()
        return

    if args.refresh_eai:
        res = refresh_eai_cache()
        if res.get("error"):
            print(f"refresh-eai: FAILED ({res['error']}) -- off-tunnel? "
                  f"cache left unchanged")
        else:
            print(f"refresh-eai: refreshed {res['refreshed']} EAI rows into {EAI_DB_PATH}")

    if args.ingest:
        p = latest_export()
        print(f"export: {p}")
        res = ingest_export(p)
        print(f"ingested: rows={res['rows']} cves={res['distinct_cves']} "
              f"clusters={res['distinct_clusters']} images={res['distinct_images']}")

    if args.reconcile:
        r = reconcile(top_limit=max(args.limit, 200))
        print(f"reconcile: overlap={r['overlap']} cgr_only={r['cgr_only']} "
              f"triage_only={r['triage_only']} cgr_only_high={r['cgr_only_high']}")
        print(f"top {args.limit} cgr-only CVEs (severity, cvss, deployments):")
        for g in r["cgr_only_top"][:args.limit]:
            pkgs = ",".join(g["packages"][:3])
            print(f"  {g['cve_id']:<20} {str(g['severity'] or '-'):<8} "
                  f"cvss={g['max_cvss']} deploy={g['deployments']} "
                  f"eai={g['distinct_eai']} base={g['any_from_base']} "
                  f"grace={g['min_grace']} [{pkgs}]")

    if args.summary:
        s = summary()
        if not s.get("available"):
            print("summary: no data ingested yet")
        else:
            print(f"source_file: {s['source_file']}")
            print(f"ingested_at: {s['ingested_at']}")
            print(f"total_rows={s['total_rows']} distinct_cves={s['distinct_cves']} "
                  f"distinct_clusters={s['distinct_clusters']} "
                  f"distinct_images={s['distinct_images']}")
            print(f"severity_distribution: {s['severity_distribution']}")
            print(f"from_base={s['cves_from_base']} not_from_base={s['cves_not_from_base']} "
                  f"overdue={s['cves_overdue']}")
            print(f"top {args.limit} clusters (eai_code: high/total CVE):")
            for c in cluster_rollup(args.limit):
                print(f"  {c['eai_code']:<20} high={c['distinct_high_cve']} "
                      f"cve={c['distinct_cve']} findings={c['total_findings']} "
                      f"images={c['distinct_images']} min_grace={c['min_grace']}")

    if args.score:
        res = score_all(fetch_preauth_cap=args.preauth_cap)
        if res.get("error"):
            print(f"score: FAILED ({res['error']})")
        else:
            print(f"scored={res['scored']} exempt={res['exempt']} "
                  f"external_facing={res['external_facing']} kev={res['kev']} "
                  f"eai_inventory={res['eai_inventory_loaded']}")
            print(f"tiers (non-exempt): {res['tiers']}")
            print(f"pre_auth: fetched={res['preauth_fetched']} "
                  f"(found={res['preauth_found']}) "
                  f"left_unknown={res['preauth_left_unknown']}")
            s = scoring_summary()
            if s.get("available"):
                print(f"SUMMARY: total_scored={s['total_scored']} "
                      f"non_exempt={s['scored_non_exempt']} exempt={s['exempt']} "
                      f"external={s['external_facing']} kev={s['kev']} "
                      f"pre_auth={s['pre_auth']}")
                print(f"  composite range: {s['composite_min']}..{s['composite_max']} "
                      f"| top-10 avg cvss: {s['top10_avg_cvss']}")

    if args.top is not None:
        n = args.top
        rows = top_composite(limit=n, include_exempt=args.include_exempt)
        if not rows:
            print("top: no scored data (run --score first)")
        else:
            print(f"top {n} composite-scored CVEs"
                  f"{' (incl. exempt)' if args.include_exempt else ''}:")
            for r in rows:
                terms = (
                    f"KEV={'Y' if r['kev'] else '-'} "
                    f"EXT={'Y' if r['external_facing'] else '-'} "
                    f"PA={'Y' if r['pre_auth'] else ('?' if not r['pre_auth_known'] else '-')} "
                    f"cvss={r['cvss']}"
                )
                exm = " EXEMPT" if r["exempt"] else ""
                print(f"  {r['cve_id']:<18} comp={r['composite']:<6} {terms} "
                      f"deploy={r['deployments']} app={r['top_eai'] or '-'}{exm}")

    if args.clusters:
        rows = cgr_cluster_eai(limit=max(args.limit, 80))
        print(f"EAI-joined clusters (internet-facing first), top {len(rows)}:")
        for c in rows:
            print(f"  {str(c['eai_code']):<10} "
                  f"if={c['internet_facing'] or '-':<4} "
                  f"crit={c['critical_ind'] or '-':<2} "
                  f"high={c['distinct_high_cve']:<4} cve={c['distinct_cve']:<5} "
                  f"findings={c['total_findings']:<6} "
                  f"app={c['app_name'] or '-'} ({c['lob'] or '-'})")


if __name__ == "__main__":
    _main()
