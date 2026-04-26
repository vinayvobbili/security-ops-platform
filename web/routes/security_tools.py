"""Security tools routes: domain lookalike, VirusTotal, RecordedFuture, S3 brand squatting."""

import json
import logging
import os
import tempfile
import threading
from datetime import datetime
from pathlib import Path

from flask import Blueprint, Response, jsonify, render_template, request, send_file

from src.utils.logging_utils import log_web_activity
from src.components.web.edit_auth import check_s3_password
from services import domain_lookalike
from services import s3_brand_scanner
from services.virustotal import VirusTotalClient

logger = logging.getLogger(__name__)
security_tools_bp = Blueprint('security_tools', __name__)

# VirusTotal client singleton
_vt_client: VirusTotalClient | None = None


def _get_vt_client() -> VirusTotalClient | None:
    """Get VirusTotal client if configured."""
    global _vt_client
    if _vt_client is None:
        _vt_client = VirusTotalClient()
    return _vt_client if _vt_client.is_configured() else None


# --- Domain Lookalike ---

@security_tools_bp.route('/domain-lookalike')
@log_web_activity
def domain_lookalike_search():
    """Display domain lookalike search page."""
    return render_template('domain_lookalike_search.html')


@security_tools_bp.route('/api/domain-lookalikes')
@log_web_activity
def api_domain_lookalikes():
    """SSE endpoint to stream domain lookalike scan with keepalive heartbeats.

    Uses Server-Sent Events to prevent reverse proxy timeouts during long
    dnstwist scans.  Sends heartbeat events every 10s while running, then
    a single 'complete' or 'error' event with the full result payload.

    Use /api/domain-lookalikes/parking SSE endpoint for streaming parking status.
    """
    domain = request.args.get('domain', '').strip()
    registered_only = request.args.get('registered_only', 'false').lower() == 'true'
    include_malicious_tlds = request.args.get('include_malicious_tlds', 'false').lower() == 'true'

    if not domain:
        return jsonify({'success': False, 'error': 'Domain parameter is required'}), 400

    def generate():
        result_holder = [None]
        error_holder = [None]

        def run_scan():
            try:
                result_holder[0] = domain_lookalike.get_domain_lookalikes(
                    domain, registered_only, include_malicious_tlds,
                    include_dictionary_combos=include_malicious_tlds,
                )
            except Exception as e:
                logger.error(f"Error in domain lookalike scan: {e}", exc_info=True)
                error_holder[0] = "An error occurred during the scan"

        thread = threading.Thread(target=run_scan, daemon=True)
        thread.start()

        # Send heartbeat every 10s to keep the connection alive through proxies
        heartbeat_count = 0
        while thread.is_alive():
            heartbeat_count += 1
            yield f"data: {json.dumps({'status': 'scanning', 'elapsed': heartbeat_count * 10})}\n\n"
            thread.join(timeout=10)

        # Send final result
        if error_holder[0]:
            yield f"data: {json.dumps({'status': 'error', 'error': error_holder[0]})}\n\n"
        elif result_holder[0] is not None:
            yield f"data: {json.dumps({'status': 'complete', 'result': result_holder[0]})}\n\n"
        else:
            yield f"data: {json.dumps({'status': 'error', 'error': 'Scan produced no results'})}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
        }
    )


@security_tools_bp.route('/api/domain-lookalikes/parking')
@log_web_activity
def api_domain_lookalikes_parking():
    """SSE endpoint to stream parking status for domains.

    Query params:
        domains: Comma-separated list of domains to check

    Returns:
        Server-Sent Events stream with parking status for each domain
    """
    domains_param = request.args.get('domains', '').strip()

    if not domains_param:
        return jsonify({'success': False, 'error': 'domains parameter is required'}), 400

    domains = [d.strip() for d in domains_param.split(',') if d.strip()]

    if not domains:
        return jsonify({'success': False, 'error': 'No valid domains provided'}), 400

    if len(domains) > 500:
        return jsonify({'success': False, 'error': 'Maximum 500 domains allowed'}), 400

    def generate():
        """Generator that yields SSE events for each domain's parking status."""
        for domain_name in domains:
            try:
                parked = domain_lookalike.check_if_parked(domain_name)
                data = json.dumps({'domain': domain_name, 'parked': parked})
                yield f"data: {data}\n\n"
            except Exception as e:
                logger.error(f"Error checking parking for {domain_name}: {e}", exc_info=True)
                data = json.dumps({'domain': domain_name, 'parked': None, 'error': 'Check failed'})
                yield f"data: {data}\n\n"

        # Send completion event
        yield "data: {\"complete\": true}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'  # Disable nginx buffering
        }
    )


@security_tools_bp.route('/api/domain-whois')
@log_web_activity
def api_domain_whois():
    """API endpoint to get WHOIS information for a domain."""
    try:
        domain = request.args.get('domain', '').strip()

        if not domain:
            return jsonify({'success': False, 'error': 'Domain parameter is required'}), 400

        result = domain_lookalike.get_domain_whois_info(domain)
        return jsonify(result)

    except Exception as exc:
        logger.error(f"Error in domain WHOIS API: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': 'An internal error occurred'}), 500


# --- S3 Brand Squatting ---

_DOMAIN_MONITORING_CONFIG = Path(__file__).resolve().parent.parent.parent / "data" / "transient" / "domain_monitoring" / "config.json"


def _load_known_good_buckets(domain: str) -> list[str]:
    """Load known-good S3 buckets for a domain from monitoring config."""
    try:
        with open(_DOMAIN_MONITORING_CONFIG) as f:
            cfg = json.load(f)
        return cfg.get("known_good_buckets", {}).get(domain, [])
    except Exception:
        return []


def _load_bucket_classifications(domain: str) -> dict:
    """Load bucket classifications (ours/investigate/unknown) from config."""
    try:
        with open(_DOMAIN_MONITORING_CONFIG) as f:
            cfg = json.load(f)
        return cfg.get("bucket_classifications", {}).get(domain, {})
    except Exception:
        return {}


def _save_bucket_classification(domain: str, bucket: str, classification: str):
    """Save a bucket classification to config. Also updates known_good_buckets."""
    try:
        with open(_DOMAIN_MONITORING_CONFIG) as f:
            cfg = json.load(f)
    except Exception:
        cfg = {}

    # Update classifications
    if "bucket_classifications" not in cfg:
        cfg["bucket_classifications"] = {}
    if domain not in cfg["bucket_classifications"]:
        cfg["bucket_classifications"][domain] = {}
    cfg["bucket_classifications"][domain][bucket] = classification

    # Sync known_good_buckets: "ours" buckets go in, others get removed
    if "known_good_buckets" not in cfg:
        cfg["known_good_buckets"] = {}
    if domain not in cfg["known_good_buckets"]:
        cfg["known_good_buckets"][domain] = []

    good = set(cfg["known_good_buckets"][domain])
    if classification == "ours":
        good.add(bucket)
    else:
        good.discard(bucket)
    cfg["known_good_buckets"][domain] = sorted(good)

    with open(_DOMAIN_MONITORING_CONFIG, "w") as f:
        json.dump(cfg, f, indent=2)


@security_tools_bp.route('/api/domain-lookalikes/s3-buckets')
@log_web_activity
def api_s3_bucket_scan():
    """SSE endpoint to stream S3 bucket brand-squatting scan results.

    Uses a background thread + queue pattern (same as domain-lookalikes SSE)
    so heartbeats keep the connection alive through Waitress/nginx.

    Query params:
        domain: Domain to scan (e.g. the company.com) — brand name extracted automatically.
    """
    domain = request.args.get('domain', '').strip().lower()
    if not domain:
        return jsonify({'success': False, 'error': 'domain parameter is required'}), 400

    brand = domain.split('.')[0]
    if not brand:
        return jsonify({'success': False, 'error': 'Could not extract brand name from domain'}), 400

    known_good = _load_known_good_buckets(domain)

    def generate():
        import queue as _q

        eq, thread, scan_info = s3_brand_scanner.start_scan_with_queue(brand, known_good)

        # Send initial progress event
        total = scan_info['total']
        filtered = scan_info['filtered_out']
        init_msg = f"Generated {total} bucket name permutations"
        init_event = json.dumps({
            'type': 'progress', 'phase': 'generating',
            'total': total, 'filtered_out': filtered, 'message': init_msg,
        })
        yield f"data: {init_event}\n\n"

        # Drain the queue with heartbeats while the scan thread is running
        while thread.is_alive() or not eq.empty():
            # Drain all available events
            drained = 0
            while True:
                try:
                    event = eq.get_nowait()
                    yield f"data: {json.dumps(event)}\n\n"
                    drained += 1
                    if event.get("type") == "complete" or event.get("type") == "error":
                        return
                except _q.Empty:
                    break

            # If no events were drained and thread is alive, send heartbeat
            if drained == 0 and thread.is_alive():
                yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"

            # Wait a bit before next drain cycle
            thread.join(timeout=1)

        # Final drain in case anything was queued after thread finished
        while not eq.empty():
            try:
                event = eq.get_nowait()
                yield f"data: {json.dumps(event)}\n\n"
            except _q.Empty:
                break

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
        },
    )


@security_tools_bp.route('/api/domain-lookalikes/s3-buckets/cached')
@log_web_activity
def api_s3_bucket_cached():
    """Return cached S3 brand scan results for a domain (if any)."""
    domain = request.args.get('domain', '').strip().lower()
    if not domain:
        return jsonify({'success': False, 'error': 'domain parameter is required'}), 400
    brand = domain.split('.')[0]
    cached = s3_brand_scanner.load_scan_cache(brand)
    if not cached:
        return jsonify({'success': True, 'cached': False})
    return jsonify({'success': True, 'cached': True, **cached})


@security_tools_bp.route('/api/domain-lookalikes/s3-buckets/history')
@log_web_activity
def api_s3_bucket_history():
    """List available historical S3 scans for a domain."""
    domain = request.args.get('domain', '').strip().lower()
    if not domain:
        return jsonify({'success': False, 'error': 'domain parameter is required'}), 400
    brand = domain.split('.')[0]
    history = s3_brand_scanner.list_scan_history(brand)
    return jsonify({'success': True, 'history': history})


@security_tools_bp.route('/api/domain-lookalikes/s3-buckets/history/<scan_id>')
@log_web_activity
def api_s3_bucket_history_load(scan_id):
    """Load a specific historical S3 scan."""
    domain = request.args.get('domain', '').strip().lower()
    if not domain:
        return jsonify({'success': False, 'error': 'domain parameter is required'}), 400
    brand = domain.split('.')[0]
    data = s3_brand_scanner.load_scan_history(brand, scan_id)
    if not data:
        return jsonify({'success': False, 'error': 'Scan not found'}), 404
    return jsonify({'success': True, 'cached': True, **data})


@security_tools_bp.route('/api/domain-lookalikes/s3-buckets/verify-auth', methods=['POST'])
@log_web_activity
def api_s3_verify_auth():
    """Verify password for S3 scan authorization."""
    if not check_s3_password(request):
        return jsonify({'success': False, 'error': 'Invalid password'}), 403
    return jsonify({'success': True})


@security_tools_bp.route('/api/domain-lookalikes/s3-buckets/classify', methods=['POST'])
@log_web_activity
def api_s3_bucket_classify():
    """Classify an S3 bucket as ours/investigate/unknown.

    JSON body: { "domain": "the company.com", "bucket": "the company-cdn", "classification": "ours" }
    Valid classifications: "ours", "investigate", "unknown"
    """
    data = request.get_json(silent=True) or {}
    domain = data.get('domain', '').strip().lower()
    bucket = data.get('bucket', '').strip().lower()
    classification = data.get('classification', '').strip().lower()

    if not domain or not bucket:
        return jsonify({'success': False, 'error': 'domain and bucket are required'}), 400
    if classification not in ('ours', 'investigate', 'unknown'):
        return jsonify({'success': False, 'error': 'classification must be ours, investigate, or unknown'}), 400

    # "Ours" requires password — it adds to the known-good allowlist
    if classification == 'ours' and not check_s3_password(request):
        return jsonify({'success': False, 'error': 'Invalid password'}), 403

    try:
        _save_bucket_classification(domain, bucket, classification)
        return jsonify({'success': True, 'bucket': bucket, 'classification': classification})
    except Exception as exc:
        logger.error("Failed to classify bucket: %s", exc, exc_info=True)
        return jsonify({'success': False, 'error': str(exc)}), 500


@security_tools_bp.route('/api/domain-lookalikes/s3-buckets/classifications')
@log_web_activity
def api_s3_bucket_classifications():
    """Get all bucket classifications for a domain."""
    domain = request.args.get('domain', '').strip().lower()
    if not domain:
        return jsonify({'success': False, 'error': 'domain parameter is required'}), 400
    return jsonify({'success': True, 'classifications': _load_bucket_classifications(domain)})


@security_tools_bp.route('/api/domain-lookalikes/s3-buckets/export')
@log_web_activity
def api_s3_bucket_export():
    """Export S3 brand squatting scan results as a professionally formatted Excel file.

    Query params:
        domain: The scanned domain (for filename)
        buckets: JSON-encoded array of found bucket dicts
    """
    domain = request.args.get('domain', 'unknown').strip()
    buckets_json = request.args.get('buckets', '[]')

    try:
        buckets = json.loads(buckets_json)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid buckets JSON'}), 400

    if not buckets:
        return jsonify({'success': False, 'error': 'No buckets to export'}), 400

    import pandas as pd
    from src.utils.excel_formatting import apply_professional_formatting

    rows = []
    for b in buckets:
        rows.append({
            'Bucket Name': b.get('bucket', ''),
            'Status': 'PUBLIC (Listable)' if b.get('listable') else 'Exists (Private)',
            'Region': b.get('region') or 'us-east-1',
            'Objects': b.get('key_count', 0) if b.get('listable') else 'N/A',
            'URL': b.get('url', ''),
        })

    df = pd.DataFrame(rows)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir='/tmp')
    tmp_path = tmp.name
    tmp.close()
    df.to_excel(tmp_path, index=False, engine='openpyxl')

    col_widths = {
        'bucket name': 40, 'status': 22, 'region': 16,
        'objects': 12, 'url': 55,
    }
    apply_professional_formatting(tmp_path, column_widths=col_widths)

    # Make URLs clickable
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = load_workbook(tmp_path)
    ws = wb.active
    url_col = 5  # URL column (E)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=url_col)
        if cell.value and str(cell.value).startswith('http'):
            cell.hyperlink = str(cell.value)
            cell.style = 'Hyperlink'
    wb.save(tmp_path)

    # Add metadata sheet
    wb = load_workbook(tmp_path)
    meta = wb.create_sheet("Scan Info", 0)
    brand = domain.split('.')[0]
    public_count = sum(1 for b in buckets if b.get('listable'))
    meta_rows = [
        ("S3 Brand Squatting Scanner — Results", ""),
        ("", ""),
        ("Brand", brand),
        ("Domain", domain),
        ("Buckets Found", str(len(buckets))),
        ("Publicly Listable", str(public_count)),
        ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_data in meta_rows:
        meta.append(row_data)
    meta["A1"].font = Font(bold=True, size=14, color="0891B2")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for row_idx in range(3, len(meta_rows) + 1):
        meta.cell(row=row_idx, column=1).font = Font(bold=True)
        for col_idx in (1, 2):
            cell = meta.cell(row=row_idx, column=col_idx)
            cell.border = thin
            cell.alignment = Alignment(vertical='top')
    meta.column_dimensions['A'].width = 22
    meta.column_dimensions['B'].width = 60

    data_ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[-1]
    data_ws.title = "Buckets"
    wb.save(tmp_path)

    ts = datetime.now().strftime('%Y-%m-%d_%H%M')
    filename = f's3_brand_squatting_{brand}_{len(buckets)}_found_{ts}.xlsx'

    return send_file(
        tmp_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# --- RecordedFuture ---

@security_tools_bp.route('/api/domain-lookalikes/rf-enrich')
@log_web_activity
def api_domain_lookalikes_rf_enrich():
    """API endpoint to enrich domains with RecordedFuture threat intelligence.

    Query params:
        domains: Comma-separated list of domains to enrich

    Returns:
        JSON with RF risk scores and evidence rules for each domain
    """
    domains_param = request.args.get('domains', '').strip()

    if not domains_param:
        return jsonify({'success': False, 'error': 'domains parameter is required'}), 400

    domains = [d.strip() for d in domains_param.split(',') if d.strip()]

    if not domains:
        return jsonify({'success': False, 'error': 'No valid domains provided'}), 400

    if len(domains) > 1000:
        return jsonify({'success': False, 'error': 'Maximum 1000 domains allowed'}), 400

    try:
        from services.recorded_future import RecordedFutureClient

        client = RecordedFutureClient()
        if not client.is_configured():
            return jsonify({
                'success': False,
                'error': 'RecordedFuture API key not configured'
            }), 503

        # Enrich domains
        result = client.enrich_domains(domains)

        if "error" in result:
            return jsonify({'success': False, 'error': result['error']}), 502

        # Extract and format results
        enriched = client.extract_enrichment_results(result)

        # Build response map
        results_map = {}
        for item in enriched:
            results_map[item['value']] = {
                'risk_score': item.get('risk_score', 0),
                'risk_level': item.get('risk_level', 'Unknown'),
                'rules': item.get('rules', []),
                'evidence_count': item.get('evidence_count', 0),
            }

        # Count high risk
        high_risk_count = sum(1 for v in results_map.values() if v.get('risk_score', 0) >= 65)

        return jsonify({
            'success': True,
            'domains_enriched': len(results_map),
            'high_risk_count': high_risk_count,
            'results': results_map
        })

    except ImportError:
        return jsonify({
            'success': False,
            'error': 'RecordedFuture client not available'
        }), 503

    except Exception as exc:
        logger.error(f"Error in RF enrichment API: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': 'An internal error occurred'}), 500


@security_tools_bp.route('/api/rf/enrich')
@log_web_activity
def api_rf_enrich():
    """General-purpose RecordedFuture enrichment API.

    Query params:
        ips: Comma-separated list of IP addresses
        domains: Comma-separated list of domains
        hashes: Comma-separated list of file hashes
        urls: Comma-separated list of URLs

    Returns:
        JSON with RF risk scores and evidence rules
    """
    try:
        from services.recorded_future import RecordedFutureClient

        client = RecordedFutureClient()
        if not client.is_configured():
            return jsonify({
                'success': False,
                'error': 'RecordedFuture API key not configured'
            }), 503

        # Parse parameters
        ips = [ip.strip() for ip in request.args.get('ips', '').split(',') if ip.strip()]
        domains = [d.strip() for d in request.args.get('domains', '').split(',') if d.strip()]
        hashes = [h.strip() for h in request.args.get('hashes', '').split(',') if h.strip()]
        urls = [u.strip() for u in request.args.get('urls', '').split(',') if u.strip()]

        if not any([ips, domains, hashes, urls]):
            return jsonify({
                'success': False,
                'error': 'At least one of ips, domains, hashes, or urls is required'
            }), 400

        total_iocs = len(ips) + len(domains) + len(hashes) + len(urls)
        if total_iocs > 1000:
            return jsonify({
                'success': False,
                'error': 'Maximum 1000 total IOCs allowed'
            }), 400

        # Call enrichment
        result = client.enrich(
            ips=ips if ips else None,
            domains=domains if domains else None,
            hashes=hashes if hashes else None,
            urls=urls if urls else None,
        )

        if "error" in result:
            return jsonify({'success': False, 'error': result['error']}), 502

        # Extract results
        enriched = client.extract_enrichment_results(result)

        return jsonify({
            'success': True,
            'total_enriched': len(enriched),
            'results': enriched
        })

    except ImportError:
        return jsonify({
            'success': False,
            'error': 'RecordedFuture client not available'
        }), 503

    except Exception as exc:
        logger.error(f"Error in RF enrichment API: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': 'An internal error occurred'}), 500


# --- VirusTotal ---

@security_tools_bp.route('/api/virustotal/domain/<domain>')
@log_web_activity
def api_virustotal_domain(domain: str):
    """API endpoint for on-demand VirusTotal domain lookup.

    Returns reputation data for a domain including threat level and detection stats.
    """
    try:
        vt = _get_vt_client()
        if not vt:
            return jsonify({
                'success': False,
                'error': 'VirusTotal not configured (missing API key)'
            }), 503

        logger.info(f"VT lookup requested for domain: {domain}")
        result = vt.lookup_domain(domain)

        if "error" in result:
            return jsonify({
                'success': False,
                'error': result['error']
            }), 400 if 'not found' in result['error'].lower() else 429 if 'rate limit' in result['error'].lower() else 500

        # Extract key reputation data
        attrs = result.get("data", {}).get("attributes", {})
        stats = attrs.get("last_analysis_stats", {})

        malicious = stats.get("malicious", 0)
        suspicious = stats.get("suspicious", 0)
        harmless = stats.get("harmless", 0)
        undetected = stats.get("undetected", 0)

        # Determine threat level
        if malicious >= 3:
            threat_level = "HIGH"
        elif malicious >= 1 or suspicious >= 3:
            threat_level = "MEDIUM"
        elif suspicious >= 1:
            threat_level = "LOW"
        else:
            threat_level = "CLEAN"

        return jsonify({
            'success': True,
            'domain': domain,
            'result': {
                'malicious': malicious,
                'suspicious': suspicious,
                'harmless': harmless,
                'undetected': undetected,
                'threat_level': threat_level,
                'categories': attrs.get("categories", {}),
                'registrar': attrs.get("registrar", ""),
                'creation_date': attrs.get("creation_date"),
                'vt_link': f"https://www.virustotal.com/gui/domain/{domain}",
            }
        })

    except Exception as exc:
        logger.error(f"VT lookup error for {domain}: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': 'An internal error occurred'}), 500
