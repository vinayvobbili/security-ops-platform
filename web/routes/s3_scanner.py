"""AWS S3 public bucket scanner — web routes."""

import json
import logging
import re
import tempfile
from datetime import datetime

from flask import Blueprint, Response, jsonify, render_template, request, send_file

from src.utils.logging_utils import get_client_ip, log_web_activity

logger = logging.getLogger(__name__)
s3_scanner_bp = Blueprint('s3_scanner', __name__)

# S3 bucket naming rules: lowercase, digits, hyphens, dots; 3-63 chars; no ".."
_BUCKET_NAME_RE = re.compile(r'^[a-z0-9][a-z0-9.\-]{1,61}[a-z0-9]$')


@s3_scanner_bp.route('/s3-scanner')
@log_web_activity
def s3_scanner_page():
    return render_template('s3_scanner.html')


@s3_scanner_bp.route('/api/s3-scanner/targets', methods=['GET'])
@log_web_activity
def api_get_targets():
    from services.s3_scanner import load_targets
    return jsonify({'success': True, 'targets': load_targets()})


@s3_scanner_bp.route('/api/s3-scanner/targets', methods=['POST'])
@log_web_activity
def api_save_targets():
    from services.s3_scanner import load_targets, save_targets

    data = request.get_json(silent=True)
    if not data:
        return jsonify({'success': False, 'error': 'Invalid JSON body'}), 400

    action = data.get('action')

    if action == 'add':
        key = data.get('key', '').strip()
        label = data.get('label', '').strip()
        buckets = data.get('buckets', [])

        if not key or not label or not buckets:
            return jsonify({'success': False, 'error': 'key, label, and buckets are required'}), 400

        # Validate bucket names
        clean_buckets = []
        for b in buckets:
            b = b.strip()
            if not b:
                continue
            if '..' in b or not _BUCKET_NAME_RE.match(b):
                return jsonify({
                    'success': False,
                    'error': f'Invalid bucket name: {b}. '
                             'Must be 3-63 chars, lowercase alphanumeric, hyphens, and dots only (no "..").',
                }), 400
            clean_buckets.append(b)

        if not clean_buckets:
            return jsonify({'success': False, 'error': 'No valid bucket names provided'}), 400

        targets = load_targets()
        targets[key] = {'label': label, 'buckets': clean_buckets}

        # Optional sample_count override
        sample_count = data.get('sample_count')
        if sample_count is not None:
            try:
                targets[key]['sample_count'] = max(1, min(int(sample_count), 50))
            except (ValueError, TypeError):
                pass

        save_targets(targets)
        return jsonify({'success': True, 'targets': targets})

    elif action == 'remove':
        key = data.get('key', '').strip()
        if not key:
            return jsonify({'success': False, 'error': 'key is required'}), 400
        targets = load_targets()
        if key in targets:
            del targets[key]
            save_targets(targets)
        return jsonify({'success': True, 'targets': targets})

    return jsonify({'success': False, 'error': f'Unknown action: {action}'}), 400


@s3_scanner_bp.route('/api/s3-scanner/scan')
@log_web_activity
def api_scan():
    """SSE endpoint — streams scan progress and results."""
    from services.s3_scanner import load_targets, scan_buckets

    targets = load_targets()
    if not targets:
        return jsonify({'success': False, 'error': 'No scan targets configured'}), 400

    # Filter to selected targets (comma-separated keys)
    targets_param = request.args.get('targets', '').strip()
    if targets_param:
        selected_keys = [k.strip() for k in targets_param.split(',') if k.strip()]
        targets = {k: v for k, v in targets.items() if k in selected_keys}
        if not targets:
            return jsonify({'success': False, 'error': 'No matching targets found'}), 400

    client_ip = get_client_ip()

    def generate():
        final_report = None
        try:
            for event_type, data in scan_buckets(targets):
                if event_type == 'complete':
                    final_report = data
                yield f"event: {event_type}\ndata: {json.dumps(data)}\n\n"
        except Exception as e:
            logger.error(f"S3 scan error: {e}", exc_info=True)
            yield f"event: error\ndata: {json.dumps({'error': str(e)})}\n\n"

        # Auto-persist scan to SQLite
        if final_report:
            try:
                from services.s3_scanner_db import save_scan
                scan_id = save_scan(final_report, ip_address=client_ip)
                yield f"event: saved\ndata: {json.dumps({'scan_id': scan_id})}\n\n"
            except Exception as e:
                logger.error(f"Failed to persist S3 scan: {e}", exc_info=True)

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
        }
    )


@s3_scanner_bp.route('/api/s3-scanner/history')
@log_web_activity
def api_scan_history():
    from services.s3_scanner_db import get_scan_history
    limit = request.args.get('limit', 50, type=int)
    return jsonify({'success': True, 'scans': get_scan_history(limit)})


@s3_scanner_bp.route('/api/s3-scanner/scan/<scan_id>')
@log_web_activity
def api_get_scan(scan_id):
    from services.s3_scanner_db import get_scan
    scan = get_scan(scan_id)
    if not scan:
        return jsonify({'success': False, 'error': 'Scan not found'}), 404
    return jsonify({'success': True, 'scan': scan})


@s3_scanner_bp.route('/api/s3-scanner/diff')
@log_web_activity
def api_scan_diff():
    from services.s3_scanner_db import get_scan_diff
    a = request.args.get('a', '').strip()
    b = request.args.get('b', '').strip()
    if not a or not b:
        return jsonify({'success': False, 'error': 'Both a and b scan IDs required'}), 400
    diff = get_scan_diff(a, b)
    return jsonify({'success': True, 'diff': diff})


@s3_scanner_bp.route('/api/s3-scanner/download-objects')
@log_web_activity
def api_download_objects():
    """SSE endpoint — lists all objects in a bucket, builds an Excel file,
    then emits a download URL."""
    from services.s3_scanner import iter_all_objects

    bucket = request.args.get('bucket', '').strip()
    if not bucket or not _BUCKET_NAME_RE.match(bucket):
        return jsonify({'success': False, 'error': 'Invalid bucket name'}), 400

    def generate():
        all_objects = []
        error_msg = None

        try:
            for event_type, data in iter_all_objects(bucket):
                if event_type == 'objects':
                    all_objects.extend(data['objects'])
                elif event_type == 'progress':
                    yield f"event: progress\ndata: {json.dumps(data)}\n\n"
                elif event_type == 'error':
                    error_msg = data.get('error', 'Unknown error')
                    break
                elif event_type == 'done':
                    pass
        except Exception as e:
            logger.error(f"S3 download error: {e}", exc_info=True)
            error_msg = str(e)

        if error_msg and not all_objects:
            yield f"event: error\ndata: {json.dumps({'error': error_msg})}\n\n"
            return

        if not all_objects:
            yield f"event: error\ndata: {json.dumps({'error': 'No objects found in bucket'})}\n\n"
            return

        yield f"event: progress\ndata: {json.dumps({'status': f'Building Excel file ({len(all_objects):,} objects)...'})}\n\n"

        import pandas as pd
        from src.utils.excel_formatting import apply_professional_formatting

        # Add direct download URL for each object
        for obj in all_objects:
            obj['url'] = f"https://{bucket}.s3.amazonaws.com/{obj['key']}"

        df = pd.DataFrame(all_objects, columns=[
            'key', 'size_human', 'size', 'last_modified', 'storage_class', 'extension', 'url',
        ])
        df.columns = [
            'Object Key', 'Size', 'Size (bytes)', 'Last Modified',
            'Storage Class', 'Extension', 'Direct URL',
        ]

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir='/tmp')
        tmp_path = tmp.name
        tmp.close()
        df.to_excel(tmp_path, index=False, engine='openpyxl')

        col_widths = {
            'object key': 60, 'size': 12, 'size (bytes)': 14,
            'last modified': 22, 'storage class': 16, 'extension': 12,
            'direct url': 70,
        }
        apply_professional_formatting(tmp_path, column_widths=col_widths)

        # Make URLs clickable hyperlinks in Excel
        from openpyxl import load_workbook as _lwb
        _wb = _lwb(tmp_path)
        _ws = _wb.active
        url_col = 7  # Direct URL column (G)
        for row_idx in range(2, _ws.max_row + 1):
            cell = _ws.cell(row=row_idx, column=url_col)
            if cell.value and str(cell.value).startswith('http'):
                cell.hyperlink = str(cell.value)
                cell.style = 'Hyperlink'
        _wb.save(tmp_path)

        # Add metadata sheet
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, Side
        wb = load_workbook(tmp_path)
        meta = wb.create_sheet("Scan Info", 0)
        meta_rows = [
            ("S3 Public Bucket Scanner — Object Listing", ""),
            ("", ""),
            ("Bucket", bucket),
            ("Objects Listed", f"{len(all_objects):,}"),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for row_data in meta_rows:
            meta.append(row_data)
        meta["A1"].font = Font(bold=True, size=14, color="0891B2")
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        for row_idx in range(3, len(meta_rows) + 1):
            meta.cell(row=row_idx, column=1).font = Font(bold=True)
            for col_idx in (1, 2):
                cell = meta.cell(row=row_idx, column=col_idx)
                cell.border = thin
                cell.alignment = Alignment(vertical='top')
        meta.column_dimensions['A'].width = 22
        meta.column_dimensions['B'].width = 60

        data_ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[-1]
        data_ws.title = "Objects"
        wb.save(tmp_path)

        import os
        file_id = os.path.basename(tmp_path)
        ts = datetime.now().strftime('%Y-%m-%d_%H%M')
        filename = f'{bucket}_{len(all_objects)}_objects_{ts}.xlsx'
        yield f"event: done\ndata: {json.dumps({'total': len(all_objects), 'file_id': file_id, 'filename': filename})}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@s3_scanner_bp.route('/api/s3-scanner/download-file')
@log_web_activity
def api_download_file():
    """Serve a previously-generated Excel file."""
    import os
    file_id = request.args.get('id', '').strip()
    filename = request.args.get('filename', 'export.xlsx').strip()

    if not file_id or '/' in file_id or '\\' in file_id or '..' in file_id:
        return jsonify({'success': False, 'error': 'Invalid file id'}), 400

    tmp_path = os.path.join('/tmp', file_id)
    if not os.path.isfile(tmp_path):
        return jsonify({'success': False, 'error': 'File not found — it may have expired'}), 404

    return send_file(
        tmp_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@s3_scanner_bp.route('/api/s3-scanner/bucket-counts')
@log_web_activity
def api_bucket_counts():
    """Return latest deep-scan object counts for given buckets."""
    from services.s3_scanner_db import get_bucket_object_counts

    buckets_param = request.args.get('buckets', '').strip()
    if not buckets_param:
        return jsonify({'success': False, 'error': 'buckets parameter required'}), 400

    buckets = [b.strip() for b in buckets_param.split(',') if b.strip()]
    counts = get_bucket_object_counts(buckets)
    return jsonify({'success': True, 'counts': counts})


@s3_scanner_bp.route('/api/s3-scanner/download-s3-object')
@log_web_activity
def api_download_s3_object():
    """Proxy-download an individual object from a public S3 bucket.

    This lets the IR team pull actual files (PDFs, PNGs, etc.) as evidence
    without needing direct S3 access or AWS credentials.
    """
    import requests as req

    bucket = request.args.get('bucket', '').strip()
    key = request.args.get('key', '').strip()

    if not bucket or not _BUCKET_NAME_RE.match(bucket):
        return jsonify({'success': False, 'error': 'Invalid bucket name'}), 400
    if not key:
        return jsonify({'success': False, 'error': 'Object key is required'}), 400
    # Safety: reject keys with path traversal
    if '..' in key:
        return jsonify({'success': False, 'error': 'Invalid object key'}), 400

    # Build the S3 URL — try virtual-hosted style
    s3_url = f"https://{bucket}.s3.amazonaws.com/{key}"

    try:
        resp = req.get(s3_url, timeout=(5, 60), stream=True, verify=True)
    except req.RequestException as e:
        logger.error(f"S3 object download failed: {e}")
        return jsonify({'success': False, 'error': f'Failed to fetch object: {e}'}), 502

    if resp.status_code != 200:
        return jsonify({
            'success': False,
            'error': f'S3 returned HTTP {resp.status_code}',
        }), resp.status_code

    content_type = resp.headers.get('Content-Type', 'application/octet-stream')
    content_length = resp.headers.get('Content-Length')

    # Derive a safe filename from the key
    import os
    filename = os.path.basename(key) or 'download'

    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"',
        'Content-Type': content_type,
    }
    if content_length:
        headers['Content-Length'] = content_length

    def stream():
        for chunk in resp.iter_content(chunk_size=65536):
            yield chunk

    return Response(stream(), status=200, headers=headers)
