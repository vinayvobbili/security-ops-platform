"""OE Detection dashboard routes."""

import logging

from flask import Blueprint, jsonify, render_template, request

from src.utils.logging_utils import log_web_activity

logger = logging.getLogger(__name__)

oe_detection_bp = Blueprint('oe_detection', __name__)


@oe_detection_bp.route('/oe-detection')
@log_web_activity
def oe_detection_page():
    return render_template('oe_detection.html')


@oe_detection_bp.route('/api/oe-detection/dashboard')
@log_web_activity
def api_oe_detection_dashboard():
    try:
        from src.components.web import oe_detection_handler

        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        data = oe_detection_handler.get_dashboard_data(start_date=start_date, end_date=end_date)
        return jsonify(data)
    except Exception as exc:
        logger.error(f"OE dashboard API error: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': str(exc)}), 500


@oe_detection_bp.route('/api/oe-detection/employee/<employee_id>')
@log_web_activity
def api_oe_detection_employee(employee_id):
    try:
        from src.components.web import oe_detection_handler

        data = oe_detection_handler.get_employee_detail(employee_id)
        return jsonify(data)
    except Exception as exc:
        logger.error(f"OE employee detail API error: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': str(exc)}), 500


@oe_detection_bp.route('/api/oe-detection/scans')
@log_web_activity
def api_oe_detection_scans():
    try:
        from src.components.web import oe_detection_handler

        data = oe_detection_handler.get_scan_history()
        return jsonify(data)
    except Exception as exc:
        logger.error(f"OE scan history API error: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': str(exc)}), 500


@oe_detection_bp.route('/api/oe-detection/scan', methods=['POST'])
@log_web_activity
def api_oe_detection_trigger_scan():
    try:
        from src.components.web import oe_detection_handler

        body = request.get_json(silent=True) or {}
        employee_id = body.get('employee_id')
        dry_run = body.get('dry_run', False)

        data = oe_detection_handler.trigger_scan(employee_id=employee_id, dry_run=dry_run)
        return jsonify(data)
    except Exception as exc:
        logger.error(f"OE scan trigger API error: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': str(exc)}), 500


@oe_detection_bp.route('/api/oe-detection/export', methods=['POST'])
@log_web_activity
def api_oe_detection_export():
    try:
        from src.components.web import oe_detection_handler
        import io
        from flask import send_file

        body = request.get_json(silent=True) or {}
        start_date = body.get('start_date')
        end_date = body.get('end_date')

        data = oe_detection_handler.get_dashboard_data(start_date=start_date, end_date=end_date)
        if not data.get('success') or not data.get('scores'):
            return jsonify({'success': False, 'error': 'No data to export'}), 404

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OE Detection Scores"

        headers = ['Employee ID', 'Employee Name', 'Score', 'Risk Level',
                    'Domains Hit', 'Signal Count', 'Correlation Multiplier', 'Last Scanned']
        ws.append(headers)

        for score in data['scores']:
            ws.append([
                score['employee_id'],
                score['employee_name'],
                score['normalized_score'],
                score['risk_level'].upper(),
                ', '.join(score.get('domains_hit', [])),
                score.get('signal_count', 0),
                score.get('correlation_multiplier', 1.0),
                score.get('calculated_at', ''),
            ])

        import os
        from openpyxl.styles import Font, Alignment
        _wm_author = os.environ.get("WATERMARK_AUTHOR", "")
        if _wm_author:
            wm_text = f"By {_wm_author}"
            wm_row = ws.max_row + 2
            wm_cell = ws.cell(row=wm_row, column=ws.max_column)
            wm_cell.value = wm_text
            wm_cell.font = Font(size=8, italic=True, color="9E9E9E")
            wm_cell.alignment = Alignment(horizontal='right')
            ws.oddFooter.right.text = wm_text
            ws.oddFooter.right.size = 8
            ws.oddFooter.right.font = "Calibri,Italic"
            ws.oddFooter.right.color = "9E9E9E"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='oe_detection_scores.xlsx',
        )
    except Exception as exc:
        logger.error(f"OE export error: {exc}", exc_info=True)
        return jsonify({'success': False, 'error': str(exc)}), 500
