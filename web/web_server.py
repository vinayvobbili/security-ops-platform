#!/usr/bin/python3

# Standard library imports
import asyncio
import http.client
import http.server
import ipaddress
import json
import logging
import os
import random
import select
import socket
import socketserver
# NEW: ensure project root on sys.path before local imports
import sys
import threading
from datetime import date, datetime, timedelta
from typing import Dict, List
from urllib.parse import urlsplit

_CURRENT_DIR = os.path.dirname(__file__)
_PROJECT_ROOT = os.path.abspath(os.path.join(_CURRENT_DIR, '..'))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# Third-party imports
import pytz
import requests
from flask import Flask, abort, jsonify, render_template, request, session

# Local imports with graceful degradation
try:
    from my_bot.core.my_model import ask, ask_stream  # type: ignore
    from my_bot.core.state_manager import get_state_manager  # type: ignore

    POKEDEX_AVAILABLE = True
except Exception as e:  # Broad except to handle partial dependency failures
    print(f"⚠️ Pokedex components unavailable: {e}. Continuing with limited functionality.")
    POKEDEX_AVAILABLE = False


    def ask(*_args, **_kwargs):  # type: ignore
        return "Model not available in this environment"


    def ask_stream(*_args, **_kwargs):  # type: ignore
        yield "Model not available in this environment"


    def get_state_manager():  # type: ignore
        return None

from my_config import get_config
from services import xsoar
from services.xsoar import XsoarEnvironment
from services.approved_testing_utils import add_approved_testing_entry
from src import secops
from src.components import apt_names_fetcher, secops_shift_metrics
from src.utils.logging_utils import is_scanner_request, log_web_activity, setup_bot_logging

CONFIG = get_config()

# Server configuration constants
SHOULD_START_PROXY = True  # Enable proxy server
USE_DEBUG_MODE = CONFIG.web_server_debug_mode_on
PROXY_PORT = 8081
# Allow environment override; default to 8080 for non-root dev usage
WEB_SERVER_PORT = CONFIG.web_server_port
BUFFER_SIZE = 16384
NUM_WORKERS = 10
MAX_CONNECTIONS = 100
COMPANY_EMAIL_DOMAIN = '@' + CONFIG.my_web_domain

# Configure logging with centralized utility
setup_bot_logging(
    bot_name='web_server',
    log_level=logging.WARNING,
    info_modules=['__main__'],
    use_colors=True
)

# Suppress noisy library logs
logging.getLogger('werkzeug').setLevel(logging.WARNING)
logging.getLogger('waitress').setLevel(logging.WARNING)

app = Flask(__name__, static_folder='static', static_url_path='/static', template_folder='templates')
app.secret_key = CONFIG.flask_secret_key if hasattr(CONFIG, 'flask_secret_key') else 'your-secret-key-change-this'
eastern = pytz.timezone('US/Eastern')

# Store server start time
SERVER_START_TIME = datetime.now(eastern)

blocked_ip_ranges = []  # ["10.49.70.0/24", "10.50.70.0/24"]

# Initialize XSOAR handlers for production environment
prod_list_handler = xsoar.ListHandler(XsoarEnvironment.PROD)
prod_ticket_handler = xsoar.TicketHandler(XsoarEnvironment.PROD)

# Add module logger (was missing previously)
logger = logging.getLogger(__name__)

# --- Server detection utilities (added) ---
from functools import lru_cache

# One-time guard for first-request logic (Flask 3 removed before_first_request)
_server_detection_ran = False


@lru_cache(maxsize=1)
def _runtime_server_info_sample() -> dict:
    """Cached placeholder before first request. Updated on first request."""
    return {
        'server_type': 'unknown_startup',
        'server_software': 'unknown',
        'debug_mode': app.debug,
        'pid': os.getpid()
    }


def detect_server_type(server_software: str) -> str:
    server_software_lower = (server_software or '').lower()
    if 'waitress' in server_software_lower:
        return 'waitress'
    if 'gunicorn' in server_software_lower:
        return 'gunicorn'
    if 'uwsgi' in server_software_lower:
        return 'uwsgi'
    if 'werkzeug' in server_software_lower:
        return 'flask-dev'
    return 'unknown'


# NOTE: Flask 3 removed app.before_first_request; emulate with a guarded before_request.

def _log_real_server():  # noqa: D401
    """Log WSGI server details (invoked once on first real request)."""
    try:
        env = request.environ  # type: ignore[attr-defined]
        server_software = env.get('SERVER_SOFTWARE', 'unknown')
        server_type = detect_server_type(server_software)
        # Update cached info
        _runtime_server_info_sample.cache_clear()  # type: ignore[attr-defined]

        @lru_cache(maxsize=1)
        def _runtime_server_info_sample_override():  # type: ignore
            return {
                'server_type': server_type,
                'server_software': server_software,
                'debug_mode': app.debug,
                'pid': os.getpid()
            }

        globals()['_runtime_server_info_sample'] = _runtime_server_info_sample_override  # swap reference for later calls
        print(f"[ServerDetect] Running under {server_type} ({server_software}) debug={app.debug} pid={os.getpid()}")
    except Exception as e:
        print(f"[ServerDetect] Failed to detect server: {e}")


@app.before_request
def _maybe_server_detect():  # type: ignore[override]
    global _server_detection_ran
    if not _server_detection_ran:
        _log_real_server()
        _server_detection_ran = True


@app.before_request
def block_ip():
    """Block requests from configured IP ranges and scanner probes."""
    if blocked_ip_ranges and any(
            ipaddress.ip_network(request.remote_addr).subnet_of(ipaddress.ip_network(blocked_ip_range))
            for blocked_ip_range in blocked_ip_ranges
    ):
        abort(403)  # Forbidden

    # Check for scanner patterns to quickly return 404 for scanner probes
    if is_scanner_request():
        # Return a very minimal 404 response without logging to reduce overhead
        # This will bypass the normal Flask request handling
        return abort(404)


def get_image_files() -> List[str]:
    """Retrieves a list of image files from the static and charts directories."""
    today_date = datetime.now(eastern).strftime('%m-%d-%Y')
    image_order = [
        "images/Company Logo.png",
        "images/DnR Welcome.png",
        f"charts/{today_date}/Threatcon Level.png",
        f"charts/{today_date}/Days Since Last Incident.png",
        "images/DnR Metrics by Peanuts.jpg",
        f"charts/{today_date}/Aging Tickets.png",
        f"charts/{today_date}/Inflow Yesterday.png",
        # f"charts/{today_date}/Inflow Past 60 Days.png",
        f"charts/{today_date}/Inflow Past 12 Months - Impact Only.png",
        f"charts/{today_date}/Inflow Past 12 Months - Ticket Type Only.png",
        f"charts/{today_date}/Outflow.png",
        f"charts/{today_date}/SLA Breaches.png",
        f"charts/{today_date}/MTTR MTTC.png",
        f"charts/{today_date}/Heat Map.png",
        # f"charts/{today_date}/CrowdStrike Detection Efficacy-Quarter.png",
        # f"charts/{today_date}/CrowdStrike Detection Efficacy-Month.png",
        f"charts/{today_date}/CrowdStrike Detection Efficacy-Week.png",
        # f"charts/{today_date}/QR Rule Efficacy-Quarter.png",
        # f"charts/{today_date}/QR Rule Efficacy-Month.png",
        f"charts/{today_date}/QR Rule Efficacy-Week.png",
        f"charts/{today_date}/Vectra Volume.png",
        # f"charts/{today_date}/CrowdStrike Volume.png",
        # f"charts/{today_date}/Lifespan.png",
        "images/Threat Hunting Intro.png",
        f"charts/{today_date}/Threat Tippers.png",
        f"charts/{today_date}/DE Stories.png",
        f"charts/{today_date}/RE Stories.png",
        "images/End of presentation.jpg",
        "images/Feedback Email.png",
        "images/Thanks.png"
    ]
    image_files = []
    # fetch files per that image order
    for image_path in image_order:
        full_path = os.path.join(app.static_folder, image_path)
        if os.path.exists(full_path):
            image_files.append(image_path)
        else:
            logger.warning(f"File not found: {full_path}")

    return image_files


@app.route("/")
@log_web_activity
def get_ir_dashboard_slide_show():
    """Renders the HTML template with the ordered list of image files."""
    image_files = get_image_files()
    return render_template("slide-show.html", image_files=image_files, show_burger=True)


@app.route("/<path:filename>.pac")
def proxy_pac_file(filename):
    """Handle PAC file requests to reduce log clutter."""
    # Return empty PAC file content for any PAC file request
    pac_content = """function FindProxyForURL(url, host) {
    return "DIRECT";
}"""
    return pac_content, 200, {'Content-Type': 'application/x-ns-proxy-autoconfig'}


@app.route("/msoc-form")
@log_web_activity
def display_msoc_form():
    """Displays the MSOC form."""
    return render_template("msoc_form.html", show_burger=False)


@app.route("/submit-msoc-form", methods=['POST'])
@log_web_activity
def handle_msoc_form_submission():
    """Handles MSOC form submissions and processes the data."""
    form = request.form.to_dict()
    form['type'] = 'MSOC Site Security Device Management'
    response = prod_ticket_handler.create_in_dev(form)
    # Return a JSON response
    return jsonify({
        'status': 'success',
        'new_incident_id': response['id'],
        'new_incident_link': f"{CONFIG.xsoar_dev_ui_base_url}/Custom/caseinfoid/{response['id']}"
    })


@app.route("/speak-up-form")
@log_web_activity
def display_speak_up_form():
    """Displays the Speak Up form."""
    return render_template("speak_up_form.html")


@app.route("/submit-speak-up-form", methods=['POST'])
@log_web_activity
def handle_speak_up_form_submission():
    """Handles the Speak Up form submissions and processes the data."""
    form = request.form.to_dict()

    # Format the date from yyyy-mm-dd to mm/dd/yyyy
    date_occurred = form.get('dateOccurred', '')
    formatted_date = ""
    if date_occurred:
        try:
            # Parse the date string from the form (likely in yyyy-mm-dd format)
            year, month, day = date_occurred.split('-')
            # Format it as mm/dd/yyyy
            formatted_date = f"{month}/{day}/{year}"
        except ValueError:
            # If there's an error parsing, use the original value
            formatted_date = date_occurred

    form['name'] = 'Speak Up Report'
    form['type'] = f'{CONFIG.team_name} Employee Reported Incident'
    form['details'] = (
        f"Date Occurred: {formatted_date} \n"
        f"Issue Type: {form.get('issueType')} \n"
        f"Description: {form.get('description')} \n"
    )
    response = prod_ticket_handler.create(form)
    # Return a JSON response
    return jsonify({
        'status': 'success',
        'new_incident_id': response['id'],
        'new_incident_link': f"{CONFIG.xsoar_dev_ui_base_url}/Custom/caseinfoid/{response['id']}"
    })


@app.route('/xsoar-ticket-import-form', methods=['GET'])
@log_web_activity
def xsoar_ticket_import_form():
    return render_template('xsoar-ticket-import-form.html')


@app.route("/import-xsoar-ticket", methods=['POST'])
@log_web_activity
def import_xsoar_ticket():
    source_ticket_number = request.form.get('source_ticket_number')
    destination_ticket_number, destination_ticket_link = xsoar.import_ticket(source_ticket_number)
    return jsonify({
        'source_ticket_number': source_ticket_number,
        'destination_ticket_number': destination_ticket_number,
        'destination_ticket_link': destination_ticket_link
    })


@app.route("/get-approved-testing-entries", methods=['GET'])
@log_web_activity
def get_approved_testing_entries():
    """Fetches approved testing records and displays them in separate HTML tables."""
    approved_testing_records = prod_list_handler.get_list_data_by_name(f'{CONFIG.team_name}_Approved_Testing')

    if not approved_testing_records:
        return "<h2>No Approved Testing Records Found</h2>"

    # Organize data for the template
    endpoints = approved_testing_records.get("ENDPOINTS", [])
    usernames = approved_testing_records.get("USERNAMES", [])
    ip_addresses = approved_testing_records.get("IP_ADDRESSES", [])
    cidr_blocks = approved_testing_records.get("CIDR_BLOCKS", [])

    # Render the template with the data
    return render_template(
        'approved_testing.html',
        ENDPOINTS=endpoints,
        USERNAMES=usernames,
        IP_ADDRESSES=ip_addresses,
        CIDR_BLOCKS=cidr_blocks
    )


def parse_date(date_str: str) -> datetime:
    """Parses a date string in multiple formats and returns a datetime object."""
    for fmt in ('%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"Date format not recognized: {date_str}")


@app.route("/get-current-upcoming-travel-records", methods=['GET'])
@log_web_activity
def get_upcoming_travel():
    """Fetches upcoming travel records and displays them."""
    upcoming_travel_records = [
        record for record in prod_list_handler.get_list_data_by_name('DnR_Upcoming_Travel')
        if parse_date(record['vacation_end_date']) >= datetime.now()
    ]

    return render_template(
        'upcoming_travel.html',
        travel_records=upcoming_travel_records
    )


@app.route("/travel-form")
@log_web_activity
def display_travel_form():
    """Displays the Upcoming Travel Notification form."""
    today = date.today().isoformat()  # Get today's date in 'YYYY-MM-DD' format
    return render_template("upcoming_travel_notification_form.html", today=today)


@app.route("/submit-travel-form", methods=['POST'])
@log_web_activity
def handle_travel_form_submission():
    """Handles the Upcoming Travel Notification form submissions and processes the data."""
    form = request.form.to_dict()

    # Submit to prod_list_handler instead of prod_incident_handler
    response = prod_list_handler.add_item_to_list('DnR_Upcoming_Travel', {
        "traveller_email_address": form.get('traveller_email_address'),
        "work_location": form.get('work_location'),
        "vacation_location": form.get('vacation_location'),
        "vacation_start_date": form.get('vacation_start_date'),
        "vacation_end_date": form.get('vacation_end_date'),
        "is_working_during_vacation": form.get('will_work_during_vacation'),
        "comments": form.get('comments'),
        "submitted_at": datetime.now(eastern).strftime("%m/%d/%Y %I:%M %p %Z"),
        "submitted_by_ip_address": request.remote_addr
    })

    return jsonify({
        'status': 'success',
        'response': response
    })


@app.route("/red-team-testing-form")
@log_web_activity
def display_red_team_testing_form():
    """Displays the Red Team Testing form."""
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    return render_template("red_team_testing_form.html", tomorrow=tomorrow)


@app.route("/submit-red-team-testing-form", methods=['POST'])
@log_web_activity
def handle_red_team_testing_form_submission():
    """Handles the Red Team Testing form submissions and processes the data."""
    form = request.form.to_dict()

    usernames = form.get('usernames', '').strip()
    tester_hosts = form.get('tester_hosts', '').strip()
    targets = form.get('targets', '').strip()
    description = form.get('description', '').strip()
    notes_scope = form.get('notes_scope', '').strip()
    keep_until = form.get('keep_until', '')
    submitter_ip_address = request.remote_addr
    submitter_email_address = form.get('email_local', '') + COMPANY_EMAIL_DOMAIN
    submit_date = datetime.now(eastern).strftime("%m/%d/%Y")

    approved_testing_list_name = f"{CONFIG.team_name}_Approved_Testing"
    approved_testing_master_list_name = f"{CONFIG.team_name}_Approved_Testing_MASTER"
    try:
        add_approved_testing_entry(
            prod_list_handler,
            approved_testing_list_name,
            approved_testing_master_list_name,
            usernames,
            tester_hosts,
            targets,
            description,
            notes_scope,
            submitter_email_address,
            keep_until,
            submit_date,
            submitter_ip_address
        )
    except ValueError as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }, 400)

    return jsonify({
        'status': 'success'
    })


@app.route('/favicon.ico')
def favicon():
    """Serve the favicon icon."""
    return app.send_static_file('icons/favicon.ico')


# HTTP connection pool for reusing connections
class ConnectionPool:
    def __init__(self, max_connections=100):
        self.pool: Dict[str, List[http.client.HTTPConnection]] = {}
        self.max_connections = max_connections
        self.lock = threading.RLock()

    def get_connection(self, host):
        with self.lock:
            if host not in self.pool:
                self.pool[host] = []

            if self.pool[host]:
                return self.pool[host].pop()
            else:
                return http.client.HTTPConnection(host)

    def release_connection(self, host, conn):
        with self.lock:
            if host not in self.pool:
                self.pool[host] = []

            if len(self.pool[host]) < self.max_connections:
                self.pool[host].append(conn)
            else:
                conn.close()


# Create a global connection pool
connection_pool = ConnectionPool(max_connections=MAX_CONNECTIONS)


# Optimized proxy class with async support
def _relay_sockets(client, target):
    """Simple synchronous socket relay that avoids HTTP processing"""
    try:
        # Set non-blocking mode to handle disconnections gracefully
        client.settimeout(1.0)
        target.settimeout(1.0)

        sockets = [client, target]

        # Keep transferring data between client and target
        while True:
            try:
                # Check if sockets are still valid before select
                if client.fileno() == -1 or target.fileno() == -1:
                    break

                # Wait until a socket is ready to be read
                readable, _, exceptional = select.select(sockets, [], sockets, 1.0)

                if exceptional:
                    break

                if not readable:
                    continue  # Timeout, try again

                for sock in readable:
                    # Check socket validity again
                    if sock.fileno() == -1:
                        continue

                    # Determine the destination socket
                    dest = target if sock is client else client

                    # Check destination socket validity
                    if dest.fileno() == -1:
                        return

                    try:
                        data = sock.recv(BUFFER_SIZE)
                        if not data:
                            return  # Connection closed
                        dest.sendall(data)
                    except (socket.error, ConnectionResetError, BrokenPipeError, OSError) as e:
                        if e.errno == 9:  # Bad file descriptor
                            return
                        return  # Any socket error means we're done
            except (OSError, ValueError) as e:
                # Handle select errors gracefully
                if hasattr(e, 'errno') and e.errno == 9:  # Bad file descriptor
                    break
                return
    except Exception:
        # Catch any other unexpected errors
        pass
    finally:
        # Ensure sockets are properly closed
        for sock in [client, target]:
            try:
                if sock and sock.fileno() != -1:
                    sock.close()
            except (OSError, AttributeError):
                pass


async def _async_select(client_sock, target_sock):
    """Async-compatible version of select operation using running loop (Python 3.13 safe)."""
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        # Fallback: no running loop (shouldn't generally happen here)
        loop = asyncio.get_event_loop_policy().new_event_loop()
    readable = []
    for sock in [client_sock, target_sock]:
        try:
            if await loop.sock_recv(sock, 1):
                readable.append(sock)
        except (ConnectionError, OSError, asyncio.CancelledError):
            pass
    return readable


def relay_data_async(client_sock, target_sock):
    """Efficiently relays data bidirectionally between client_sock and target_sock.
    Uses a dedicated event loop to avoid deprecated get_event_loop() patterns.
    """
    loop = asyncio.new_event_loop()
    try:
        asyncio.set_event_loop(loop)
        # Use two separate buffers for better performance
        client_to_target = bytearray(BUFFER_SIZE)
        target_to_client = bytearray(BUFFER_SIZE)

        while True:
            try:
                r = loop.run_until_complete(
                    asyncio.wait_for(_async_select(client_sock, target_sock), timeout=2.0)
                )
            except asyncio.TimeoutError:
                r = []
            except Exception:
                break

            if not r:
                # Check if connections are still alive
                if client_sock.fileno() == -1 or target_sock.fileno() == -1:
                    break

            if client_sock in r:
                view = memoryview(client_to_target)
                bytes_read = client_sock.recv_into(view)
                if not bytes_read:
                    break
                target_sock.sendall(view[:bytes_read])

            if target_sock in r:
                view = memoryview(target_to_client)
                bytes_read = target_sock.recv_into(view)
                if not bytes_read:
                    break
                client_sock.sendall(view[:bytes_read])
    except Exception as e:
        logger.error(f"Error during relay: {e}")
    finally:
        try:
            loop.close()
        except Exception:
            pass
        for sock in [client_sock, target_sock]:
            try:
                sock.close()
            except (OSError, AttributeError):
                pass


class OptimizedProxy(http.server.SimpleHTTPRequestHandler):
    protocol_version = 'HTTP/1.1'  # Enable keep-alive

    def do_GET(self):
        self.proxy_http_request()

    def do_POST(self):
        self.proxy_http_request()

    def do_PUT(self):
        self.proxy_http_request()

    def do_DELETE(self):
        self.proxy_http_request()

    def do_HEAD(self):
        self.proxy_http_request()

    def do_CONNECT(self):
        target_host = "unknown"
        target_port = "unknown"

        try:
            # Parse target address
            target_host, target_port = self.path.split(':', 1)
            target_port = int(target_port)

            timestamp = datetime.now(eastern).strftime('%Y-%m-%d %H:%M:%S %Z')
            client_ip = self.client_address[0] if hasattr(self, 'client_address') else 'Unknown'
            logger.info(f"[{timestamp}] CONNECT request from {client_ip} to {target_host}:{target_port}")

            # Connect to target server
            target_sock = socket.create_connection((target_host, target_port), timeout=30)

            # Send 200 Connection Established response
            try:
                self.wfile.write(b"HTTP/1.1 200 Connection established\r\n\r\n")
                self.wfile.flush()
            except (BrokenPipeError, ConnectionResetError, OSError):
                target_sock.close()
                return

            # Create connection pipes
            client_socket = self.connection

            # Set reasonable timeouts
            try:
                client_socket.settimeout(30)
                target_sock.settimeout(30)
            except (OSError, AttributeError):
                target_sock.close()
                return

            # Start bidirectional relay
            _relay_sockets(client_socket, target_sock)

        except ValueError:
            # Invalid target address format
            try:
                self.send_error(400, "Bad Request: Invalid target address")
            except (BrokenPipeError, ConnectionResetError, OSError):
                pass
        except (socket.timeout, socket.gaierror):
            # Connection timeout or DNS resolution error
            try:
                self.send_error(502, f"Cannot connect to {target_host}:{target_port}")
            except (BrokenPipeError, ConnectionResetError, OSError):
                pass
        except (ConnectionRefusedError, OSError) as e:
            # Connection refused or other OS-level errors
            try:
                if hasattr(e, 'errno') and e.errno == 9:  # Bad file descriptor
                    return  # Client already disconnected
                self.send_error(502, f"Cannot connect to {target_host}:{target_port}")
            except (BrokenPipeError, ConnectionResetError, OSError):
                pass
        except Exception as e:
            # Catch any other unexpected errors
            try:
                logger.error(f"CONNECT error: {e}")
                self.send_error(502, f"Cannot connect to {target_host}:{target_port}")
            except (BrokenPipeError, ConnectionResetError, OSError):
                pass

    def proxy_http_request(self):
        # This part handles regular HTTP requests (not HTTPS via CONNECT)
        url = self.path

        if url.startswith('https://'):
            logger.warning(f"Client tried to send HTTPS directly. Use CONNECT for HTTPS tunneling")
            self.send_error(501, "HTTPS GET/POST proxy not implemented (use CONNECT)")
            return

        try:
            parts = urlsplit(url)
            netloc = parts.netloc
            path = parts.path
            query = parts.query
            fragment = parts.fragment

            full_path = path
            if query:
                full_path += '?' + query
            if fragment:
                full_path += '#' + fragment

            # Check if client accepts gzip encoding
            accept_encoding = self.headers.get('Accept-Encoding', '')
            supports_gzip = 'gzip' in accept_encoding.lower()

            headers = {}
            for h in self.headers:
                if h.lower() not in ['proxy-connection', 'transfer-encoding', 'connection']:
                    headers[h] = self.headers[h]

            # Add support for gzip if client accepts it
            if supports_gzip and 'Accept-Encoding' not in headers:
                headers['Accept-Encoding'] = 'gzip'

            # Use connection pool
            conn = connection_pool.get_connection(netloc)

            if self.command == 'GET':
                conn.request(self.command, full_path, headers=headers)
            else:
                content_length = int(self.headers.get('Content-Length', 0))
                body = self.rfile.read(content_length) if content_length > 0 else None
                conn.request(self.command, full_path, body=body, headers=headers)

            response = conn.getresponse()

            self.send_response(response.status)

            # Prepare for potential gzip compression
            for h, v in response.getheaders():
                if h.lower() not in ['transfer-encoding', 'connection']:
                    self.send_header(h, v)

            self.end_headers()

            # Read response content in larger chunks for better performance
            content = b''
            while True:
                chunk = response.read(BUFFER_SIZE)
                if not chunk:
                    break
                content += chunk

            # Return the connection to the pool
            connection_pool.release_connection(netloc, conn)

            # Send response data to client
            self.wfile.write(content)

        except (BrokenPipeError, ConnectionResetError, OSError) as e:
            if hasattr(e, 'errno') and e.errno == 9:  # Bad file descriptor
                return  # Client already disconnected, no need to respond
            logger.error(f"Connection error during HTTP proxy request: {e}")
            try:
                self.send_error(502, "Bad Gateway")
            except (BrokenPipeError, ConnectionResetError, OSError):
                pass
        except Exception as e:
            logger.error(f"Error during HTTP proxy request: {e}")
            try:
                self.send_error(502, "Bad Gateway")
            except (BrokenPipeError, ConnectionResetError, OSError):
                pass


# Add a function to start the optimized proxy server
def start_proxy_server():
    handler = OptimizedProxy
    print(f"Starting optimized proxy on port {PROXY_PORT}")
    try:
        # Enable address reuse to avoid "address already in use" errors
        socketserver.TCPServer.allow_reuse_address = True
        with socketserver.ThreadingTCPServer(("", PROXY_PORT), handler) as httpd:  # type: ignore[arg-type]
            httpd.serve_forever()
    except Exception as e:
        print(f"Failed to start proxy: {e}")


@app.route("/api/apt-names", methods=["GET"])
@log_web_activity
def api_apt_names():
    """API endpoint to get APT workbook summary (region sheets only)."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, '../data/transient/de/APTAKAcleaned.xlsx')
    file_path = os.path.abspath(file_path)
    info = apt_names_fetcher.get_workbook_info(file_path)
    return jsonify(info)


@app.route("/api/apt-other-names", methods=["GET"])
@log_web_activity
def api_apt_other_names():
    """API endpoint to get other names for a given APT common name."""
    common_name = request.args.get("common_name")
    app.logger.info(f"[APT API] Received common_name: '{common_name}' from request")
    if not common_name:
        return jsonify({"error": "Missing required parameter: common_name"}), 400
    should_include_metadata = request.args.get("should_include_metadata")
    if should_include_metadata is not None:
        should_include_metadata = should_include_metadata.lower() == "true"
    else:
        should_include_metadata = False
    # Support both 'format' and 'response_format' query parameters
    response_format = request.args.get("response_format")
    if not response_format:
        response_format = request.args.get("format", "html")
    response_format = response_format.lower()
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, '../data/transient/de/APTAKAcleaned.xlsx')
    file_path = os.path.abspath(file_path)
    app.logger.info(f"[APT API] Calling get_other_names_for_common_name with common_name: '{common_name}', file_path: '{file_path}', should_include_metadata: {should_include_metadata}")
    results = apt_names_fetcher.get_other_names_for_common_name(common_name, file_path, should_include_metadata)
    app.logger.info(f"[APT API] Results returned: {results}")
    if response_format == "json":
        return jsonify(results)
    else:
        return render_template("apt_other_names_results.html", common_name=common_name, results=results, should_include_metadata=should_include_metadata)


@app.route("/apt-other-names-search", methods=["GET"])
@log_web_activity
def apt_other_names_search():
    """Render the APT Other Names search form page."""
    # Get all APT names for dropdown
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, '../data/transient/de/APTAKAcleaned.xlsx')
    file_path = os.path.abspath(file_path)

    try:
        apt_names = apt_names_fetcher.get_all_apt_names(file_path)
        app.logger.info(f"[APT Search] Loaded {len(apt_names)} APT names for dropdown")
    except Exception as e:
        app.logger.error(f"[APT Search] Error loading APT names: {str(e)}")
        apt_names = []

    return render_template("apt_other_names_search.html", apt_names=apt_names)


@app.route('/api/random-audio', methods=['GET'])
def random_audio():
    """Return a random mp3 filename from the static/audio directory."""
    audio_dir = os.path.join(os.path.dirname(__file__), 'static', 'audio')
    files = [f for f in os.listdir(audio_dir) if f.endswith('.mp3')]
    if not files:
        return jsonify({'error': 'No audio files found'}), 404
    return jsonify({'filename': random.choice(files)})


@app.route('/pokedex')
@log_web_activity
def pokedex_chat():
    """Pokedex AI chat interface"""
    return render_template('pokedex_chat.html')


@app.route('/api/pokedex-status')
def api_pokedex_status():
    """Health check endpoint for Pokédex chat availability"""
    try:

        state_manager = get_state_manager()

        if not state_manager or not state_manager.is_initialized:
            return jsonify({
                'ready': False,
                'status': 'not_initialized',
                'message': 'Pokedex chat is not available. Please ensure all components are initialized.',
                'instructions': [
                    'Restart the web server with ENABLE_POKEDEX_CHAT = True'
                ]
            })

        # Perform health check
        health = state_manager.health_check()
        components = health.get('components', {})

        # Check if core LLM components are working (LLM and embeddings)
        # RAG is optional - chat works fine without it
        core_ready = components.get('llm', False) and components.get('embeddings', False)

        if health.get('status') == 'initialized':
            return jsonify({
                'ready': True,
                'status': 'healthy',
                'message': 'Pokedex chat is ready',
                'components': components
            })
        elif core_ready:
            # LLM works, but RAG might be missing - still usable!
            return jsonify({
                'ready': True,
                'status': 'partial',
                'message': 'Pokedex chat is ready (without document search)',
                'components': components
            })
        else:
            return jsonify({
                'ready': False,
                'status': 'partial',
                'message': 'Pokedex chat core components not ready',
                'components': components
            })

    except Exception as e:
        logger.error(f"Error checking Pokedex status: {e}")
        return jsonify({
            'ready': False,
            'status': 'error',
            'message': 'Error checking chat status',
            'error': str(e)
        })


@app.route('/api/pokedex-chat', methods=['POST'])
@log_web_activity
def api_pokedex_chat():
    """API endpoint for Pokedex chat messages"""
    try:
        data = request.get_json()
        user_message = data.get('message', '').strip()
        session_id = data.get('session_id', '')

        if not user_message:
            return jsonify({'success': False, 'error': 'Message is required'}), 400

        if not session_id:
            return jsonify({'success': False, 'error': 'Session ID is required'}), 400

        # Use IP address + session ID as identifier for context management
        # This mimics how Webex bot uses person_id + room_id
        user_ip = request.remote_addr
        user_identifier = f"web_{user_ip}_{session_id}"

        # Use the ask function to get LLM response
        # The ask function handles conversation context via session management
        try:
            response_text = ask(
                user_message,
                user_id=user_identifier,
                room_id="web_chat"  # Use a common room_id for web chats
            )
        except Exception as e:
            logger.error(f"Error getting LLM response: {e}", exc_info=True)
            return jsonify({
                'success': False,
                'error': 'Failed to get response from AI. Please try again.'
            }), 500

        return jsonify({
            'success': True,
            'response': response_text
        })

    except Exception as e:
        logger.error(f"Error in Pokedex chat API: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'An unexpected error occurred. Please try again.'
        }), 500


@app.route('/api/pokedex-chat-stream', methods=['POST'])
@log_web_activity
def api_pokedex_chat_stream():
    """Streaming API endpoint for Pokédex chat messages using Server-Sent Events"""
    try:
        data = request.get_json()
        user_message = data.get('message', '').strip()
        session_id = data.get('session_id', '')

        if not user_message:
            return jsonify({'success': False, 'error': 'Message is required'}), 400

        if not session_id:
            return jsonify({'success': False, 'error': 'Session ID is required'}), 400

        # Clean up old sessions before processing (removes messages older than 2 hours)
        state_manager = get_state_manager()
        if state_manager and hasattr(state_manager, 'session_manager'):
            state_manager.session_manager.cleanup_old_sessions()

        # Use IP address + session ID as identifier
        user_ip = request.remote_addr
        user_identifier = f"web_{user_ip}_{session_id}"

        def generate():
            """Generator function for Server-Sent Events"""
            try:
                for token in ask_stream(
                        user_message,
                        user_id=user_identifier,
                        room_id="web_chat"
                ):
                    # Format as Server-Sent Event
                    yield f"data: {json.dumps({'token': token})}\n\n"

                # Send completion event
                yield f"data: {json.dumps({'done': True})}\n\n"

            except Exception as e:
                logger.error(f"Error in streaming response: {e}", exc_info=True)
                yield f"data: {json.dumps({'error': 'Streaming error occurred'})}\n\n"

        # Return SSE response
        # Note: 'Connection' header is removed as it's a hop-by-hop header
        # not allowed in WSGI applications (PEP 3333)
        return app.response_class(
            generate(),
            mimetype='text/event-stream',
            headers={
                'Cache-Control': 'no-cache',
                'X-Accel-Buffering': 'no'  # Disable buffering in nginx
            }
        )

    except Exception as e:
        logger.error(f"Error in Pokedex streaming chat API: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'An unexpected error occurred. Please try again.'
        }), 500


@app.route('/xsoar')
@log_web_activity
def xsoar_dashboard():
    """XSOAR incident dashboard"""
    return render_template('xsoar_dashboard.html')


@app.route('/api/xsoar/incidents')
@log_web_activity
def api_xsoar_incidents():
    """API to get XSOAR incidents with search and pagination"""
    query = request.args.get('query', '')
    period = request.args.get('period')
    size = int(request.args.get('size', 50))

    try:
        incidents = prod_ticket_handler.get_tickets(query, period, size)
        return jsonify({'success': True, 'incidents': incidents})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/xsoar/incident/<incident_id>')
@log_web_activity
def api_xsoar_incident_detail(incident_id):
    """API to get XSOAR incident details"""
    try:
        incident = prod_ticket_handler.get_case_data(incident_id)
        entries = prod_ticket_handler.get_entries(incident_id)
        return jsonify({'success': True, 'incident': incident, 'entries': entries})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/xsoar/incident/<incident_id>')
@log_web_activity
def xsoar_incident_detail(incident_id):
    """XSOAR incident detail view"""
    try:
        incident = prod_ticket_handler.get_case_data(incident_id)
        entries = prod_ticket_handler.get_entries(incident_id)
        return render_template('xsoar_incident_detail.html',
                               incident=incident, entries=entries)
    except requests.exceptions.HTTPError as e:
        return f"XSOAR API Error for incident {incident_id}: HTTP {e.response.status_code} - {e.response.text}", 500
    except requests.exceptions.ConnectionError as e:
        return f"Connection Error for incident {incident_id}: {str(e)}", 500
    except ValueError as e:
        return f"Invalid JSON response for incident {incident_id}: {str(e)}", 500
    except Exception as e:
        return f"Error loading incident {incident_id}: {str(e)}", 500


@app.route('/api/xsoar/incident/<incident_id>/entries')
@log_web_activity
def api_xsoar_incident_entries(incident_id):
    """API to get incident entries/comments"""
    try:
        entries = prod_ticket_handler.get_entries(incident_id)
        return jsonify({'success': True, 'entries': entries})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/xsoar/incident/<incident_id>/link', methods=['POST'])
@log_web_activity
def api_xsoar_link_incident(incident_id):
    """API to link incidents"""
    link_incident_id = request.json.get('link_incident_id')
    try:
        result = prod_ticket_handler.link_tickets(incident_id, link_incident_id)
        return jsonify({'success': True, 'result': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/xsoar/incident/<incident_id>/participant', methods=['POST'])
@log_web_activity
def api_xsoar_add_participant(incident_id):
    """API to add participant to incident"""
    email = request.json.get('email')
    try:
        result = prod_ticket_handler.add_participant(incident_id, email)
        return jsonify({'success': True, 'result': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/shift-performance')
@log_web_activity
def shift_performance_dashboard():
    """Display shift performance page - loads instantly with empty structure"""
    return render_template('shift_performance.html', xsoar_prod_ui_base=getattr(CONFIG, 'xsoar_prod_ui_base_url', 'https://msoar.crtx.us.paloaltonetworks.com'))


def _shift_has_passed(shift_name: str, current_shift: str) -> bool:
    """Check if a shift has already passed based on current shift."""
    shift_order = ['morning', 'afternoon', 'night']
    try:
        shift_index = shift_order.index(shift_name)
        current_index = shift_order.index(current_shift)
        return shift_index < current_index
    except ValueError:
        return False


@app.route('/api/shift-list')
@log_web_activity
def get_shift_list():
    """Single source of truth for shift performance data.

    Returns comprehensive shift data for the past week including:
    - Ticket counts (inflow/outflow)
    - Full ticket details (inflow_tickets, outflow_tickets arrays)
    - Metrics (MTTR, MTTC, SLA breaches)
    - Staffing info
    - Performance scores

    Architecture:
    - Frontend loads this data ONCE on page load
    - Frontend stores data in globalShiftData
    - Table displays summary counts from this data
    - Details modal slices/dices the SAME data (no additional API calls)
    - All counts derived from actual ticket arrays for consistency
    """
    status = 'unknown'
    try:
        shift_data = []
        shifts = ['morning', 'afternoon', 'night']
        for days_back in range(7):
            target_date = datetime.now(eastern) - timedelta(days=days_back)
            day_name = target_date.strftime('%A')
            date_str = target_date.strftime('%Y-%m-%d')
            for shift_name in shifts:
                # Special handling for night shift date
                # Night shift runs from 20:30 to 04:30 next day
                # If we're between 00:00 and 04:30, the night shift started yesterday
                shift_date = target_date
                shift_day_name = day_name
                shift_date_str = date_str
                if shift_name == 'night' and days_back == 0:
                    now = datetime.now(eastern)
                    if now.hour < 4 or (now.hour == 4 and now.minute < 30):
                        # We're in the night shift that started yesterday
                        shift_date = target_date - timedelta(days=1)
                        shift_day_name = shift_date.strftime('%A')
                        shift_date_str = shift_date.strftime('%Y-%m-%d')
                try:
                    staffing = secops.get_staffing_data(shift_day_name, shift_name)
                    total_staff = sum(len(staff) for team, staff in staffing.items() if team != 'On-Call' and staff != ['N/A (Excel file missing)'])
                    shift_id = f"{shift_date_str}_{shift_name}"
                    current_shift = secops.get_current_shift()
                    if days_back > 0:
                        status = 'completed'
                        show_shift = True
                    elif days_back == 0:
                        if shift_name.lower() == current_shift:
                            status = 'active'
                            show_shift = True
                        elif _shift_has_passed(shift_name.lower(), current_shift):
                            status = 'completed'
                            show_shift = True
                        else:
                            show_shift = False
                    else:
                        show_shift = False
                    if show_shift:
                        # Get staffing and security actions
                        shift_hour_map = {'morning': 4.5, 'afternoon': 12.5, 'night': 20.5}
                        shift_start_hour = shift_hour_map.get(shift_name.lower(), 4.5)
                        # Adjust days_back for night shift metrics when we're after midnight
                        adjusted_days_back = days_back
                        if shift_name == 'night' and days_back == 0:
                            now = datetime.now(eastern)
                            if now.hour < 4 or (now.hour == 4 and now.minute < 30):
                                adjusted_days_back = 1
                        security_actions = secops.get_shift_security_actions(adjusted_days_back, shift_start_hour)

                        detailed_staffing = secops.get_staffing_data(shift_day_name, shift_name)
                        basic_staffing = secops.get_basic_shift_staffing(shift_day_name, shift_name.lower())

                        # Determine shift lead (prefer first SA)
                        sa_list = detailed_staffing.get('SA') or detailed_staffing.get('senior_analysts') or []
                        shift_lead = None
                        if isinstance(sa_list, list) and sa_list:
                            first_sa = sa_list[0]
                            if first_sa and 'N/A' not in str(first_sa):
                                shift_lead = str(first_sa)
                        if not shift_lead:
                            shift_lead = secops.get_shift_lead(shift_day_name, shift_name)
                        if not shift_lead or 'N/A' in str(shift_lead):
                            shift_lead = 'N/A'

                        # Calculate all shift metrics using the component
                        base_date = datetime(shift_date.year, shift_date.month, shift_date.day)
                        metrics = secops_shift_metrics.get_shift_metrics(
                            date_obj=base_date,
                            shift_name=shift_name,
                            ticket_handler=prod_ticket_handler,
                            day_name=shift_day_name,
                            total_staff=total_staff,
                            security_actions=security_actions,
                            shift_lead=shift_lead,
                            basic_staffing=basic_staffing,
                            detailed_staffing=detailed_staffing
                        )

                        # Build shift data entry
                        shift_data.append({
                            'id': shift_id,
                            'date': shift_date_str,
                            'day': metrics['day'],
                            'shift': shift_name.title(),
                            'total_staff': metrics['total_staff'],
                            'actual_staff': metrics['actual_staff'],
                            'status': status,
                            'tickets_acknowledged': metrics['tickets_acknowledged'],
                            'tickets_closed': metrics['tickets_closed'],
                            'response_time_minutes': metrics['response_time_minutes'],
                            'contain_time_minutes': metrics['contain_time_minutes'],
                            'response_sla_breaches': metrics['response_sla_breaches'],
                            'containment_sla_breaches': metrics['containment_sla_breaches'],
                            'security_actions': metrics['security_actions'],
                            'mtp_ticket_ids': metrics['mtp_ticket_ids'],
                            'inflow_tickets': metrics['inflow_tickets'],
                            'outflow_tickets': metrics['outflow_tickets'],
                            'shift_lead': metrics['shift_lead'],
                            'basic_staffing': metrics['basic_staffing'],
                            'detailed_staffing': metrics['detailed_staffing'],
                            'score': metrics['score']
                        })
                except Exception as e:
                    logger.error(f"Error getting staffing or metrics for {shift_day_name} {shift_name}: {e}")
                    current_shift = secops.get_current_shift()
                    if days_back > 0:
                        show_shift = True
                        status = 'error'
                    elif days_back == 0:
                        if shift_name.lower() == current_shift:
                            show_shift = True
                            status = 'error'
                        elif _shift_has_passed(shift_name.lower(), current_shift):
                            show_shift = True
                            status = 'error'
                        else:
                            show_shift = False
                    else:
                        show_shift = False
                    if show_shift:
                        shift_id = f"{shift_date_str}_{shift_name}"
                        shift_data.append({
                            'id': shift_id,
                            'date': shift_date_str,
                            'day': shift_day_name,
                            'shift': shift_name.title(),
                            'total_staff': 0,
                            'actual_staff': 0,
                            'status': status,
                            'tickets_acknowledged': 0,
                            'tickets_closed': 0,
                            'response_time_minutes': 0,
                            'contain_time_minutes': 0,
                            'response_sla_breaches': 0,
                            'containment_sla_breaches': 0,
                            'mtp_ticket_ids': '',
                            'inflow_tickets': [],
                            'outflow_tickets': [],
                            'shift_lead': 'N/A',
                            'basic_staffing': {'total_staff': 0, 'teams': {}},
                            'detailed_staffing': {},
                            'security_actions': {'iocs_blocked': 0, 'domains_blocked': 0, 'malicious_true_positives': 0},
                            'score': 0
                        })
        result = {'success': True, 'data': shift_data}
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/clear-cache', methods=['POST'])
@log_web_activity
def clear_shift_cache():
    """No-op endpoint for compatibility with frontend cache clearing.
    Backend caching has been removed - only frontend localStorage cache exists."""
    return jsonify({'success': True, 'message': 'No backend cache (frontend-only caching)'})


@app.route('/meaningful-metrics')
@log_web_activity
def meaningful_metrics():
    """Meaningful Metrics Dashboard"""
    return render_template('meaningful_metrics.html')


@app.route('/api/meaningful-metrics/data')
@log_web_activity
def api_meaningful_metrics_data():
    """API to get cached security incident data for dashboard."""
    import json
    from pathlib import Path
    try:
        today_date = datetime.now(eastern).strftime('%m-%d-%Y')
        root_directory = Path(__file__).parent.parent
        cache_file = root_directory / 'data' / 'transient' / 'secOps' / today_date / 'past_90_days_tickets.json'
        if not cache_file.exists():
            return jsonify({'success': False, 'error': 'Cache file not found'}), 404
        with open(cache_file, 'r') as f:
            cached_data = json.load(f)
        if isinstance(cached_data, dict) and 'data' in cached_data:
            return jsonify({
                'success': True,
                'data': cached_data['data'],
                'total_count': cached_data.get('total_count', len(cached_data['data'])),
                'data_generated_at': cached_data.get('data_generated_at')
            })
        else:
            return jsonify({
                'success': True,
                'data': cached_data,
                'total_count': len(cached_data),
                'data_generated_at': None
            })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/meaningful-metrics/export', methods=['POST'])
@log_web_activity
def api_meaningful_metrics_export():
    """Server-side Excel export with professional formatting."""
    import pandas as pd
    from flask import send_file
    from src.utils.excel_formatting import apply_professional_formatting
    import tempfile

    try:
        # Get the filtered data from the request
        data = request.get_json()
        if not data or 'incidents' not in data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        incidents = data['incidents']
        visible_columns = data.get('visible_columns', [])
        column_labels = data.get('column_labels', {})

        if not incidents:
            return jsonify({'success': False, 'error': 'No incidents to export'}), 400

        # Process the data similar to frontend
        severity_map = {0: 'Unknown', 1: 'Low', 2: 'Medium', 3: 'High', 4: 'Critical'}
        status_map = {0: 'Pending', 1: 'Active', 2: 'Closed'}

        # Prepare rows for export
        max_cell_length = 32767
        rows = []

        for incident in incidents:
            row = {}
            for col_id in visible_columns:
                # Get the value based on column path
                value = incident.get(col_id)
                # Use label if provided, otherwise use col_id
                col_label = column_labels.get(col_id, col_id)

                # Handle special formatting
                if col_id == 'notes':
                    # Format notes with truncation (notes are already pre-formatted from xsoar.get_user_notes)
                    if isinstance(value, list) and value:
                        truncation_message = '\n\n[... Content truncated due to Excel cell size limit. Please view full notes in the web interface ...]'
                        reserved_length = len(truncation_message) + 100

                        notes_text = ''
                        total_length = 0
                        truncated = False

                        for idx, note in enumerate(value):
                            # Notes are pre-formatted with note_text, author, created_at from xsoar.get_user_notes()
                            note_text = note.get('note_text', '')
                            author = note.get('author', '')
                            timestamp = note.get('created_at', '')
                            formatted_note = f"{idx + 1}. Note: {note_text}\nAuthor: {author}\nTimestamp: {timestamp}"
                            separator = '\n\n' if idx > 0 else ''
                            next_chunk = separator + formatted_note

                            if total_length + len(next_chunk) + reserved_length > max_cell_length:
                                truncated = True
                                break

                            notes_text += next_chunk
                            total_length += len(next_chunk)

                        if truncated:
                            notes_text += truncation_message

                        value = notes_text
                    else:
                        value = ''
                elif col_id == 'severity':
                    value = severity_map.get(value, 'Unknown')
                elif col_id == 'status':
                    value = status_map.get(value, 'Unknown')
                elif col_id in ['created', 'modified', 'closed', 'updated'] and value:
                    # Format date as MM/DD/YYYY HH:MM AM ET
                    try:
                        from datetime import datetime
                        import pytz
                        # Parse the date (handle various formats)
                        if isinstance(value, str):
                            # Try to parse ISO format or other common formats
                            try:
                                dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                            except:
                                dt = datetime.strptime(value, '%Y-%m-%dT%H:%M:%S.%fZ')
                        else:
                            dt = value

                        # Convert to Eastern Time
                        if dt.tzinfo is None:
                            dt = pytz.utc.localize(dt)
                        et_tz = pytz.timezone('US/Eastern')
                        dt_et = dt.astimezone(et_tz)

                        # Format as MM/DD/YYYY HH:MM AM ET
                        value = dt_et.strftime('%m/%d/%Y %I:%M %p ET')
                    except Exception as e:
                        logger.warning(f"Could not format date {value}: {e}")
                        # Keep original value if formatting fails
                        pass
                elif isinstance(value, list):
                    value = ', '.join(str(v) for v in value)

                # Truncate any overly long text
                if isinstance(value, str) and len(value) > max_cell_length:
                    truncation_msg = '\n\n[... Content truncated due to Excel cell size limit ...]'
                    value = value[:max_cell_length - len(truncation_msg)] + truncation_msg

                row[col_label] = value if value is not None else ''

            rows.append(row)

        # Create DataFrame and export to Excel
        df = pd.DataFrame(rows)

        # Create temporary file
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
            temp_path = tmp.name
            df.to_excel(temp_path, index=False, engine='openpyxl')

        # Add hyperlinks to ID column
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        wb = load_workbook(temp_path)
        ws = wb.active

        # Find the ID column index
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        id_col_idx = None
        for idx, header in enumerate(header_row, 1):
            if header and header.lower() == 'id':
                id_col_idx = idx
                break

        # Add hyperlinks to ID cells
        if id_col_idx:
            blue_font = Font(color="0046AD", underline="single")
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=id_col_idx)
                if cell.value:
                    ticket_id = cell.value
                    ticket_url = f"https://msoar.crtx.us.paloaltonetworks.com/Custom/caseinfoid/{ticket_id}"
                    cell.hyperlink = ticket_url
                    cell.font = blue_font
                    cell.value = ticket_id  # Keep the ID as the display value

        wb.save(temp_path)

        # Apply professional formatting
        column_widths = {
            'id': 15,
            'name': 30,
            'severity': 15,
            'status': 15,
            'country': 20,
            'impact': 25,
            'type': 25,
            'owner': 25,
            'created': 25,
            'user notes': 80,
            'notes': 80
        }

        wrap_columns = {'notes', 'impact', 'name', 'user notes'}
        # Note: dates are pre-formatted as strings, so no date_columns needed

        apply_professional_formatting(temp_path, column_widths, wrap_columns, date_columns=set())

        # Send file
        return send_file(
            temp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='security_incidents.xlsx'
        )

    except Exception as e:
        logger.error(f"Error exporting meaningful metrics: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/healthz")
def healthz():
    """Lightweight health probe endpoint for load balancers / monitoring.
    Augmented with server details to verify runtime (waitress vs flask-dev).
    """
    try:
        # Use Eastern Time for all timestamps
        current_time = datetime.now(eastern)
        timestamp = current_time.strftime('%Y-%m-%d %H:%M:%S %Z')

        # Calculate uptime
        uptime_delta = current_time - SERVER_START_TIME
        uptime_seconds = int(uptime_delta.total_seconds())
        uptime_hours = uptime_seconds // 3600
        uptime_minutes = (uptime_seconds % 3600) // 60
        uptime_str = f"{uptime_hours}h {uptime_minutes}m"

        try:
            # Try to get live request environ server software if available
            server_software = request.environ.get('SERVER_SOFTWARE', 'unknown')  # type: ignore[attr-defined]
            server_name = request.environ.get('SERVER_NAME', 'unknown')  # type: ignore[attr-defined]
            server_port = request.environ.get('SERVER_PORT', WEB_SERVER_PORT)  # type: ignore[attr-defined]
            server_info = {
                'server_type': detect_server_type(server_software),
                'server_software': server_software,
                'host': server_name,
                'port': int(server_port) if server_port else WEB_SERVER_PORT,
                'start_time': SERVER_START_TIME.strftime('%Y-%m-%d %H:%M:%S %Z'),
                'uptime': uptime_str,
                'uptime_seconds': uptime_seconds
            }
        except Exception:
            # Fallback to cached sample established at first request
            server_info = _runtime_server_info_sample()
            server_info['host'] = 'unknown'
            server_info['port'] = WEB_SERVER_PORT
            server_info['start_time'] = SERVER_START_TIME.strftime('%Y-%m-%d %H:%M:%S %Z')
            server_info['uptime'] = uptime_str
            server_info['uptime_seconds'] = uptime_seconds
        return jsonify({
            "status": "ok",
            "timestamp": timestamp,
            "team": getattr(CONFIG, 'team_name', 'unknown'),
            "service": "ir_web_server",
            "server": server_info
        }), 200
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route('/toodles')
@log_web_activity
def toodles_chat():
    """Toodles chat interface - password protected"""
    # The template handles authentication via modal, so just render it
    # Session check happens on the frontend via API calls
    return render_template('toodles_chat.html')


@app.route('/api/toodles/login', methods=['POST'])
def api_toodles_login():
    """API endpoint for Toodles authentication"""
    try:
        data = request.get_json()
        password = data.get('password', '').strip()
        email = data.get('email', '').strip()

        # Check password against config
        configured_password = CONFIG.toodles_password
        if not configured_password:
            logger.error("TOODLES_PASSWORD not configured in .env file")
            return jsonify({'success': False, 'error': 'Authentication system not configured'}), 500

        if password == configured_password:
            session['toodles_authenticated'] = True
            session['toodles_user_email'] = email
            session.permanent = True  # Make session persist
            return jsonify({'success': True, 'message': 'Authentication successful'})
        else:
            return jsonify({'success': False, 'error': 'Invalid password'}), 401

    except Exception as e:
        logger.error(f"Error in Toodles login: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/toodles/logout', methods=['POST'])
def api_toodles_logout():
    """API endpoint to logout from Toodles"""
    session.pop('toodles_authenticated', None)
    return jsonify({'success': True, 'message': 'Logged out successfully'})


@app.route('/api/toodles/create-x-ticket', methods=['POST'])
@log_web_activity
def api_create_x_ticket():
    """API endpoint to create X ticket"""
    try:
        data = request.get_json()
        title = data.get('title', '').strip()
        details = data.get('details', '').strip()
        detection_source = data.get('detection_source', '').strip()
        user_email = data.get('user_email', '').strip()

        if not title or not details or not detection_source:
            return jsonify({'success': False, 'error': 'All fields are required'}), 400

        # Add submitter info to details
        submitter_ip = request.remote_addr
        details += f"\n\nSubmitted by: {user_email}"
        details += f"\nSubmitted from: {submitter_ip}"

        incident = {
            'name': title,
            'details': details,
            'CustomFields': {
                'detectionsource': detection_source,
                'isusercontacted': False,
                'securitycategory': 'CAT-5: Scans/Probes/Attempted Access'
            }
        }

        result = prod_ticket_handler.create(incident)
        new_incident_id = result.get('id')
        incident_url = CONFIG.xsoar_prod_ui_base_url + '/Custom/caseinfoid/' + new_incident_id

        return jsonify({
            'success': True,
            'message': f'Ticket [#{new_incident_id}]({incident_url}) has been created in XSOAR Prod.'
        })

    except Exception as e:
        logger.error(f"Error creating X ticket: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/toodles/approved-testing', methods=['POST'])
@log_web_activity
def api_approved_testing():
    """API endpoint to add approved testing entry"""
    try:
        data = request.get_json()
        usernames = data.get('usernames', '').strip()
        tester_hosts = data.get('tester_hosts', '').strip()
        targets = data.get('targets', '').strip()
        description = data.get('description', '').strip()
        notes_scope = data.get('notes_scope', '').strip()
        keep_until = data.get('keep_until', '')
        user_email = data.get('user_email', '').strip()

        submitter_ip = request.remote_addr
        # Default to tomorrow if no date provided
        if not keep_until:
            keep_until = (datetime.now(eastern) + timedelta(days=1)).strftime("%Y-%m-%d")
        submit_date = datetime.now(eastern).strftime("%m/%d/%Y")

        # Use user email if provided, otherwise use IP-based identifier
        submitter_email = user_email if user_email else f"web_user@{submitter_ip}"

        approved_testing_list_name = f"{CONFIG.team_name}_Approved_Testing"
        approved_testing_master_list_name = f"{CONFIG.team_name}_Approved_Testing_MASTER"

        try:
            add_approved_testing_entry(
                prod_list_handler,
                approved_testing_list_name,
                approved_testing_master_list_name,
                usernames,
                tester_hosts,
                targets,
                description,
                notes_scope,
                submitter_email,
                keep_until,
                submit_date,
                submitter_ip
            )

            return jsonify({
                'success': True,
                'message': f'Your approved testing entry has been added. Expires at 5 PM ET on {keep_until}.'
            })

        except ValueError as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    except Exception as e:
        logger.error(f"Error adding approved testing: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/toodles/ioc-hunt', methods=['POST'])
@log_web_activity
def api_ioc_hunt():
    """API endpoint to create IOC hunt"""
    try:
        data = request.get_json()
        ioc_title = data.get('ioc_title', '').strip()
        iocs = data.get('iocs', '').strip()
        user_email = data.get('user_email', '').strip()

        if not ioc_title or not iocs:
            return jsonify({'success': False, 'error': 'All fields are required'}), 400

        submitter_ip = request.remote_addr
        details = iocs
        if user_email:
            details += f"\n\nSubmitted by: {user_email}"
        details += f"\nSubmitted from: {submitter_ip}"

        incident = {
            'name': ioc_title,
            'details': details,
            'type': "METCIRT IOC Hunt",
            'CustomFields': {
                'huntsource': 'Other'
            }
        }

        result = prod_ticket_handler.create(incident)
        ticket_no = result.get('id')
        incident_url = CONFIG.xsoar_prod_ui_base_url + '/Custom/caseinfoid/' + ticket_no

        return jsonify({
            'success': True,
            'message': f'A New IOC Hunt has been created in XSOAR. Ticket: [#{ticket_no}]({incident_url})'
        })

    except Exception as e:
        logger.error(f"Error creating IOC hunt: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/toodles/threat-hunt', methods=['POST'])
@log_web_activity
def api_threat_hunt():
    """API endpoint to create threat hunt"""
    try:
        data = request.get_json()
        threat_title = data.get('threat_title', '').strip()
        threat_description = data.get('threat_description', '').strip()
        user_email = data.get('user_email', '').strip()

        if not threat_title or not threat_description:
            return jsonify({'success': False, 'error': 'All fields are required'}), 400

        submitter_ip = request.remote_addr
        details = threat_description
        if user_email:
            details += f"\n\nSubmitted by: {user_email}"
        details += f"\nSubmitted from: {submitter_ip}"

        incident = {
            'name': threat_title,
            'details': details,
            'type': "Threat Hunt"
        }

        result = prod_ticket_handler.create(incident)
        ticket_no = result.get('id')
        incident_url = CONFIG.xsoar_prod_ui_base_url + '/Custom/caseinfoid/' + ticket_no

        return jsonify({
            'success': True,
            'message': f'A new Threat Hunt has been created in XSOAR. Ticket: [#{ticket_no}]({incident_url})'
        })

    except Exception as e:
        logger.error(f"Error creating threat hunt: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/toodles/oncall', methods=['GET'])
@log_web_activity
def api_oncall():
    """API endpoint to get on-call information"""
    try:
        import src.components.oncall as oncall
        on_call_person = oncall.get_on_call_person()

        return jsonify({
            'success': True,
            'data': on_call_person
        })

    except Exception as e:
        logger.error(f"Error getting on-call info: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/verify-command')
@log_web_activity
def verify_command_form():
    """Display command verification form"""
    # Get command details from query parameters
    command = request.args.get('command', 'N/A')
    timestamp = request.args.get('timestamp', 'N/A')
    system = request.args.get('system', 'N/A')
    ticket_id = request.args.get('ticket_id', '')
    task_id = request.args.get('task_id', '')

    return render_template('command_verification_form.html',
                           command=command,
                           timestamp=timestamp,
                           system=system,
                           ticket_id=ticket_id,
                           task_id=task_id)


@app.route('/submit-command-verification', methods=['POST'])
@log_web_activity
def submit_command_verification():
    """Handle command verification form submission"""
    try:
        data = request.form.to_dict()
        recognized = data.get('recognized')  # 'yes' or 'no'
        ticket_id = data.get('ticket_id', '')
        task_id = data.get('task_id', '')
        command = data.get('command', '')
        timestamp = data.get('timestamp', '')
        system = data.get('system', '')

        # Log the response
        logger.info(f"Command verification response: recognized={recognized}, ticket_id={ticket_id}, task_id={task_id}, command={command}")

        # Complete the XSOAR task if ticket_id and task_id are provided
        if ticket_id and task_id:
            try:
                result = prod_ticket_handler.complete_task(ticket_id, task_id, recognized)
                logger.info(f"Successfully completed XSOAR task {task_id} in ticket {ticket_id} with response: {recognized}")

                return jsonify({
                    'status': 'success',
                    'recognized': recognized,
                    'ticket_id': ticket_id,
                    'task_id': task_id,
                    'message': f'Thank you for your response. The task has been completed in XSOAR ticket #{ticket_id}.'
                })
            except Exception as xsoar_error:
                logger.error(f"Error completing XSOAR task: {xsoar_error}")
                return jsonify({
                    'status': 'error',
                    'error': f'Failed to complete XSOAR task: {str(xsoar_error)}'
                }), 500
        else:
            # If no ticket_id or task_id, just log the response
            logger.warning(f"No ticket_id or task_id provided. Response logged but not sent to XSOAR.")
            return jsonify({
                'status': 'success',
                'recognized': recognized,
                'message': 'Thank you for your response. Your verification has been recorded.'
            })

    except Exception as e:
        logger.error(f"Error submitting command verification: {e}")
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500


def main():
    """Entry point for launching the web server.

    Supports configuration via CLI args or environment variables:
      --host / WEB_HOST (default 0.0.0.0 for LAN access)
      --port / WEB_PORT (default 8080)
      --proxy flag enables internal proxy (default disabled)
    """

    # Initialize Pokédex bot components
    try:
        if os.environ.get('SKIP_POKEDEX_WARMUP', '').lower() == 'true':
            print("⏭️ Skipping Pokedex initialization (SKIP_POKEDEX_WARMUP=true)")
        else:
            print("🤖 Initializing Pokedex chat components...")
            from my_bot.core.my_model import initialize_model_and_agent
            if initialize_model_and_agent():
                print("✅ Pokedex chat components initialized!")

                # Warm up the model to preload it into memory
                print("🔥 Warming up LLM (this will load the model into memory)...")
                state_manager = get_state_manager()
                if state_manager and hasattr(state_manager, 'fast_warmup') and state_manager.fast_warmup():
                    print("✅ LLM warmed up and ready! Model is now loaded in memory.")
                else:
                    print("⚠️ LLM warmup skipped or failed - model will load on first request")
            else:
                print("⚠️ Pokedex chat initialization failed - chat endpoint will return errors")
    except Exception as e:
        print(f"⚠️ Failed to initialize Pokedex chat: {e}")
        print("   Chat endpoint will be available but may return errors")

    charts_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'static/charts'))
    app.config['CHARTS_DIR'] = charts_dir

    host = '0.0.0.0'
    port = WEB_SERVER_PORT

    # Only start proxy server in main process (not in reloader child process) and if enabled
    if SHOULD_START_PROXY and os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        proxy_thread = threading.Thread(target=start_proxy_server, daemon=True)
        proxy_thread.start()
        print(f"High-performance proxy server thread started on port {PROXY_PORT}")
    elif not SHOULD_START_PROXY:
        print(f"Proxy server disabled (pass --proxy to enable)")

    print(f"Attempting to start web server on http://{host}:{WEB_SERVER_PORT}")

    if USE_DEBUG_MODE:
        print("Using Flask dev server with auto-reload (debug mode)")
        try:
            app.run(debug=True, host=host, port=WEB_SERVER_PORT, threaded=True, use_reloader=True)
        except OSError as e:
            if WEB_SERVER_PORT < 1024 and e.errno == 13:
                fallback_port = 8080
                print(f"\n{'=' * 70}")
                print(f"❌ ERROR: Port {WEB_SERVER_PORT} is LOCKED/UNAVAILABLE")
                print(f"{'=' * 70}")
                print(f"Port {WEB_SERVER_PORT} requires elevated privileges (sudo/root).")
                print(f"This is typically because ports below 1024 are privileged ports.")
                print(f"\nTo fix this:")
                print(f"  1. Run with sudo: sudo python3 web/web_server.py")
                print(f"  2. Grant capability: sudo setcap 'cap_net_bind_service=+ep' $(which python3)")
                print(f"  3. Use a different port in .env: WEB_SERVER_PORT=8080")
                print(f"\nFalling back to port {fallback_port}...")
                print(f"{'=' * 70}\n")
                app.run(debug=True, host=host, port=fallback_port, threaded=True, use_reloader=True)
            elif e.errno == 48 or e.errno == 98:  # Address already in use
                print(f"\n{'=' * 70}")
                print(f"❌ ERROR: Port {WEB_SERVER_PORT} is LOCKED/IN USE")
                print(f"{'=' * 70}")
                print(f"Port {WEB_SERVER_PORT} is already being used by another process.")
                print(f"\nTo fix this:")
                print(f"  1. Find the process: sudo lsof -i :{WEB_SERVER_PORT}")
                print(f"  2. Stop the process using the port")
                print(f"  3. Or use a different port in .env: WEB_SERVER_PORT=8080")
                print(f"{'=' * 70}\n")
                raise
            else:
                raise
    else:
        try:
            from waitress import serve
            print("Using Waitress WSGI server for production deployment")
            try:
                serve(app, host=host, port=port, threads=20, channel_timeout=120)
            except OSError as e:
                # Permission denied for privileged port without sudo/capability
                if port < 1024 and e.errno == 13:  # Permission denied
                    fallback_port = 8080
                    print(f"\n{'=' * 70}")
                    print(f"❌ ERROR: Port {port} is LOCKED/UNAVAILABLE")
                    print(f"{'=' * 70}")
                    print(f"Port {port} requires elevated privileges (sudo/root).")
                    print(f"This is typically because ports below 1024 are privileged ports.")
                    print(f"\nTo fix this:")
                    print(f"  1. Run with sudo: sudo python3 web/web_server.py")
                    print(f"  2. Grant capability: sudo setcap 'cap_net_bind_service=+ep' $(which python3)")
                    print(f"  3. Use a different port in .env: WEB_SERVER_PORT=8080")
                    print(f"\nFalling back to port {fallback_port}...")
                    print(f"{'=' * 70}\n")
                    serve(app, host=host, port=fallback_port, threads=20, channel_timeout=120)
                elif e.errno == 48 or e.errno == 98:  # Address already in use
                    print(f"\n{'=' * 70}")
                    print(f"❌ ERROR: Port {port} is LOCKED/IN USE")
                    print(f"{'=' * 70}")
                    print(f"Port {port} is already being used by another process.")
                    print(f"\nTo fix this:")
                    print(f"  1. Find the process: sudo lsof -i :{port}")
                    print(f"  2. Stop the process using the port")
                    print(f"  3. Or use a different port in .env: WEB_SERVER_PORT=8080")
                    print(f"{'=' * 70}\n")
                    raise
                else:
                    raise
        except ImportError:
            print("Waitress not available, falling back to Flask dev server")
            try:
                app.run(debug=True, host=host, port=port, threaded=True, use_reloader=True)
            except OSError as e:
                if port < 1024 and e.errno == 13:
                    fallback_port = 8080
                    print(f"\n{'=' * 70}")
                    print(f"❌ ERROR: Port {port} is LOCKED/UNAVAILABLE")
                    print(f"{'=' * 70}")
                    print(f"Port {port} requires elevated privileges (sudo/root).")
                    print(f"This is typically because ports below 1024 are privileged ports.")
                    print(f"\nTo fix this:")
                    print(f"  1. Run with sudo: sudo python3 web/web_server.py")
                    print(f"  2. Grant capability: sudo setcap 'cap_net_bind_service=+ep' $(which python3)")
                    print(f"  3. Use a different port in .env: WEB_SERVER_PORT=8080")
                    print(f"\nFalling back to port {fallback_port}...")
                    print(f"{'=' * 70}\n")
                    app.run(debug=True, host=host, port=fallback_port, threaded=True, use_reloader=True)
                elif e.errno == 48 or e.errno == 98:  # Address already in use
                    print(f"\n{'=' * 70}")
                    print(f"❌ ERROR: Port {port} is LOCKED/IN USE")
                    print(f"{'=' * 70}")
                    print(f"Port {port} is already being used by another process.")
                    print(f"\nTo fix this:")
                    print(f"  1. Find the process: sudo lsof -i :{port}")
                    print(f"  2. Stop the process using the port")
                    print(f"  3. Or use a different port in .env: WEB_SERVER_PORT=8080")
                    print(f"{'=' * 70}\n")
                    raise
                else:
                    raise


if __name__ == "__main__":
    main()
