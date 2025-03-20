import json
import math
import time
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Optional, List, Union

import openpyxl
from falconpy import OAuth2, Hosts, RealTimeResponse
from pytz import timezone
from webexpythonsdk import WebexAPI

from config import get_config
from services.service_now import ServiceNowClient

# Load configuration
config = get_config()

# Constants
DATA_DIR = Path("data")
TRANSIENT_DIR = DATA_DIR / "transient"
REGIONS_FILE = DATA_DIR / "regions_by_country.json"
COUNTRIES_FILE = DATA_DIR / "countries_by_code.json"
INPUT_FILE = TRANSIENT_DIR / "EPP-Falcon ring tagging.xlsx"

# Ring distribution percentages
RING_1_PERCENT = 0.1
RING_2_PERCENT = 0.2
RING_3_PERCENT = 0.3
# Ring 4 is the remainder

# Server environment mappings
RING_1_ENVS = {"dev", "poc", "lab", "integration"}
RING_2_ENVS = {"qa", "test"}
RING_3_ENVS = {"dr"}
# Ring 4 is for production or unknown environments

# Ensure directories exist
TRANSIENT_DIR.mkdir(parents=True, exist_ok=True)

# Load mapping data
with open(REGIONS_FILE, 'r') as f:
    REGIONS_BY_COUNTRY = json.load(f)

with open(COUNTRIES_FILE, 'r') as f:
    COUNTRY_NAMES_BY_ABBREVIATION = json.load(f)

# Initialize API clients
falcon_auth = OAuth2(
    client_id=config.cs_rtr_client_id,
    client_secret=config.cs_rtr_client_secret,
    base_url="api.us-2.crowdstrike.com",
    ssl_verify=False
)
falcon_rtr = RealTimeResponse(auth_object=falcon_auth)
falcon_hosts = Hosts(auth_object=falcon_auth)
service_now = ServiceNowClient(
    config.snow_base_url,
    config.snow_functional_account_id,
    config.snow_functional_account_password,
    config.snow_client_key
)


class HostCategory(Enum):
    """Enumeration of possible host categories."""
    WORKSTATION = "Workstation"
    SERVER = "Server"


@dataclass
class Host:
    """
    Represents a host with all its relevant attributes and tagging information.
    """
    name: str
    device_id: str = ""
    country: str = ""
    region: str = ""
    category: Optional[HostCategory] = None
    environment: Union[str, List[str]] = ""
    current_crowd_strike_tags: List[str] = field(default_factory=list)
    new_crowd_strike_tag: str = ""
    was_country_guessed: bool = False
    life_cycle_status: str = ""
    status_message: str = ""

    @classmethod
    def create_and_initialize(cls, name: str) -> 'Host':
        """Factory method to create and initialize a host with its data."""
        host = cls(name)
        host._initialize_host_data()
        return host

    def _initialize_host_data(self) -> None:
        """Initialize all host data in one method."""
        self._set_cs_device_id()

        if not self.device_id:
            self.status_message = "No CrowdStrike device ID found"
            return

        self._set_host_details_from_snow()
        self._normalize_country_data()
        self._determine_region()

    def _set_cs_device_id(self) -> None:
        """Retrieve device ID and tags from CrowdStrike."""
        try:
            host_filter = f"hostname:'{self.name}'"
            response = falcon_hosts.query_devices_by_filter(filter=host_filter, sort_by="last_seen.desc", limit=1)

            if response.get("status_code") != 200:
                self.status_message = f"CrowdStrike API error: {response.get('errors', ['Unknown error'])}"
                return

            resources = response["body"].get("resources", [])

            if not resources:
                self.status_message = "Host not found in CrowdStrike"
                return

            self.device_id = resources[0]

            # Fetch device details including tags
            device_response = falcon_hosts.get_device_details(ids=self.device_id)

            if device_response.get("status_code") != 200:
                self.status_message = f"Error fetching device details: {device_response.get('errors', ['Unknown error'])}"
                return

            device_resources = device_response["body"].get("resources", [])

            if device_resources:
                self.current_crowd_strike_tags = device_resources[0].get('tags', [])

        except Exception as e:
            self.status_message = f"Error retrieving CrowdStrike data: {str(e)}"

    def _set_host_details_from_snow(self) -> None:
        """Retrieve host details from ServiceNow."""
        try:
            host_details = service_now.get_host_details(self.name)

            if not host_details:
                self.status_message = "Host not found in ServiceNow"
                return

            # Map category
            category = host_details.get('category', '').lower()
            if category == 'workstation':
                self.category = HostCategory.WORKSTATION
            elif category == 'server':
                self.category = HostCategory.SERVER
            else:
                self.status_message = f"Unknown host category: {category}"

            self.environment = host_details.get('environment', '')
            self.country = host_details.get('country', '')
            self.life_cycle_status = host_details.get('lifecycleStatus', '')

        except Exception as e:
            self.status_message = f"Error retrieving ServiceNow data: {str(e)}"

    def _normalize_country_data(self) -> None:
        """Normalize country data"""

        if not self.country:
            # Try to infer country from hostname prefix
            country_code = self.name[:2].upper()
            self.country = COUNTRY_NAMES_BY_ABBREVIATION.get(country_code, '')

            if self.country:
                self.was_country_guessed = True
                self.status_message = f"Country guessed from hostname: {self.country}"
            else:
                self.status_message = "Country unknown and couldn't be guessed"

    def _determine_region(self) -> None:
        """Determine the region based on country with special case handling."""
        self.region = REGIONS_BY_COUNTRY.get(self.country, '')

        # Special cases per requirements
        if self.country == 'US':
            self.region = 'US'
        elif self.country == 'Japan':
            self.region = 'Japan'

        if not self.region:
            self.status_message = "Region could not be determined"

    def add_tag_to_crowd_strike(self) -> bool:
        """Add the generated tag to CrowdStrike."""
        try:
            if not self.new_crowd_strike_tag or not self.device_id:
                return False

            response = falcon_hosts.update_device_tags(
                action_name='add',
                ids=self.device_id,
                tags=[self.new_crowd_strike_tag]
            )

            if response.get("status_code") == 200:
                return True
            else:
                self.status_message = f"Failed to add tag: {response.get('errors', ['Unknown error'])}"
                return False

        except Exception as e:
            self.status_message = f"Error adding tag: {str(e)}"
            return False

    @staticmethod
    def needs_tagging(host: 'Host') -> bool:
        """Check if host needs tagging (no existing ring tags)."""
        return not any(
            tag.startswith('Falcon') and 'ring' in tag.lower()
            for tag in host.current_crowd_strike_tags
        )

    @staticmethod
    def generate_tags(hosts: List['Host']) -> None:
        """Assign rings to hosts based on predefined distribution."""
        # Filter hosts to only include those without existing ring tags
        hosts_to_tag = [host for host in hosts if Host.needs_tagging(host)]

        # Skip hosts with missing critical data
        valid_hosts = [
            host for host in hosts_to_tag
            if host.device_id and host.category and host.region
        ]

        # Separate workstations and servers
        workstations = [host for host in valid_hosts if host.category == HostCategory.WORKSTATION]
        servers = [host for host in valid_hosts if host.category == HostCategory.SERVER]

        # Get unique regions and countries
        regions = {host.region for host in valid_hosts if host.region}
        countries = {host.country for host in valid_hosts if host.country}

        # Process workstations
        Host._process_workstations(workstations, regions, countries)

        # Process servers
        Host._process_servers(servers)

    @staticmethod
    def _process_workstations(workstations: List['Host'], regions: set, countries: set) -> None:
        """Process workstations and assign ring tags based on distribution."""
        for region in regions:
            for country in countries:
                # Filter workstations by region and country
                ws_group = [
                    ws for ws in workstations
                    if ws.region == region and ws.country == country
                ]

                total = len(ws_group)
                if total == 0:
                    continue

                # Calculate ring sizes
                ring_1_size = max(1, int(RING_1_PERCENT * total)) if total >= 10 else 0
                ring_2_size = max(1, int(RING_2_PERCENT * total)) if total >= 5 else 0
                ring_3_size = max(1, int(RING_3_PERCENT * total)) if total >= 3 else 0
                ring_4_size = total - ring_1_size - ring_2_size - ring_3_size

                # Ensure at least one host in each ring when possible
                ring_sizes = [ring_1_size, ring_2_size, ring_3_size, ring_4_size]

                # Distribute workstations into rings
                current_index = 0
                for ring, size in enumerate(ring_sizes, start=1):
                    if size <= 0:
                        continue

                    for i in range(size):
                        if current_index < len(ws_group):
                            ws_group[current_index].new_crowd_strike_tag = f"{region}WksRing{ring}"
                            current_index += 1

    @staticmethod
    def _process_servers(servers: List['Host']) -> None:
        """Process servers and assign ring tags based on environment."""
        for server in servers:
            # Normalize environment data
            env = Host._normalize_environment(server.environment)

            # Determine ring based on environment
            if env in RING_1_ENVS:
                ring = 1
            elif env in RING_2_ENVS:
                ring = 2
            elif env in RING_3_ENVS:
                ring = 3
            else:  # production or unknown
                ring = 4
                if not env:
                    server.status_message = "No environment data, assigned to Ring 4"

            server.new_crowd_strike_tag = f"{server.region}SRVRing{ring}"

    @staticmethod
    def _normalize_environment(environment: Union[str, List[str]]) -> str | None:
        """Normalize environment data to a single string value."""
        if isinstance(environment, list):
            if not environment:
                return ""
            elif len(environment) >= 1:
                return environment[0].lower().strip()
        else:
            return environment.lower().strip()


def parse_timestamp(date_str: Optional[str]) -> Optional[datetime]:
    """Parse an ISO 8601 timestamp string into a datetime object."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%SZ")
    except ValueError:
        return None


def apply_tags(hosts: List[Host]) -> List[Host]:
    """Apply tags to hosts and update their status."""
    successfully_tagged = []

    for host in hosts:
        if host.new_crowd_strike_tag:
            success = host.add_tag_to_crowd_strike()
            if success:
                host.status_message = f"Successfully tagged with {host.new_crowd_strike_tag}"
                successfully_tagged.append(host)
            else:
                host.status_message = f"Failed to tag with {host.new_crowd_strike_tag}"

    return successfully_tagged


def format_duration(seconds: float) -> str:
    """Format seconds into a human-readable duration string."""
    minutes, seconds = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)

    seconds = math.ceil(seconds)  # Round up seconds

    parts = []
    if hours > 0:
        parts.append(f"{int(hours)} hour{'s' if hours != 1 else ''}")
    if minutes > 0:
        parts.append(f"{int(minutes)} minute{'s' if minutes != 1 else ''}")
    if seconds > 0 or not parts:
        parts.append(f"{int(seconds)} second{'s' if seconds != 1 else ''}")

    return " ".join(parts)


def write_results_to_file(hosts: List[Host]) -> str:
    """
    Writes the results to a new Excel sheet with timestamps.

    Args:
        hosts: List of Host objects to write to the sheet.

    Returns:
        str: Path to the output file
    """
    # Get the current date and time in ET
    et_timezone = timezone('US/Eastern')
    current_time_et = datetime.now(et_timezone).strftime("%m_%d_%Y %I:%M %p %Z")
    output_file = TRANSIENT_DIR / f'EPP-Falcon ring tagging {current_time_et}.xlsx'

    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers
    headers = [
        "Name", "CS Device ID", "Category", "Environment", "Life Cycle Status",
        "Country", "Region", "Was Country Guessed", "Current CS Tags",
        "Generated CS Tag", "Status"
    ]
    sheet.append(headers)

    # Write the results to the sheet
    for host in hosts:
        sheet.append([
            host.name,
            host.device_id,
            host.category.value if host.category else '',
            host.environment if isinstance(host.environment, str) else ', '.join(host.environment),
            host.life_cycle_status,
            host.country,
            host.region,
            'Yes' if host.was_country_guessed else 'No',
            ', '.join(host.current_crowd_strike_tags),
            host.new_crowd_strike_tag,
            host.status_message or ('Ring tag generated' if host.new_crowd_strike_tag else 'No tag needed'),
        ])

    # Save the workbook
    try:
        output_file.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_file)
        print(f"Results written to new file: {output_file}")
    except Exception as e:
        print(f"An error occurred while saving the workbook: {e}")
        return ""

    return str(output_file)


def get_hostnames(input_file=INPUT_FILE) -> List[str]:
    """
    Retrieves hostnames from an Excel file.

    Args:
        input_file: The path to the Excel file containing hostnames.

    Returns:
        List of hostnames, or an empty list if the file is not found.
    """
    try:
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active
        hostnames = [row[0].value for row in sheet.iter_rows(min_row=2) if row[0].value]
        print(f'Found {len(hostnames)} hostnames in input file')
        return hostnames
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found.")
        return []
    except Exception as e:
        print(f"Error reading input file: {e}")
        return []


def send_report(output_filename: str, time_report) -> bool:
    """Send a report via Webex with the result file attached."""
    try:
        webex_api = WebexAPI(config.webex_bot_access_token_moneyball)
        response = webex_api.messages.create(
            roomId=config.webex_room_id_vinay_test_space,
            markdown=f"EPP-Falcon ring tagging results are attached.\n\n```{time_report}",
            files=[output_filename]
        )
        return bool(response)
    except Exception as e:
        print(f"Error sending report: {e}")
        return False


def main() -> None:
    """Main execution function."""
    start_time = time.time()

    # Fetch and initialize hosts
    fetch_start = time.time()
    hostnames = get_hostnames()

    if not hostnames:
        print("No hostnames found in input file. Exiting.")
        return

    hosts = [Host.create_and_initialize(name) for name in hostnames]
    fetch_end = time.time()
    fetch_duration = fetch_end - fetch_start

    # Generate tags
    generate_tag_start = time.time()
    Host.generate_tags(hosts)
    generate_tag_end = time.time()
    generate_tag_duration = generate_tag_end - generate_tag_start

    # Filter statistics
    total_hosts = len(hosts)
    hosts_with_device_id = sum(1 for host in hosts if host.device_id)
    hosts_with_tags = sum(1 for host in hosts if host.new_crowd_strike_tag)

    # Write results before applying tags (for documentation)
    output_filename = write_results_to_file(hosts)

    # Apply tags
    apply_tag_start = time.time()
    apply_tag_end = time.time()
    apply_tag_duration = apply_tag_end - apply_tag_start

    # Generate a timing report
    end_time = time.time()
    total_duration = end_time - start_time

    time_report = f"""
Summary:
- Total hosts processed: {total_hosts}
- Hosts found in CrowdStrike: {hosts_with_device_id}
- Hosts tagged: {hosts_with_tags}

Timing:
- Fetching and initializing hosts: {format_duration(fetch_duration)}
- Generating tags: {format_duration(generate_tag_duration)}
- Applying tags: {format_duration(apply_tag_duration)}
- Total execution time: {format_duration(total_duration)}
        """

    send_report_success = send_report(output_filename, time_report)
    if send_report_success:
        print(f"Report successfully sent to Webex")
    else:
        print("Failed to send report to Webex")

    print(time_report)


if __name__ in ('__main__', '__builtin__', 'builtins'):
    main()
